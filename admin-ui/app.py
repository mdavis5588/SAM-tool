import io
import json
import os
import smtplib
import ssl
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

import bcrypt
import psycopg2
import psycopg2.extras
import psycopg2.sql
import requests
from flask import (Flask, Response, render_template, request, redirect,
                   url_for, session, flash, jsonify, abort)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import threading
import time

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "change-me-in-production")

# ---------------------------------------------------------------------------
# OCI public pricing cache
# Oracle publishes a full SKU list at this endpoint (no auth required).
# We cache it in-process for OCI_CACHE_TTL seconds to avoid hitting it on
# every page load.
# ---------------------------------------------------------------------------
_OCI_CACHE: dict = {"data": None, "fetched_at": 0.0, "lock": threading.Lock()}
OCI_CACHE_TTL   = 3600  # 1 hour
OCI_PRICING_URL = (
    "https://apexapps.oracle.com/pls/apex/cetools/api/v1/products/"
    "?currencyCode=USD"
)

# Keywords used to identify Oracle Database SKUs in the OCI catalogue.
_OCI_DB_KEYWORDS = [
    "oracle database enterprise edition",
    "oracle database standard edition",
    "oracle base database",
    "database cloud service",
    "exadata database",
    "exadata cloud@customer",
    "exacc",
]

# Metric labels returned by the OCI API that map to per-OCPU billing.
_OCI_OCPU_METRICS = {"ocpu per hour", "ocpu-hour", "ocpu hour"}


# Oracle published OCI Database hourly OCPU list prices (USD, as of 2024).
# Used as fallback when the live API is unreachable.
# Source: https://www.oracle.com/cloud/price-list/
_OCI_STATIC_SKUS = [
    {
        "name":                 "Oracle Database Enterprise Edition — BYOL",
        "part_number":          "B90453",
        "metric":               "ocpu per hour",
        "unit_price":           0.4480,
        "is_byol":              True,
        "is_licence_included":  False,
        "is_exacc":             False,
    },
    {
        "name":                 "Oracle Database Enterprise Edition — Licence Included",
        "part_number":          "B90454",
        "metric":               "ocpu per hour",
        "unit_price":           2.9008,
        "is_byol":              False,
        "is_licence_included":  True,
        "is_exacc":             False,
    },
    {
        "name":                 "Oracle Database Standard Edition 2 — BYOL",
        "part_number":          "B90455",
        "metric":               "ocpu per hour",
        "unit_price":           0.1344,
        "is_byol":              True,
        "is_licence_included":  False,
        "is_exacc":             False,
    },
    {
        "name":                 "Oracle Database Standard Edition 2 — Licence Included",
        "part_number":          "B90456",
        "metric":               "ocpu per hour",
        "unit_price":           0.2688,
        "is_byol":              False,
        "is_licence_included":  True,
        "is_exacc":             False,
    },
    # ExaCC (Exadata Cloud@Customer) X9M — Oracle-managed Exadata on-premises
    # Source: Oracle price list 2024. Note: infrastructure subscription fees apply
    # separately on top of these per-OCPU software rates.
    {
        "name":                 "Exadata Cloud@Customer X9M — BYOL",
        "part_number":          "B93189",
        "metric":               "ocpu per hour",
        "unit_price":           0.042279956,   # per OCPU per hour, BYOL (Oracle list price Aug 2024)
        "is_byol":              True,
        "is_licence_included":  False,
        "is_exacc":             True,
    },
    {
        "name":                 "Exadata Cloud@Customer X9M — Licence Included",
        "part_number":          "B93190",
        "metric":               "ocpu per hour",
        "unit_price":           1.76157746,   # per OCPU per hour, EE licence bundled (Oracle list price Aug 2024)
        "is_byol":              False,
        "is_licence_included":  True,
        "is_exacc":             True,
    },
]


_NO_PROXY = {"http": None, "https": None}  # bypass any HTTPS_PROXY env var


def _fetch_oci_raw() -> list:
    """Download all items from the OCI public pricing API."""
    resp = requests.get(OCI_PRICING_URL, timeout=15, proxies=_NO_PROXY)
    resp.raise_for_status()
    return resp.json().get("items", [])


def get_oci_prices() -> list:
    """
    Return cached list of OCI Database SKUs, refreshing if stale.
    Tries the live Oracle public API first; falls back to the static price
    table if the API is unreachable (proxy restrictions, network errors).
    Each item: {name, part_number, metric, unit_price,
                is_byol, is_licence_included}.
    """
    cache = _OCI_CACHE
    now = time.time()
    with cache["lock"]:
        if cache["data"] is not None and (now - cache["fetched_at"]) < OCI_CACHE_TTL:
            return cache["data"]

    try:
        raw = _fetch_oci_raw()
        db_skus = []
        for item in raw:
            name_lower = (item.get("displayName") or "").lower()
            if not any(kw in name_lower for kw in _OCI_DB_KEYWORDS):
                continue
            prices = item.get("currencyCodeLocalizations", [])
            usd_price = None
            for p in prices:
                if (p.get("currencyCode") or "").upper() == "USD":
                    usd_price = p.get("localizedPrice")
                    break
            if usd_price is None:
                continue
            metric = (item.get("metricName") or "").lower()
            db_skus.append({
                "name":                item.get("displayName", ""),
                "part_number":         item.get("partNumber", ""),
                "metric":              metric,
                "unit_price":          float(usd_price),
                "is_byol":             "byol" in name_lower or "bring your own" in name_lower,
                "is_licence_included": "license included" in name_lower
                                       or "licence included" in name_lower,
                "is_exacc":            "cloud@customer" in name_lower or "exacc" in name_lower,
            })
        if db_skus:
            with cache["lock"]:
                cache["data"] = db_skus
                cache["fetched_at"] = now
            return db_skus
    except Exception:
        pass

    # Live API unavailable — use static published prices
    return _OCI_STATIC_SKUS


# ---------------------------------------------------------------------------
# Azure pricing — static fallback (Azure Retail Prices API often unreachable)
# Source: Azure pricing calculator / retail prices, East US, Linux, PAYG, 2024
# ---------------------------------------------------------------------------
_AZURE_PRICING_URL = (
    "https://prices.azure.com/api/retail/prices"
    "?api-version=2023-01-01-preview"
    "&$filter=serviceName+eq+'Virtual+Machines'"
    "+and+armRegionName+eq+'eastus'"
    "+and+priceType+eq+'Consumption'"
    "+and+contains(skuName,'Esv5')"
    "+and+not+contains(skuName,'Windows')"
    "+and+not+contains(skuName,'Spot')"
)

# Oracle on Azure: E-series v5 (Intel Xeon, memory-optimised, common for Oracle DB)
# Per-vCPU price is consistent across E-series v5 sizes at ~$0.063/hr (Linux, PAYG, East US)
# Oracle on Azure: core factor = 0.5 → processor licences = vCPUs × 0.5
# So vCPUs needed = physical_cores × 2  (same net licence count as on-prem x86)
_AZURE_STATIC_SKUS = [
    {
        "name":        "Azure VM — Standard Esv5 series (Linux, PAYG, East US)",
        "sku_name":    "Standard_E-series v5",
        "metric":      "vCPU per hour",
        "unit_price":  0.0630,   # per vCPU per hour (E2s–E64s v5, Linux, East US)
        "is_byol":     True,
        "is_li":       False,
        "note":        "Compute cost only. Bring your own Oracle licence. "
                       "Oracle core factor 0.5 on Azure Intel VMs → 2 vCPUs per processor licence.",
    },
    {
        "name":        "Oracle Database@Azure — Exadata X9M (Licence Included)",
        "sku_name":    "Oracle Exadata X9M",
        "metric":      "OCPU per hour",
        "unit_price":  3.0272,   # per OCPU per hour (Oracle@Azure Exadata X9M base shape, 2024)
        "is_byol":     False,
        "is_li":       True,
        "note":        "Oracle Exadata running natively on Azure infrastructure. "
                       "Licence included — no separate Oracle software contract needed.",
    },
]


def _fetch_azure_raw() -> list:
    resp = requests.get(_AZURE_PRICING_URL, timeout=15, proxies=_NO_PROXY)
    resp.raise_for_status()
    return resp.json().get("Items", [])


def get_azure_prices() -> list:
    """
    Try to fetch Azure VM prices from the public Retail Prices API.
    Falls back to static published prices on any error.
    Returns list of SKU dicts with same shape as _AZURE_STATIC_SKUS.
    """
    try:
        items = _fetch_azure_raw()
        if not items:
            return _AZURE_STATIC_SKUS

        # Compute average per-vCPU price across returned E-series sizes
        total, count = 0.0, 0
        for it in items:
            vcpus = it.get("armSkuName", "").count("_") and None  # not reliable
            retail = it.get("retailPrice")
            vcpu_count_str = ""
            # Parse vCPU count from skuName e.g. "E4s v5"
            import re as _re
            m = _re.search(r"E(\d+)s", it.get("skuName", ""))
            if m and retail:
                total += float(retail) / int(m.group(1))
                count += 1
        if count == 0:
            return _AZURE_STATIC_SKUS
        per_vcpu = round(total / count, 4)
        dynamic = [dict(_AZURE_STATIC_SKUS[0])]
        dynamic[0]["unit_price"] = per_vcpu
        dynamic[0]["name"] += " (live price)"
        return dynamic + [_AZURE_STATIC_SKUS[1]]
    except Exception:
        return _AZURE_STATIC_SKUS


def build_azure_comparison(azure_skus: list, physical_cores: float,
                           horizon: int) -> dict | None:
    """
    Build year-by-year Azure cost comparison.

    Azure BYOL (VM):
      - Oracle core factor on Azure Intel VMs = 0.5
      - vCPUs needed = physical_cores × 2  (2 vCPUs per licence at 0.5 factor)
      - Annual = vcpus × hourly_rate × 8760

    Oracle@Azure (Exadata, Licence Included):
      - Billed per OCPU; 1 physical core ≈ 1 OCPU (same mapping as OCI)
      - Annual = physical_cores × hourly_rate × 8760
    """
    if not azure_skus or physical_cores <= 0:
        return None

    byol_sku = li_sku = None
    for sku in azure_skus:
        if sku["is_byol"] and byol_sku is None:
            byol_sku = sku
        if sku["is_li"] and li_sku is None:
            li_sku = sku

    if not byol_sku and not li_sku:
        return None

    vcpus = physical_cores * 2   # 2 vCPUs per physical core on Azure Intel VMs

    byol_annual = round(byol_sku["unit_price"] * vcpus * 8760, 2) if byol_sku else None
    li_annual   = round(li_sku["unit_price"]   * physical_cores * 8760, 2) if li_sku else None

    yearly = []
    for y in range(1, horizon + 1):
        yearly.append({
            "year":     y,
            "byol_cum": round(byol_annual * y, 2) if byol_annual is not None else None,
            "li_cum":   round(li_annual   * y, 2) if li_annual   is not None else None,
        })

    prices_static = all(s in _AZURE_STATIC_SKUS for s in azure_skus)

    return {
        "vcpus":         vcpus,
        "ocpus":         physical_cores,
        "byol_sku":      byol_sku,
        "li_sku":        li_sku,
        "byol_annual":   byol_annual,
        "li_annual":     li_annual,
        "yearly":        yearly,
        "prices_static": prices_static,
    }


def build_oci_comparison(oci_skus: list, total_processor_cores: float,
                         horizon: int) -> dict | None:
    """
    Build year-by-year OCI cost comparison including ExaCC.
    Picks the first matching EE SKU for each of:
      - OCI BYOL (standard cloud), OCI Licence Included, ExaCC BYOL, ExaCC LI
    Annual cost = ocpu_count × hourly_rate × 8760.
    """
    if not oci_skus or total_processor_cores <= 0:
        return None

    byol_sku = li_sku = exacc_byol_sku = exacc_li_sku = None
    for sku in oci_skus:
        if sku["metric"] not in _OCI_OCPU_METRICS:
            continue
        name_l = sku["name"].lower()
        is_exacc = sku.get("is_exacc", False)
        # ExaCC SKUs don't carry "enterprise" in their name; standard OCI ones do
        if not is_exacc and "enterprise" not in name_l:
            continue
        if is_exacc:
            if sku["is_byol"] and exacc_byol_sku is None:
                exacc_byol_sku = sku
            if sku["is_licence_included"] and exacc_li_sku is None:
                exacc_li_sku = sku
        else:
            if sku["is_byol"] and byol_sku is None:
                byol_sku = sku
            if sku["is_licence_included"] and li_sku is None:
                li_sku = sku

    if not any([byol_sku, li_sku, exacc_byol_sku, exacc_li_sku]):
        return None

    ocpus = total_processor_cores  # 1 physical core = 1 OCPU

    def annual(sku):
        return round(sku["unit_price"] * ocpus * 8760, 2) if sku else None

    byol_annual      = annual(byol_sku)
    li_annual        = annual(li_sku)
    exacc_byol_annual = annual(exacc_byol_sku)
    exacc_li_annual   = annual(exacc_li_sku)

    yearly = []
    for y in range(1, horizon + 1):
        yearly.append({
            "year":            y,
            "byol_cum":        round(byol_annual * y, 2)       if byol_annual is not None else None,
            "li_cum":          round(li_annual * y, 2)         if li_annual is not None else None,
            "exacc_byol_cum":  round(exacc_byol_annual * y, 2) if exacc_byol_annual is not None else None,
            "exacc_li_cum":    round(exacc_li_annual * y, 2)   if exacc_li_annual is not None else None,
        })

    return {
        "ocpus":             ocpus,
        "byol_sku":          byol_sku,
        "li_sku":            li_sku,
        "exacc_byol_sku":    exacc_byol_sku,
        "exacc_li_sku":      exacc_li_sku,
        "byol_annual":       byol_annual,
        "li_annual":         li_annual,
        "exacc_byol_annual": exacc_byol_annual,
        "exacc_li_annual":   exacc_li_annual,
        "yearly":            yearly,
        "note": (
            "OCI costs are indicative (list price, compute only). "
            "Actual costs depend on shape, storage, networking, and negotiated discounts. "
            "BYOL assumes existing perpetual EE licences. "
            "ExaCC has additional infrastructure subscription fees not included here."
        ),
    }

# ---------------------------------------------------------------------------
# i18n — simple JSON-based translations
# ---------------------------------------------------------------------------
_TRANSLATIONS = {}

def _load_translations():
    trans_dir = os.path.join(os.path.dirname(__file__), "translations")
    for lang in ("en", "fr"):
        path = os.path.join(trans_dir, f"{lang}.json")
        with open(path, encoding="utf-8") as f:
            _TRANSLATIONS[lang] = json.load(f)

_load_translations()

@app.route("/set-lang/<lang>")
def set_lang(lang):
    if lang in ("en", "fr"):
        session["lang"] = lang
    return redirect(request.referrer or url_for("servers"))

@app.context_processor
def inject_i18n():
    lang = session.get("lang", "en")
    strings = _TRANSLATIONS.get(lang, _TRANSLATIONS["en"])
    def t(key, **kwargs):
        val = strings.get(key, _TRANSLATIONS["en"].get(key, key))
        return val.format(**kwargs) if kwargs else val
    return {"t": t, "current_lang": lang}

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------
DB_CONFIG = {
    "host":     os.environ.get("DB_HOST", "localhost"),
    "port":     int(os.environ.get("DB_PORT", 5432)),
    "dbname":   os.environ.get("DB_NAME", "samdb"),
    "user":     os.environ.get("DB_USER", "sam_admin"),
    "password": os.environ.get("DB_PASSWORD", ""),
}
DEFAULT_CLIENT_SCHEMA = os.environ.get("SAM_CLIENT_SCHEMA", "client_acme")
DISPATCH_KEY = os.environ.get("DISPATCH_KEY", "")

ADMIN_USER     = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "changeme")


def get_db():
    return psycopg2.connect(**DB_CONFIG)


def query(sql, params=None, fetchall=True):
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params or ())
            return cur.fetchall() if fetchall else cur.fetchone()


def execute(sql, params=None):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
        conn.commit()


def get_schema():
    """Active client schema for this session."""
    return session.get("client_schema", DEFAULT_CLIENT_SCHEMA)


def get_clients():
    """All active clients — used by the navbar switcher."""
    try:
        return query(
            "SELECT client_id, client_code, client_name, schema_name "
            "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
        )
    except Exception:
        return []


# ---------------------------------------------------------------------------
# CSI <-> licence line compatibility
#
# A server's licence position line (product_family + product_detail) needs a
# CSI whose entitlement lines actually cover that same product/edition — a
# Standard Edition 2 CSI must never be usable to cover an Enterprise Edition
# requirement, and vice versa. Entitlement line product names are free text
# (e.g. "Oracle Database Enterprise Edition"), so edition rows are matched by
# edition class and everything else (options like "Partitioning") is matched
# by substring containment.
# ---------------------------------------------------------------------------
def _edition_class(text):
    if not text:
        return None
    t = text.lower()
    if "enterprise" in t:
        return "enterprise"
    if "standard edition 2" in t or "se2" in t:
        return "standard2"
    if "standard" in t:
        return "standard"
    return None


# Oracle Database options available for manual licence analysis.
# Each entry: (param_name, display_label, product_label_for_price_lookup, metric)
# product_label must match (or be a substring of) a row in oracle_product_list_prices.
MANUAL_DB_OPTIONS = [
    ("opt_diagnostics",     "Diagnostics Pack",              "Oracle Database Diagnostics Pack",              "processor"),
    ("opt_tuning",          "Tuning Pack",                   "Oracle Database Tuning Pack",                   "processor"),
    ("opt_partitioning",    "Partitioning",                  "Oracle Partitioning",                           "processor"),
    ("opt_rac",             "Real Application Clusters (RAC)", "Oracle Real Application Clusters",            "processor"),
    ("opt_multitenant",     "Multitenant",                   "Oracle Multitenant",                            "processor"),
    ("opt_adg",             "Active Data Guard",             "Oracle Active Data Guard",                      "processor"),
    ("opt_advanced_sec",    "Advanced Security",             "Oracle Advanced Security",                      "processor"),
    ("opt_label_sec",       "Label Security",                "Oracle Label Security",                         "processor"),
    ("opt_db_vault",        "Database Vault",                "Oracle Database Vault",                         "processor"),
    ("opt_olap",            "OLAP",                          "Oracle OLAP",                                   "processor"),
    ("opt_spatial",         "Spatial and Graph",             "Oracle Spatial and Graph",                      "processor"),
    ("opt_goldengate",      "GoldenGate",                    "Oracle GoldenGate",                             "processor"),
]

_DB_ULA_KEYWORDS = frozenset([
    "database", "oracle db",
    "enterprise edition", "standard edition", "personal edition",
    "tuning pack", "diagnostics pack", "diagnostic pack",
    "partitioning", "real application clusters", "multitenant",
    "active data guard", "data guard", "goldengate",
    "advanced security", "label security", "database vault",
    "olap", "spatial", "spatial and graph",
])

def _ula_product_matches_family(ula_product_name, server_family):
    """Return True if a ULA covered-product name applies to the server's product family."""
    pn = (ula_product_name or "").lower()
    if server_family == "oracle_database":
        return any(kw in pn for kw in _DB_ULA_KEYWORDS)
    if server_family == "oracle_weblogic":
        return "weblogic" in pn
    return server_family.replace("_", " ") in pn


def _is_compatible_product(family, product_detail, line_family, line_product_name):
    if family != line_family:
        return False
    detail_class = _edition_class(product_detail)
    line_class = _edition_class(line_product_name)
    if detail_class or line_class:
        return detail_class is not None and detail_class == line_class
    if not product_detail:
        return True
    return product_detail.lower() in (line_product_name or "").lower()


def _check_ula_coverage(csi_id, server_id, schema):
    """Check whether every product the server requires is covered by the ULA.

    Returns:
      None   — ULA has no covered_products rows configured; caller should warn.
      []     — All products are covered (or server has no specific product_detail).
      [str]  — List of product_detail strings that are NOT covered by the ULA.
    """
    covered_rows = query(
        "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s",
        (csi_id,)
    )
    if not covered_rows:
        return None  # No coverage defined — can't validate

    covered_lower = [r["product_name"].lower() for r in covered_rows]

    # Pull every distinct product_detail the server requires
    lp_details = query(
        f"SELECT DISTINCT product_detail "
        f"FROM {schema}.license_position "
        f"WHERE server_id = %s AND licences_required > 0 AND product_detail IS NOT NULL",
        (server_id,)
    )

    uncovered = []
    for row in lp_details:
        detail = (row["product_detail"] or "").strip()
        if not detail:
            continue
        detail_lower = detail.lower()
        # A product is covered if any ULA product name contains it or vice-versa
        if not any(cl in detail_lower or detail_lower in cl for cl in covered_lower):
            uncovered.append(detail)

    return uncovered


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Auth & RBAC
# ---------------------------------------------------------------------------

def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()

def _check_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode(), hashed.encode())
    except Exception:
        return False

def _ensure_bootstrap_admin():
    """On startup: upsert the env-var admin account into app_users with a bcrypt hash."""
    try:
        existing = query(
            "SELECT user_id, password_hash FROM sam_admin.app_users WHERE username = %s",
            (ADMIN_USER,), fetchall=False
        )
        hashed = _hash_password(ADMIN_PASSWORD)
        if existing is None:
            execute(
                """INSERT INTO sam_admin.app_users
                   (username, display_name, role, auth_method, password_hash, created_by)
                   VALUES (%s, 'Administrator', 'superadmin', 'local', %s, 'bootstrap')""",
                (ADMIN_USER, hashed)
            )
        elif not existing["password_hash"]:
            execute(
                "UPDATE sam_admin.app_users SET password_hash = %s WHERE username = %s",
                (hashed, ADMIN_USER)
            )
    except Exception:
        pass  # Table may not exist yet (pre-migration); fall back to env-var auth below

def _get_user_from_db(username: str):
    try:
        return query(
            """SELECT u.user_id, u.username, u.display_name, u.email,
                      u.password_hash, u.role, u.client_id, u.is_active,
                      u.auth_method, u.ad_username, u.force_password_change,
                      c.schema_name AS client_schema, c.client_code, c.client_name
               FROM sam_admin.app_users u
               LEFT JOIN sam_admin.clients c ON c.client_id = u.client_id
               WHERE u.username = %s AND u.is_active""",
            (username,), fetchall=False
        )
    except Exception:
        return None

def current_user():
    """Return the full user dict for the logged-in user (from session cache)."""
    return session.get("user") or {}

def current_role():
    return current_user().get("role", "")

def is_superadmin():
    return current_role() == "superadmin"

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated

def roles_required(*roles):
    """Decorator: allow only users whose role is in *roles."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get("logged_in"):
                return redirect(url_for("login", next=request.path))
            if current_role() not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated
    return decorator

def superadmin_required(f):
    return roles_required("superadmin")(f)

def can_write_contracts():
    return current_role() in ("superadmin", "contracting")

def can_write_licences():
    return current_role() in ("superadmin", "dba")

def _enforce_client_scope():
    """For client-scoped users, force the active schema to their assigned client."""
    user = current_user()
    if user.get("role") == "client" and user.get("client_schema"):
        session["client_schema"] = user["client_schema"]

@app.before_request
def _apply_client_scope():
    if session.get("logged_in"):
        _enforce_client_scope()

@app.context_processor
def _inject_user():
    u    = current_user()
    role = u.get("role", "") if u else ""
    pending_count = 0
    if u:
        try:
            if role in ("superadmin", "contracting"):
                row = query(
                    "SELECT COUNT(*) AS n FROM sam_admin.assignment_requests WHERE status='pending'",
                    fetchall=False
                )
            else:
                row = query(
                    "SELECT COUNT(*) AS n FROM sam_admin.assignment_requests "
                    "WHERE status='pending' AND proposed_by=%s",
                    (u.get("username"),), fetchall=False
                )
            pending_count = (row or {}).get("n", 0)
        except Exception:
            pass
    active_alerts_count = 0
    if u:
        try:
            row = query(
                "SELECT COUNT(*) AS n FROM shared.compliance_alerts WHERE severity = 'HIGH'",
                fetchall=False
            )
            active_alerts_count = (row or {}).get("n", 0)
        except Exception:
            pass

    return {
        "current_user":              u,
        "current_role":              role,
        "can_write_contracts":       can_write_contracts(),
        "can_write_licences":        can_write_licences(),
        "is_superadmin":             is_superadmin(),
        "pending_assignments_count": pending_count,
        "active_alerts_count":       active_alerts_count,
    }

with app.app_context():
    try:
        _ensure_bootstrap_admin()
    except Exception:
        pass


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""

        user = _get_user_from_db(username)
        if user:
            # DB-backed login
            authed = False
            if user["auth_method"] == "local" and user["password_hash"]:
                authed = _check_password(password, user["password_hash"])
            # AD placeholder: when auth_method='active_directory', integrate LDAP here
            if authed:
                execute(
                    "UPDATE sam_admin.app_users SET last_login = NOW() WHERE user_id = %s",
                    (user["user_id"],)
                )
                session["logged_in"] = True
                session["user"] = dict(user)
                # Client-scoped users are locked to their client
                if user["role"] == "client" and user["client_schema"]:
                    session["client_schema"] = user["client_schema"]
                return redirect(request.args.get("next") or url_for("servers"))
        else:
            # Legacy fallback: env-var admin (covers pre-migration deployments)
            if username == ADMIN_USER and password == ADMIN_PASSWORD:
                session["logged_in"] = True
                session["user"] = {
                    "username": ADMIN_USER, "display_name": "Administrator",
                    "role": "superadmin", "client_id": None,
                    "client_schema": None, "client_code": None,
                }
                return redirect(request.args.get("next") or url_for("servers"))

        flash("Invalid username or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# User management  (superadmin only)
# ---------------------------------------------------------------------------
@app.route("/admin/users")
@superadmin_required
def admin_users():
    users = query("""
        SELECT u.user_id, u.username, u.display_name, u.email,
               u.role, u.is_active, u.auth_method, u.ad_username,
               u.last_login::DATE AS last_login, u.created_at::DATE AS created_at,
               c.client_name
        FROM sam_admin.app_users u
        LEFT JOIN sam_admin.clients c ON c.client_id = u.client_id
        ORDER BY u.role, u.username
    """)
    clients = query(
        "SELECT client_id, client_name, client_code FROM sam_admin.clients "
        "WHERE is_active ORDER BY client_name"
    )
    return render_template("admin_users.html", users=users, clients=clients)


@app.route("/admin/users/new", methods=["POST"])
@superadmin_required
def admin_user_create():
    username     = (request.form.get("username") or "").strip()
    display_name = (request.form.get("display_name") or "").strip() or None
    email        = (request.form.get("email") or "").strip() or None
    role         = request.form.get("role", "client")
    client_id    = request.form.get("client_id") or None
    auth_method  = request.form.get("auth_method", "local")
    ad_username  = (request.form.get("ad_username") or "").strip() or None
    password     = request.form.get("password") or ""

    if not username:
        flash("Username is required.", "danger")
        return redirect(url_for("admin_users"))

    if role not in ("superadmin", "contracting", "dba", "client"):
        flash("Invalid role.", "danger")
        return redirect(url_for("admin_users"))

    if role == "client" and not client_id:
        flash("A client must be selected for client-scoped users.", "danger")
        return redirect(url_for("admin_users"))

    if role != "client":
        client_id = None

    pw_hash = _hash_password(password) if auth_method == "local" and password else None

    try:
        execute(
            """INSERT INTO sam_admin.app_users
               (username, display_name, email, role, client_id, auth_method,
                ad_username, password_hash, created_by)
               VALUES (%s,%s,%s,%s::sam_admin.app_role,%s,%s::sam_admin.auth_method,%s,%s,%s)""",
            (username, display_name, email, role, client_id, auth_method,
             ad_username, pw_hash, current_user().get("username", "admin"))
        )
        _audit("user.create", entity_type="user", entity_name=username,
               new_values={"role": role, "auth_method": auth_method},
               client_schema="sam_admin")
        flash(f"User '{username}' created.", "success")
    except Exception as e:
        if "unique" in str(e).lower():
            flash(f"Username '{username}' already exists.", "danger")
        else:
            flash(f"Error creating user: {e}", "danger")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/edit", methods=["POST"])
@superadmin_required
def admin_user_edit(user_id):
    action = request.form.get("action")

    if action == "toggle_active":
        target_u = query(
            "SELECT username, is_active FROM sam_admin.app_users WHERE user_id = %s",
            (user_id,), fetchall=False
        )
        execute(
            "UPDATE sam_admin.app_users SET is_active = NOT is_active WHERE user_id = %s",
            (user_id,)
        )
        _audit("user.toggle_active", entity_type="user", entity_id=user_id,
               entity_name=(target_u or {}).get("username"),
               old_values={"is_active": (target_u or {}).get("is_active")},
               client_schema="sam_admin")
        flash("User status updated.", "success")

    elif action == "reset_password":
        new_pw = request.form.get("new_password") or ""
        if len(new_pw) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for("admin_users"))
        target_u = query(
            "SELECT username FROM sam_admin.app_users WHERE user_id = %s",
            (user_id,), fetchall=False
        )
        execute(
            "UPDATE sam_admin.app_users SET password_hash = %s, force_password_change = TRUE "
            "WHERE user_id = %s",
            (_hash_password(new_pw), user_id)
        )
        _audit("user.reset_password", entity_type="user", entity_id=user_id,
               entity_name=(target_u or {}).get("username"), client_schema="sam_admin")
        flash("Password reset. User will be prompted to change it on next login.", "success")

    elif action == "update":
        display_name = (request.form.get("display_name") or "").strip() or None
        email        = (request.form.get("email") or "").strip() or None
        role         = request.form.get("role", "client")
        client_id    = request.form.get("client_id") or None
        auth_method  = request.form.get("auth_method", "local")
        ad_username  = (request.form.get("ad_username") or "").strip() or None

        if role != "client":
            client_id = None

        # Prevent locking out the bootstrap admin
        target = query(
            "SELECT username FROM sam_admin.app_users WHERE user_id = %s",
            (user_id,), fetchall=False
        )
        if target and target["username"] == ADMIN_USER and role != "superadmin":
            flash("Cannot change the role of the bootstrap admin account.", "danger")
            return redirect(url_for("admin_users"))

        old_u = query(
            "SELECT username, role::TEXT AS role, auth_method::TEXT FROM sam_admin.app_users "
            "WHERE user_id = %s", (user_id,), fetchall=False
        )
        execute(
            """UPDATE sam_admin.app_users
               SET display_name=%s, email=%s, role=%s::sam_admin.app_role,
                   client_id=%s, auth_method=%s::sam_admin.auth_method, ad_username=%s
               WHERE user_id=%s""",
            (display_name, email, role, client_id, auth_method, ad_username, user_id)
        )
        _audit("user.update", entity_type="user", entity_id=user_id,
               entity_name=(old_u or {}).get("username"),
               old_values={"role": (old_u or {}).get("role")},
               new_values={"role": role, "auth_method": auth_method},
               client_schema="sam_admin")
        flash("User updated.", "success")

    elif action == "delete":
        target = query(
            "SELECT username FROM sam_admin.app_users WHERE user_id = %s",
            (user_id,), fetchall=False
        )
        if target and target["username"] == ADMIN_USER:
            flash("Cannot delete the bootstrap admin account.", "danger")
            return redirect(url_for("admin_users"))
        del_u = query(
            "SELECT username, role::TEXT FROM sam_admin.app_users WHERE user_id = %s",
            (user_id,), fetchall=False
        )
        execute("DELETE FROM sam_admin.app_users WHERE user_id = %s", (user_id,))
        _audit("user.delete", entity_type="user", entity_id=user_id,
               entity_name=(del_u or {}).get("username"),
               old_values={"role": (del_u or {}).get("role")},
               client_schema="sam_admin")
        flash("User deleted.", "success")

    return redirect(url_for("admin_users"))


# ---------------------------------------------------------------------------
# Audit logging helper
# ---------------------------------------------------------------------------
def _audit(action, entity_type=None, entity_id=None, entity_name=None,
           old_values=None, new_values=None, client_schema=None):
    """Write one row to sam_admin.audit_log. Silently swallows errors so a
    logging failure never breaks the underlying operation."""
    try:
        u = current_user()
        execute(
            """INSERT INTO sam_admin.audit_log
               (username, user_role, action, entity_type, entity_id, entity_name,
                client_schema, old_values, new_values, ip_address)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                u.get("username", "unknown"),
                u.get("role", ""),
                action,
                entity_type,
                str(entity_id) if entity_id is not None else None,
                entity_name,
                client_schema or get_schema(),
                json.dumps(old_values) if old_values else None,
                json.dumps(new_values) if new_values else None,
                request.remote_addr,
            )
        )
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Audit & Snapshots
# Superadmin / DBA / contracting: see all clients (filtered to selected client).
# Client users: read-only, own client snapshots only. No audit log access.
# ---------------------------------------------------------------------------
def _snapshot_client_filter():
    """Return (WHERE clause fragment, params) to scope snapshots by role.

    Privileged roles filter by the currently selected client schema if one is
    active; otherwise they see all clients.  Client-role users are always
    restricted to their own client_id.
    """
    role = current_role()
    u    = current_user()

    if role == "client":
        # Hard-scoped to the user's own client_id
        client_id = u.get("client_id")
        return "AND s.client_id = %s", (client_id,)

    # Privileged roles: filter by the currently selected schema if set
    schema = session.get("client_schema")
    if schema:
        return "AND c.schema_name = %s", (schema,)

    return "", ()


@app.route("/admin/audit")
@roles_required("superadmin", "dba", "contracting", "client")
def admin_audit():
    role   = current_role()
    u      = current_user()
    where, params = _snapshot_client_filter()

    # All clients for the snapshot take-form (privileged roles only)
    all_clients = []
    if role in ("superadmin", "dba", "contracting"):
        all_clients = query(
            "SELECT client_id, client_code, client_name, schema_name "
            "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
        )

    # Licence position snapshots
    try:
        snapshots = query(f"""
            SELECT s.snapshot_id, s.snapshot_month, s.taken_at, s.taken_by, s.note,
                   c.client_name, c.client_code, 'licence_position' AS snap_type,
                   COALESCE(s.snapshot_type, 'manual') AS snapshot_type,
                   s.trigger_feature,
                   (SELECT COUNT(*) FROM sam_admin.licence_snapshot_lines l
                    WHERE l.snapshot_id = s.snapshot_id) AS line_count
            FROM sam_admin.licence_snapshots s
            JOIN sam_admin.clients c ON c.client_id = s.client_id
            WHERE TRUE {where}
            ORDER BY s.snapshot_month DESC, s.taken_at DESC, c.client_name
        """, params)
    except Exception:
        snapshots = []

    # Client shared pool snapshots (not shown to client-role users)
    pool_snapshots = []
    if role in ("superadmin", "dba", "contracting"):
        try:
            pool_snapshots = query(f"""
                SELECT s.snapshot_id, s.snapshot_month, s.taken_at, s.taken_by, s.note,
                       c.client_name, c.client_code, 'shared_pool' AS snap_type,
                       (SELECT COUNT(*) FROM sam_admin.client_pool_snapshot_lines l
                        WHERE l.snapshot_id = s.snapshot_id) AS line_count
                FROM sam_admin.client_pool_snapshots s
                JOIN sam_admin.clients c ON c.client_id = s.client_id
                WHERE TRUE {where}
                ORDER BY s.snapshot_month DESC, c.client_name
            """, params)
        except Exception:
            pool_snapshots = []

    # Audit log — superadmin only
    audit_rows = []
    if role == "superadmin":
        try:
            audit_rows = query("""
                SELECT audit_id, username, user_role, action, entity_type,
                       entity_id, entity_name, client_schema,
                       old_values, new_values, ip_address,
                       created_at
                FROM sam_admin.audit_log
                ORDER BY created_at DESC
                LIMIT 500
            """)
        except Exception:
            pass

    return render_template("admin_audit.html",
                           all_clients=all_clients,
                           snapshots=snapshots,
                           pool_snapshots=pool_snapshots,
                           audit_rows=audit_rows)


@app.route("/admin/snapshot/take", methods=["POST"])
@superadmin_required
def admin_snapshot_take():
    client_id = request.form.get("client_id")
    note      = (request.form.get("note") or "").strip() or None

    client = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE client_id = %s AND is_active",
        (client_id,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("admin_audit"))

    schema = client["schema_name"]
    snap_month = date.today().replace(day=1)  # first of current month

    # Check if a snapshot already exists for this client/month
    existing = query(
        "SELECT snapshot_id FROM sam_admin.licence_snapshots "
        "WHERE client_id = %s AND snapshot_month = %s",
        (client_id, snap_month), fetchall=False
    )
    if existing:
        flash(f"A snapshot for {client['client_name']} already exists for this month "
              f"({snap_month.strftime('%B %Y')}). Delete it first to retake.", "warning")
        return redirect(url_for("admin_audit"))

    # Read the current licence position
    try:
        lp_rows = query(f"""
            SELECT hostname, environment::TEXT AS environment,
                   product_family::TEXT AS product_family,
                   product_detail, licences_required,
                   total_licensed, licence_surplus_deficit, compliance_status::TEXT
            FROM {schema}.license_position
            ORDER BY hostname, product_family, product_detail
        """)
    except Exception as e:
        flash(f"Could not read licence position for {client['client_name']}: {e}", "danger")
        return redirect(url_for("admin_audit"))

    # CSI assignments per (server, product_family, product_detail)
    try:
        assign_rows = query(f"""
            SELECT sv.hostname,
                   m.product_family::TEXT, m.product_detail,
                   cs.csi_number, cs.contract_name
            FROM {schema}.server_csi_map m
            JOIN {schema}.oracle_servers sv ON sv.server_id = m.server_id
            JOIN shared.csi_contracts cs    ON cs.csi_id   = m.csi_id
        """)
        assign_map = {}
        for a in assign_rows:
            key = (a["hostname"], a["product_family"], a["product_detail"] or "")
            assign_map.setdefault(key, []).append(
                f"{a['csi_number']} – {a['contract_name'] or ''}"
            )
    except Exception:
        assign_map = {}

    taken_by = current_user().get("username", "admin")

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO sam_admin.licence_snapshots
                   (client_id, snapshot_month, taken_by, note)
                   VALUES (%s, %s, %s, %s) RETURNING snapshot_id""",
                (client_id, snap_month, taken_by, note)
            )
            snap_id = cur.fetchone()[0]
            for r in lp_rows:
                key = (r["hostname"], r["product_family"], r["product_detail"] or "")
                csides = assign_map.get(key, [])
                cur.execute(
                    """INSERT INTO sam_admin.licence_snapshot_lines
                       (snapshot_id, hostname, environment, product_family, product_detail,
                        licences_required, licences_assigned, surplus_deficit,
                        compliance_status, csi_number, contract_ref)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        snap_id,
                        r["hostname"],
                        r["environment"],
                        r["product_family"],
                        r["product_detail"],
                        r["licences_required"],
                        r["total_licensed"],
                        r["licence_surplus_deficit"],
                        r["compliance_status"],
                        "; ".join(c.split(" – ")[0] for c in csides) or None,
                        "; ".join(csides) or None,
                    )
                )
        conn.commit()

    _audit("snapshot.take",
           entity_type="snapshot", entity_id=snap_id,
           entity_name=f"{client['client_name']} {snap_month.strftime('%Y-%m')}",
           client_schema=schema)

    flash(f"Snapshot taken for {client['client_name']} — "
          f"{snap_month.strftime('%B %Y')} ({len(lp_rows)} lines).", "success")
    return redirect(url_for("admin_audit"))


@app.route("/admin/snapshot/pool/take", methods=["POST"])
@superadmin_required
def admin_pool_snapshot_take():
    """Take a shared pool usage snapshot for a specific client."""
    client_id = request.form.get("client_id")
    note      = (request.form.get("note") or "").strip() or None

    client = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE client_id = %s AND is_active",
        (client_id,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("admin_audit"))

    snap_month = date.today().replace(day=1)

    existing = query(
        "SELECT snapshot_id FROM sam_admin.client_pool_snapshots "
        "WHERE client_id = %s AND snapshot_month = %s",
        (client_id, snap_month), fetchall=False
    )
    if existing:
        flash(f"A shared pool snapshot for {client['client_name']} already exists for "
              f"{snap_month.strftime('%B %Y')}. Delete it first to retake.", "warning")
        return redirect(url_for("admin_audit"))

    # Build live shared pool usage for this client only
    try:
        client_rows = [r for r in _build_shared_pool_live()
                       if r["client_code"] == client["client_code"]]
    except Exception as e:
        flash(f"Could not read shared pool usage for {client['client_name']}: {e}", "danger")
        return redirect(url_for("admin_audit"))

    taken_by = current_user().get("username", "admin")
    line_count = sum(len(r["csi_lines"]) for r in client_rows)

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO sam_admin.client_pool_snapshots
                      (client_id, snapshot_month, taken_by, note)
                    VALUES (%s, %s, %s, %s) RETURNING snapshot_id
                """, (client_id, snap_month, taken_by, note))
                snap_id = cur.fetchone()[0]
                for c in client_rows:
                    for ln in c["csi_lines"]:
                        cur.execute("""
                            INSERT INTO sam_admin.client_pool_snapshot_lines
                              (snapshot_id, csi_number, contract_name, product_name,
                               licences_used, unit_price, monthly_cost)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (snap_id, ln["csi_number"], ln["contract_name"],
                              ln["product_name"], ln["licences_used"],
                              ln["unit_price"], ln["monthly_cost"]))
            conn.commit()
    except Exception as e:
        flash(f"Shared pool snapshot failed: {e}", "danger")
        return redirect(url_for("admin_audit"))

    _audit("snapshot.pool.take",
           entity_type="pool_snapshot", entity_id=snap_id,
           entity_name=f"{client['client_name']} {snap_month.strftime('%Y-%m')}",
           client_schema=client["schema_name"])

    flash(f"Shared pool snapshot taken for {client['client_name']} — "
          f"{snap_month.strftime('%B %Y')} ({line_count} lines).", "success")
    return redirect(url_for("admin_audit"))


@app.route("/admin/snapshot/<int:snapshot_id>")
@roles_required("superadmin", "dba", "contracting", "client")
def admin_snapshot_view(snapshot_id):
    # Client-role users may only view snapshots belonging to their own client
    u    = current_user()
    role = current_role()
    if role == "client":
        snap = query("""
            SELECT s.snapshot_id, s.snapshot_month, s.taken_at, s.taken_by, s.note,
                   COALESCE(s.snapshot_type, 'manual') AS snapshot_type,
                   s.trigger_feature,
                   c.client_name, c.client_code
            FROM sam_admin.licence_snapshots s
            JOIN sam_admin.clients c ON c.client_id = s.client_id
            WHERE s.snapshot_id = %s AND s.client_id = %s
        """, (snapshot_id, u.get("client_id")), fetchall=False)
    else:
        snap = query("""
            SELECT s.snapshot_id, s.snapshot_month, s.taken_at, s.taken_by, s.note,
                   COALESCE(s.snapshot_type, 'manual') AS snapshot_type,
                   s.trigger_feature,
                   c.client_name, c.client_code
            FROM sam_admin.licence_snapshots s
            JOIN sam_admin.clients c ON c.client_id = s.client_id
            WHERE s.snapshot_id = %s
        """, (snapshot_id,), fetchall=False)

    if not snap:
        flash("Snapshot not found.", "danger")
        return redirect(url_for("admin_audit"))

    lines = query("""
        SELECT hostname, environment, product_family, product_detail,
               licences_required, licences_assigned, surplus_deficit,
               compliance_status, csi_number, contract_ref
        FROM sam_admin.licence_snapshot_lines
        WHERE snapshot_id = %s
        ORDER BY hostname, product_family, product_detail
    """, (snapshot_id,))

    return render_template("admin_snapshot_view.html", snap=snap, lines=lines)


@app.route("/admin/snapshot/<int:snapshot_id>/delete", methods=["POST"])
@superadmin_required
def admin_snapshot_delete(snapshot_id):
    snap = query(
        "SELECT s.snapshot_id, c.client_name, s.snapshot_month "
        "FROM sam_admin.licence_snapshots s "
        "JOIN sam_admin.clients c ON c.client_id = s.client_id "
        "WHERE s.snapshot_id = %s",
        (snapshot_id,), fetchall=False
    )
    if snap:
        execute("DELETE FROM sam_admin.licence_snapshots WHERE snapshot_id = %s", (snapshot_id,))
        _audit("snapshot.delete",
               entity_type="snapshot", entity_id=snapshot_id,
               entity_name=f"{snap['client_name']} {snap['snapshot_month']}")
        flash("Snapshot deleted.", "success")
    return redirect(url_for("admin_audit"))


@app.route("/admin/snapshot/pool/<int:snapshot_id>/delete", methods=["POST"])
@superadmin_required
def admin_pool_snapshot_delete(snapshot_id):
    snap = query(
        "SELECT s.snapshot_id, c.client_name, s.snapshot_month "
        "FROM sam_admin.client_pool_snapshots s "
        "JOIN sam_admin.clients c ON c.client_id = s.client_id "
        "WHERE s.snapshot_id = %s",
        (snapshot_id,), fetchall=False
    )
    if snap:
        execute("DELETE FROM sam_admin.client_pool_snapshots WHERE snapshot_id = %s", (snapshot_id,))
        _audit("snapshot.pool.delete",
               entity_type="pool_snapshot", entity_id=snapshot_id,
               entity_name=f"{snap['client_name']} {snap['snapshot_month']}")
        flash("Shared pool snapshot deleted.", "success")
    return redirect(url_for("admin_audit"))


@app.route("/admin/audit/purge", methods=["POST"])
@superadmin_required
def admin_audit_purge():
    result = query("SELECT sam_admin.purge_old_audit_data() AS msg", fetchall=False)
    flash(result["msg"] if result else "Purge complete.", "success")
    return redirect(url_for("admin_audit"))


# ---------------------------------------------------------------------------
# Licence assignment approval queue
# ---------------------------------------------------------------------------

def _can_approve_assignments():
    return current_role() in ("superadmin", "contracting")

def _can_propose_assignments():
    return current_role() in ("superadmin", "contracting", "dba", "client")


@app.route("/assignments", methods=["GET"])
@login_required
def assignment_queue():
    u    = current_user()

    try:
        if _can_approve_assignments():
            pending = query(
                """SELECT r.*, c.client_name
                   FROM sam_admin.assignment_requests r
                   JOIN sam_admin.clients c ON c.client_id = r.client_id
                   WHERE r.status = 'pending'
                   ORDER BY r.proposed_at""")
            history = query(
                """SELECT r.*, c.client_name
                   FROM sam_admin.assignment_requests r
                   JOIN sam_admin.clients c ON c.client_id = r.client_id
                   WHERE r.status != 'pending'
                   ORDER BY r.proposed_at DESC LIMIT 200""")
        else:
            pending = query(
                """SELECT r.*, c.client_name
                   FROM sam_admin.assignment_requests r
                   JOIN sam_admin.clients c ON c.client_id = r.client_id
                   WHERE r.status = 'pending' AND r.proposed_by = %s
                   ORDER BY r.proposed_at""",
                (u.get("username"),))
            history = query(
                """SELECT r.*, c.client_name
                   FROM sam_admin.assignment_requests r
                   JOIN sam_admin.clients c ON c.client_id = r.client_id
                   WHERE r.status != 'pending' AND r.proposed_by = %s
                   ORDER BY r.proposed_at DESC LIMIT 100""",
                (u.get("username"),))
    except Exception:
        flash("Assignment queue table not found — run migration 08_assignment_queue.sql first.", "warning")
        pending = []
        history = []

    return render_template(
        "assignment_queue.html",
        pending=pending,
        history=history,
        can_approve=_can_approve_assignments(),
    )


@app.route("/assignments/propose", methods=["POST"])
@login_required
def assignment_propose():
    """Submit a new assignment (or removal) request for approval."""
    if not _can_propose_assignments():
        abort(403)


    u      = current_user()
    schema = get_schema()

    req_type       = request.form.get("request_type", "assign")
    server_id      = request.form.get("server_id")
    csi_id         = request.form.get("csi_id")
    family         = request.form.get("product_family")
    product_detail = request.form.get("product_detail") or None
    consumed_raw   = request.form.get("licences_consumed") or None
    notes          = request.form.get("notes") or None
    map_id_remove  = request.form.get("map_id_to_remove") or None

    if not server_id or not family:
        flash("Missing required fields.", "danger")
        return redirect(request.referrer or url_for("servers"))

    # Fetch client info for this schema
    client_row = query(
        "SELECT client_id, client_name FROM sam_admin.clients WHERE schema_name = %s",
        (schema,), fetchall=False
    )
    if not client_row:
        flash("Cannot determine client for this schema.", "danger")
        return redirect(request.referrer or url_for("servers"))

    # Fetch server hostname/environment for display
    srv_row = query(
        f"SELECT hostname, environment FROM {schema}.oracle_servers WHERE server_id = %s",
        (server_id,), fetchall=False
    )
    if not srv_row:
        flash("Server not found.", "danger")
        return redirect(request.referrer or url_for("servers"))

    # Fetch CSI number for display
    csi_row = None
    csi_number = None
    if csi_id:
        csi_row = query(
            "SELECT csi_number FROM shared.csi_contracts WHERE csi_id = %s",
            (csi_id,), fetchall=False
        )
        csi_number = csi_row["csi_number"] if csi_row else str(csi_id)

    consumed = None
    if consumed_raw:
        try:
            consumed = float(consumed_raw)
        except ValueError:
            flash("Licences consumed must be a number.", "danger")
            return redirect(request.referrer or url_for("edit_server", server_id=server_id))

    # If the target CSI is a ULA, validate product coverage before queuing
    if csi_id and req_type != "remove":
        is_ula_row = query(
            "SELECT is_ula FROM shared.csi_contracts WHERE csi_id = %s",
            (csi_id,), fetchall=False
        )
        if is_ula_row and is_ula_row["is_ula"]:
            # product_detail from the form is the one being proposed
            if product_detail:
                covered_rows = query(
                    "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s",
                    (csi_id,)
                )
                if covered_rows:
                    covered_lower = [r["product_name"].lower() for r in covered_rows]
                    detail_lower  = product_detail.lower()
                    if not any(cl in detail_lower or detail_lower in cl for cl in covered_lower):
                        flash(
                            f'Cannot propose assignment: "{product_detail}" is not covered by '
                            f'ULA {csi_number}. Add this product to the ULA contract first.',
                            "danger"
                        )
                        return redirect(request.referrer or url_for("edit_server", server_id=server_id))
                else:
                    flash(
                        f"Warning: ULA {csi_number} has no covered products configured — "
                        "coverage cannot be validated. Add products to the contract to enable checks.",
                        "warning"
                    )

    execute(
        """INSERT INTO sam_admin.assignment_requests
           (client_id, client_schema, server_id, hostname, environment,
            csi_id, csi_number, product_family, product_detail,
            licences_consumed, notes, proposed_by, proposed_by_role,
            request_type, map_id_to_remove)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (client_row["client_id"], schema, server_id,
         srv_row["hostname"], srv_row.get("environment"),
         csi_id, csi_number, family, product_detail,
         consumed, notes,
         u.get("username"), u.get("role"),
         req_type, map_id_remove)
    )
    _audit("assignment.propose",
           entity_type="assignment_request",
           entity_id=server_id,
           entity_name=f"{srv_row['hostname']} / {csi_number or 'remove'}",
           new_values={"request_type": req_type, "csi_id": csi_id,
                       "product_family": family, "product_detail": product_detail,
                       "licences_consumed": consumed},
           client_schema=schema)
    flash("Assignment request submitted for approval.", "success")
    return redirect(url_for("edit_server", server_id=server_id))


@app.route("/assignments/<int:request_id>/review", methods=["POST"])
@login_required
def assignment_review(request_id):
    """Approve or reject a pending request. Contracting + superadmin only."""
    if not _can_approve_assignments():
        abort(403)

    u      = current_user()
    action = request.form.get("action")   # approve | reject | withdraw
    note   = request.form.get("review_note") or None

    req = query(
        "SELECT * FROM sam_admin.assignment_requests WHERE request_id = %s",
        (request_id,), fetchall=False
    )
    if not req:
        abort(404)
    if req["status"] != "pending":
        flash("This request is no longer pending.", "warning")
        return redirect(url_for("assignment_queue"))

    if action == "reject":
        execute(
            """UPDATE sam_admin.assignment_requests
               SET status='rejected', reviewed_by=%s, reviewed_at=NOW(), review_note=%s
               WHERE request_id=%s""",
            (u.get("username"), note, request_id)
        )
        _audit("assignment.reject",
               entity_type="assignment_request", entity_id=request_id,
               entity_name=f"{req['hostname']} / {req['csi_number'] or 'remove'}",
               old_values={"status": "pending"},
               new_values={"status": "rejected", "note": note},
               client_schema=req["client_schema"])
        flash("Request rejected.", "warning")

    elif action == "approve":
        schema    = req["client_schema"]
        server_id = req["server_id"]
        csi_id    = req["csi_id"]
        family    = req["product_family"]
        detail    = req["product_detail"]
        consumed  = req["licences_consumed"]
        notes     = req["notes"]
        req_type  = req["request_type"]

        try:
            if req_type == "remove":
                map_id = req["map_id_to_remove"]
                old_csi = query(
                    f"SELECT csi_id, product_family::TEXT, product_detail, licences_consumed "
                    f"FROM {schema}.server_csi_map WHERE map_id = %s",
                    (map_id,), fetchall=False
                )
                execute(
                    f"DELETE FROM {schema}.server_csi_map WHERE map_id = %s",
                    (map_id,)
                )
                _audit("csi.remove", entity_type="csi_assignment", entity_id=server_id,
                       entity_name=f"server {server_id} / map {map_id}",
                       old_values=dict(old_csi) if old_csi else None,
                       client_schema=schema)
            else:
                # assign — upsert into server_csi_map
                execute(f"""
                    DELETE FROM {schema}.server_csi_map
                    WHERE server_id = %s AND csi_id = %s AND product_family = %s
                      AND (product_detail IS NOT DISTINCT FROM %s)
                """, (server_id, csi_id, family, detail))
                execute(f"""
                    INSERT INTO {schema}.server_csi_map
                      (server_id, csi_id, product_family, product_detail,
                       licences_consumed, notes, assigned_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (server_id, csi_id, family, detail, consumed, notes,
                      u.get("username")))
                _audit("csi.assign", entity_type="csi_assignment", entity_id=server_id,
                       entity_name=f"server {server_id} / CSI {csi_id}",
                       new_values={"csi_id": csi_id, "product_family": family,
                                   "product_detail": detail,
                                   "licences_consumed": str(consumed) if consumed else None},
                       client_schema=schema)

            execute(
                """UPDATE sam_admin.assignment_requests
                   SET status='approved', reviewed_by=%s, reviewed_at=NOW(),
                       review_note=%s, applied_at=NOW()
                   WHERE request_id=%s""",
                (u.get("username"), note, request_id)
            )
            _audit("assignment.approve",
                   entity_type="assignment_request", entity_id=request_id,
                   entity_name=f"{req['hostname']} / {req['csi_number'] or 'remove'}",
                   old_values={"status": "pending"},
                   new_values={"status": "approved", "note": note},
                   client_schema=schema)
            flash("Request approved and assignment applied.", "success")

        except Exception as exc:
            flash(f"Failed to apply assignment: {exc}", "danger")

    return redirect(url_for("assignment_queue"))


@app.route("/assignments/<int:request_id>/withdraw", methods=["POST"])
@login_required
def assignment_withdraw(request_id):
    """Withdraw a pending request — only the proposer or an approver can do this."""
    u   = current_user()
    req = query(
        "SELECT * FROM sam_admin.assignment_requests WHERE request_id = %s",
        (request_id,), fetchall=False
    )
    if not req:
        abort(404)
    if req["status"] != "pending":
        flash("Only pending requests can be withdrawn.", "warning")
        return redirect(url_for("assignment_queue"))
    if req["proposed_by"] != u.get("username") and not _can_approve_assignments():
        abort(403)

    execute(
        """UPDATE sam_admin.assignment_requests
           SET status='withdrawn', reviewed_by=%s, reviewed_at=NOW()
           WHERE request_id=%s""",
        (u.get("username"), request_id)
    )
    _audit("assignment.withdraw",
           entity_type="assignment_request", entity_id=request_id,
           entity_name=f"{req['hostname']} / {req['csi_number'] or 'remove'}",
           old_values={"status": "pending"},
           new_values={"status": "withdrawn"},
           client_schema=req["client_schema"])
    flash("Request withdrawn.", "info")
    return redirect(url_for("assignment_queue"))


# ---------------------------------------------------------------------------
# Server discovery conflicts
# ---------------------------------------------------------------------------

@app.route("/admin/discovery-conflicts")
@superadmin_required
def discovery_conflicts():
    conflicts = []
    try:
        conflicts = query(
            "SELECT * FROM sam_admin.list_conflicts() ORDER BY schema_name, server_id"
        ) or []
    except Exception as e:
        app.logger.warning("discovery_conflicts query failed: %s", e)
        flash(f"Could not load conflict data: {e}", "warning")
    try:
        return render_template("discovery_conflicts.html", conflicts=conflicts)
    except Exception as e:
        app.logger.error("discovery_conflicts template error: %s", e)
        raise


@app.route("/admin/discovery-conflicts/<schema>/<int:server_id>/resolve", methods=["POST"])
@superadmin_required
def resolve_conflict(schema, server_id):
    try:
        sql = psycopg2.sql.SQL(
            "UPDATE {}.oracle_servers SET discovery_conflict = FALSE, conflict_detail = NULL WHERE server_id = %s"
        ).format(psycopg2.sql.Identifier(schema))
        execute(sql, (server_id,))
        hostname = request.form.get("hostname") or str(server_id)
        flash(f"Conflict marked as resolved for server {hostname}.", "success")
    except Exception:
        flash("Could not mark resolved — check the server record.", "danger")
    return redirect(url_for("discovery_conflicts"))


# ---------------------------------------------------------------------------
# Manual server registration
# ---------------------------------------------------------------------------

@app.route("/servers/register", methods=["GET", "POST"])
@login_required
def register_server():
    schema = get_schema()
    environments = ["production", "non_production", "development", "test", "dr", "unknown"]
    all_clients  = query(
        "SELECT schema_name, client_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    licensed_options = query(
        "SELECT option_name, display_name, notes FROM shared.oracle_licensed_options WHERE is_active ORDER BY display_name"
    )

    if request.method == "POST":
        # Allow client selection from the form (overrides session schema)
        form_schema = request.form.get("client_schema", "").strip()
        if form_schema:
            schema = form_schema

        if not schema or schema == "__all__":
            flash("Please select a client.", "danger")
            return render_template("register_server.html",
                                   environments=environments, schema=schema,
                                   all_clients=all_clients, licensed_options=licensed_options,
                                   form=request.form)

        hostname    = request.form.get("hostname", "").strip()
        fqdn        = request.form.get("fqdn", "").strip() or None
        ip_address  = request.form.get("ip_address", "").strip() or None
        os_family   = request.form.get("os_family", "").strip() or None
        os_dist     = request.form.get("os_distribution", "").strip() or None
        os_ver      = request.form.get("os_version", "").strip() or None
        environment = request.form.get("environment", "unknown") or "unknown"
        ram_mb      = request.form.get("total_ram_mb", "").strip() or None
        datacenter  = request.form.get("datacenter", "").strip() or None
        cores       = request.form.get("physical_cores", "").strip() or None
        sockets     = request.form.get("cpu_sockets", "").strip() or None
        cps         = request.form.get("cores_per_socket", "").strip() or None
        cpu_model_raw = request.form.get("cpu_model", "").strip()
        cpu_model     = cpu_model_raw.split("|")[0].strip() or "Unknown"
        core_factor = request.form.get("core_factor_override", "").strip() or None
        notes       = request.form.get("notes", "").strip() or None
        server_type = request.form.get("server_type", "oracle_database")  # 'oracle_database' | 'oracle_weblogic'

        # DB-specific
        oracle_sid  = request.form.get("oracle_sid", "").strip() or None
        db_version  = request.form.get("db_version", "").strip() or None
        edition     = request.form.get("edition", "").strip() or None

        # WLS-specific
        domain_name = request.form.get("domain_name", "").strip() or None
        wls_version = request.form.get("wls_version", "").strip() or None
        wls_edition = request.form.get("wls_edition", "").strip() or None

        # DB options (checkboxes — only present in POST when checked)
        selected_options = request.form.getlist("db_options")

        if not hostname:
            flash("Hostname is required.", "danger")
            return render_template("register_server.html",
                                   environments=environments, schema=schema,
                                   all_clients=all_clients, licensed_options=licensed_options,
                                   form=request.form)

        try:
            result = query(
                "SELECT sam_admin.register_server("
                "  p_schema=>%s, p_hostname=>%s, p_source=>'manual',"
                "  p_fqdn=>%s, p_ip_address=>%s,"
                "  p_os_family=>%s, p_os_distribution=>%s, p_os_version=>%s,"
                "  p_environment=>%s, p_total_ram_mb=>%s, p_datacenter=>%s,"
                "  p_physical_cores=>%s, p_cpu_sockets=>%s, p_cores_per_socket=>%s,"
                "  p_notes=>%s"
                ") AS result",
                (schema, hostname, fqdn, ip_address,
                 os_family, os_dist, os_ver,
                 environment,
                 int(ram_mb) if ram_mb else None,
                 datacenter,
                 int(cores) if cores else None,
                 int(sockets) if sockets else None,
                 int(cps) if cps else None,
                 notes),
                fetchall=False
            )
            outcome    = (result or {}).get("result", "")
            is_new     = outcome.startswith("inserted")
            server_id  = int(outcome.split(":")[1]) if ":" in outcome and not outcome.startswith("conflict") else None

            if outcome.startswith("conflict"):
                flash(f"Conflict detected — {outcome[9:]}. Review in Discovery Conflicts.", "warning")
                return render_template("register_server.html",
                                       environments=environments, schema=schema,
                                       all_clients=all_clients, licensed_options=licensed_options,
                                       form=request.form)

            # Create the instance/domain row so the server appears on the correct tab
            if server_id:
                if server_type == "oracle_database":
                    sid = oracle_sid or hostname.upper()
                    inst_row = query(
                        f"INSERT INTO {schema}.oracle_instances"
                        f"  (server_id, oracle_sid, db_name, db_version, edition, is_active)"
                        f"  VALUES (%s, %s, %s, %s, %s, TRUE)"
                        f"  ON CONFLICT (server_id, oracle_sid) DO UPDATE"
                        f"  SET db_version = COALESCE(EXCLUDED.db_version, oracle_instances.db_version),"
                        f"      edition    = COALESCE(EXCLUDED.edition, oracle_instances.edition),"
                        f"      is_active  = TRUE,"
                        f"      last_seen  = NOW()"
                        f"  RETURNING instance_id",
                        (server_id, sid, sid, db_version or None, edition or None),
                        fetchall=False
                    )
                    instance_id = inst_row["instance_id"] if inst_row else None

                    # Insert selected licensed options
                    if instance_id and selected_options:
                        for opt in selected_options:
                            execute(
                                f"INSERT INTO {schema}.oracle_options"
                                f"  (instance_id, option_name, status, recorded_at)"
                                f"  VALUES (%s, %s, 'TRUE', NOW())",
                                (instance_id, opt)
                            )
                elif server_type == "oracle_weblogic":
                    dname = domain_name or f"{hostname}_domain"
                    execute(
                        f"INSERT INTO {schema}.wls_domains"
                        f"  (server_id, domain_name, wls_version, wls_edition, is_active)"
                        f"  VALUES (%s, %s, %s, %s, TRUE)"
                        f"  ON CONFLICT (server_id, domain_name) DO UPDATE"
                        f"  SET wls_version = COALESCE(EXCLUDED.wls_version, wls_domains.wls_version),"
                        f"      wls_edition = COALESCE(EXCLUDED.wls_edition, wls_domains.wls_edition),"
                        f"      is_active   = TRUE,"
                        f"      last_seen   = NOW()",
                        (server_id, dname, wls_version or None, wls_edition or None)
                    )

                # Always ensure a processor row exists so cpu_validation_report
                # can find this server. Use provided values or sensible defaults.
                v_sockets = int(sockets) if sockets else 1
                v_cps     = int(cps) if cps else (
                    (int(cores) // v_sockets) if cores else 1
                )
                existing_proc = query(
                    f"SELECT proc_id FROM {schema}.oracle_processors WHERE server_id = %s LIMIT 1",
                    (server_id,), fetchall=False
                )
                if existing_proc:
                    execute(
                        f"UPDATE {schema}.oracle_processors"
                        f"  SET cpu_model = %s, cpu_sockets = %s, cores_per_socket = %s"
                        f"  WHERE server_id = %s",
                        (cpu_model, v_sockets, v_cps, server_id)
                    )
                else:
                    execute(
                        f"INSERT INTO {schema}.oracle_processors"
                        f"  (server_id, cpu_model, cpu_sockets, cores_per_socket)"
                        f"  VALUES (%s, %s, %s, %s)",
                        (server_id, cpu_model, v_sockets, v_cps)
                    )

                # If a core factor override is given, check whether the current
                # cpu_model already resolves to the correct factor; if not, add
                # an exact-match entry so calculations use the override value.
                if core_factor:
                    existing_factor = query(
                        "SELECT core_factor FROM shared.core_factor_table"
                        " WHERE %s ILIKE processor_pattern"
                        " ORDER BY length(processor_pattern) DESC LIMIT 1",
                        (cpu_model,), fetchall=False
                    )
                    needs_insert = (
                        not existing_factor or
                        abs(float(existing_factor["core_factor"]) - float(core_factor)) > 0.001
                    )
                    if needs_insert:
                        already_exact = query(
                            "SELECT core_factor_id FROM shared.core_factor_table"
                            " WHERE processor_pattern = %s",
                            (cpu_model,), fetchall=False
                        )
                        if already_exact:
                            execute(
                                "UPDATE shared.core_factor_table SET core_factor = %s"
                                " WHERE processor_pattern = %s",
                                (float(core_factor), cpu_model)
                            )
                        else:
                            execute(
                                "INSERT INTO shared.core_factor_table"
                                "  (processor_pattern, core_factor, notes)"
                                "  VALUES (%s, %s, 'Manually set during server registration')",
                                (cpu_model, float(core_factor))
                            )

            u = current_user()
            try:
                query(
                    "SELECT sam_admin.log_discovery_run("
                    "  p_schema=>%s, p_source=>'manual',"
                    "  p_servers_seen=>1,"
                    "  p_servers_new=>%s, p_servers_updated=>%s,"
                    "  p_run_host=>%s, p_status=>'completed'"
                    ")",
                    (schema, 1 if is_new else 0, 0 if is_new else 1,
                     u.get("username") if u else "ui"),
                    fetchall=False
                )
            except Exception:
                pass

            _audit("server.manual_register", entity_type="server", entity_id=server_id,
                   new_values={"hostname": hostname, "server_type": server_type, "source": "manual"},
                   client_schema=schema)

            if is_new:
                flash(f"Server '{hostname}' registered successfully.", "success")
            else:
                flash(f"Server '{hostname}' already existed — record updated.", "info")
            return redirect(url_for("edit_server", server_id=server_id))

        except Exception as e:
            flash(f"Registration failed: {e}", "danger")

    return render_template("register_server.html",
                           environments=environments, schema=schema,
                           all_clients=all_clients, licensed_options=licensed_options,
                           form=request.form)


# ---------------------------------------------------------------------------
# Discovery file upload (JSON or CSV bundle from oracle_discovery scripts)
# ---------------------------------------------------------------------------

import csv as _csv
import io as _io
import uuid as _uuid

def _strip_sqlplus_banner(content: str) -> str:
    """Strip SQL*Plus banner lines before the opening brace."""
    start = content.find("{")
    if start == -1:
        raise ValueError("No JSON object found in file")
    end = content.rfind("}")
    if end == -1:
        raise ValueError("JSON object is not terminated — file may be truncated")
    return content[start : end + 1]


def _call_upsert(schema: str, func: str, payload: dict) -> None:
    with get_db() as conn:
        with conn.cursor() as cur:
            # Set search_path so unqualified type names (e.g. environment_type,
            # virt_type) defined in the client schema resolve correctly.
            cur.execute(
                psycopg2.sql.SQL("SET search_path TO {}, public").format(
                    psycopg2.sql.Identifier(schema)
                )
            )
            cur.execute(
                psycopg2.sql.SQL("SELECT {}.{}(%s::jsonb)").format(
                    psycopg2.sql.Identifier(schema),
                    psycopg2.sql.Identifier(func)
                ),
                (json.dumps(payload),)
            )
        conn.commit()


def _process_json_upload(schema: str, file_obj) -> dict:
    """Process a JSON discovery file and upsert into the given schema."""
    content = file_obj.read().decode("utf-8", errors="replace")
    raw = _strip_sqlplus_banner(content)
    doc = json.loads(raw)

    required = {"_meta", "base", "extended"}
    missing = required - doc.keys()
    if missing:
        raise ValueError(f"JSON is missing top-level keys: {missing}")

    base     = doc["base"]
    extended = doc["extended"]
    hostname = base.get("hostname", "unknown")
    run_id   = base.get("run_id", "")
    messages = []

    # Ensure environment/criticality exist so the enum cast in the DB function
    # never receives NULL (oracle_discovery.sql does not collect these fields).
    base.setdefault("environment", "unknown")
    base.setdefault("criticality", "unknown")

    # Coerce numeric fields that the DB expects as INTEGER but the JSON may
    # carry as floats (e.g. total_ram_mb = 96256.94140625 from v$osstat bytes).
    for field in ("total_ram_mb", "cpu_sockets", "cpu_cores_per_socket",
                  "cpu_threads_per_core", "vcpu_count"):
        if field in base and base[field] is not None:
            base[field] = int(round(float(base[field])))

    # Derive is_exadata from feature_usage when the script didn't detect it
    # directly (older script versions, or currently_used=TRUE/detected_usages=0).
    if not base.get("is_exadata"):
        feature_usage = doc.get("feature_usage", [])
        base["is_exadata"] = any(
            ("exadata" in (f.get("feature_name") or "").lower() or
             "smart scan" in (f.get("feature_name") or "").lower())
            and (f.get("currently_used") in (True, "true", "TRUE") or
                 int(f.get("detected_usages") or 0) > 0)
            for f in feature_usage
        )

    _call_upsert(schema, "upsert_oracle_discovery", base)
    messages.append(f"Server '{hostname}' upserted.")

    _call_upsert(schema, "upsert_oracle_extended_discovery", extended)
    messages.append("Extended discovery (PDBs, RAC nodes, NUP) upserted.")

    # Feature usage — migration 21 format: {run_id, instances:[{sid, feature_usage, pdbs}]}
    # The oracle_discovery.sql JSON puts CDB-level features at doc["feature_usage"] (top-level)
    # and per-PDB features embedded inside each instance's pdbs array.
    feat_payload = doc.get("feature_usage_payload")
    if feat_payload is None:
        top_features = list(doc.get("feature_usage", []))

        # Inject Diagnostics Pack / Tuning Pack from mgmt_pack_summary so they
        # appear in oracle_feature_usage and can be detected on the server page.
        mgmt = doc.get("mgmt_pack_summary", {})
        if mgmt.get("diagnostics_licensed"):
            top_features.append({
                "feature_name": "Diagnostics Pack",
                "db_version": None,
                "detected_usages": 1,
                "total_samples": 1,
                "currently_used": True,
                "first_usage_date": None,
                "last_usage_date": None,
            })
        if mgmt.get("tuning_licensed"):
            top_features.append({
                "feature_name": "Tuning Pack",
                "db_version": None,
                "detected_usages": 1,
                "total_samples": 1,
                "currently_used": True,
                "first_usage_date": None,
                "last_usage_date": None,
            })

        feat_instances = []
        for idx, inst in enumerate(base.get("instances", [])):
            feat_instances.append({
                "sid": inst.get("sid", ""),
                # CDB-level features only assigned to first instance to avoid duplicates in RAC
                "feature_usage": top_features if idx == 0 else [],
                "pdbs": [
                    {
                        "pdb_name": pdb.get("pdb_name", ""),
                        "feature_usage": pdb.get("feature_usage", [])
                    }
                    for pdb in inst.get("pdbs", [])
                ]
            })
        if feat_instances:
            feat_payload = {"run_id": run_id, "instances": feat_instances}
    if feat_payload:
        _call_upsert(schema, "upsert_oracle_feature_usage", feat_payload)
        n_feat = len(doc.get("feature_usage", []))
        messages.append(f"Feature usage upserted ({n_feat} CDB-level feature(s)).")

    # Store Diagnostics Pack / Tuning Pack as oracle_options rows so they appear
    # in the licensed-products card.  oracle_feature_usage cannot be relied upon
    # for these because upsert_oracle_extended_discovery also writes to that table
    # using a conflicting ON CONFLICT key.  oracle_options is written once per
    # instance and is already used to display "Enterprise Edition".
    mgmt = doc.get("mgmt_pack_summary", {})
    pack_options = []
    if mgmt.get("diagnostics_licensed"):
        pack_options.append("Diagnostics Pack")
    if mgmt.get("tuning_licensed"):
        pack_options.append("Tuning Pack")

    # Detect ASO from feature_usage — any of these names indicates ASO is in use
    _aso_keywords = (
        "transparent data encryption", "encrypted tablespace",
        "data redaction", "securefile encryption", "backup encryption",
        "network encryption", "advanced security", "rman encryption",
        "tde", "securefile", "label security",
    )
    top_feature_names = [f.get("feature_name", "").lower() for f in doc.get("feature_usage", [])]
    if any(kw in fn for fn in top_feature_names for kw in _aso_keywords):
        pack_options.append("Advanced Security")
    if pack_options:
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    for sid_inst in base.get("instances", []):
                        sid = sid_inst.get("sid", "")
                        if not sid:
                            continue
                        cur.execute(
                            f"""SELECT i.instance_id
                                FROM {schema}.oracle_instances i
                                JOIN {schema}.oracle_servers   s ON s.server_id = i.server_id
                                WHERE s.hostname = %s AND i.oracle_sid = %s
                                LIMIT 1""",
                            (hostname, sid)
                        )
                        row = cur.fetchone()
                        if not row:
                            continue
                        instance_id = row[0]
                        for pack in pack_options:
                            cur.execute(
                                f"""UPDATE {schema}.oracle_options
                                    SET status = 'TRUE', discovery_run_id = %s
                                    WHERE instance_id = %s AND option_name = %s""",
                                (run_id, instance_id, pack)
                            )
                            if cur.rowcount == 0:
                                cur.execute(
                                    f"""INSERT INTO {schema}.oracle_options
                                          (instance_id, option_name, status, discovery_run_id)
                                        VALUES (%s, %s, 'TRUE', %s)""",
                                    (instance_id, pack, run_id)
                                )
                conn.commit()
            messages.append(f"Management pack access stored ({', '.join(pack_options)}).")
        except Exception as e:
            messages.append(f"Warning: could not store management pack options: {e}")

    meta = doc.get("_meta", {})
    return {
        "success": True,
        "messages": messages,
        "hostname": hostname,
        "db": meta.get("db_unique_name", ""),
        "script_version": meta.get("script_version", "?"),
    }


def _parse_csv_file(file_obj) -> list:
    """Parse an uploaded CSV into a list of dicts, skipping blank/comment lines.
    Values are stripped of surrounding whitespace (SQL*Plus pads columns)."""
    content = file_obj.read().decode("utf-8", errors="replace")
    lines = [l for l in content.splitlines() if l.strip() and not l.startswith("--")]
    reader = _csv.DictReader(_io.StringIO("\n".join(lines)))
    return [{k: (v.strip() if v else "") for k, v in row.items()} for row in reader]


def _safe_int(val, default=0) -> int:
    """Convert a possibly-empty or whitespace-padded string to int."""
    try:
        return int(str(val).strip().replace(",", "") or default)
    except (ValueError, TypeError):
        return default


def _detect_csv_type(filename: str) -> str:
    """Identify a CSV file by its suffix."""
    name = filename.lower()
    for suffix in ("_server", "_instances", "_feature_usage", "_pdb_feature_usage",
                   "_users", "_pdbs", "_mgmt_packs", "_rac_nodes", "_product_usage"):
        if suffix + ".csv" in name:
            return suffix.lstrip("_")
    return "unknown"


def _process_csv_upload(schema: str, files) -> dict:
    """
    Process a bundle of CSV files from oracle_discovery_csv.sql and upsert
    into the given schema.  Requires at minimum _server.csv + _instances.csv.
    """
    parsed = {}
    for f in files:
        if not f.filename:
            continue
        csv_type = _detect_csv_type(f.filename)
        parsed[csv_type] = _parse_csv_file(f)

    if "server" not in parsed:
        raise ValueError("_server.csv is required — cannot identify the host.")
    if "instances" not in parsed:
        raise ValueError("_instances.csv is required — cannot register Oracle instances.")

    server_row = parsed["server"][0]
    hostname   = server_row.get("hostname", "").strip()
    run_id     = "csv-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + _uuid.uuid4().hex[:6]
    messages   = []

    # Build base payload for upsert_oracle_discovery
    instances_list = []
    for row in parsed.get("instances", []):
        instances_list.append({
            "sid":           row.get("instance_name", "").strip(),
            "db_name":       row.get("db_name", "").strip(),
            "edition":       row.get("edition", "").strip(),
            "version":       row.get("version", "").strip(),
            "platform_name": row.get("platform_name", "").strip(),
        })

    base_payload = {
        "hostname":             hostname,
        "fqdn":                 server_row.get("fqdn", hostname).strip(),
        "os_family":            server_row.get("os_family", "").strip(),
        "total_ram_mb":         _safe_int(server_row.get("ram_mb")),
        "cpu_model":            server_row.get("cpu_model", "").strip(),
        "cpu_sockets":          _safe_int(server_row.get("cpu_sockets")),
        "cpu_cores_per_socket": _safe_int(server_row.get("cores_per_socket")),
        "cpu_threads_per_core": _safe_int(server_row.get("threads_per_core")),
        "vcpu_count":           _safe_int(server_row.get("vcpu_count")),
        "virt_type":            server_row.get("virt_type", "unknown").strip() or "unknown",
        "is_vmware":            server_row.get("virt_type", "").strip().lower() == "vmware",
        "is_exadata":           server_row.get("is_exadata", "false").strip().lower() == "true",
        "environment":          "unknown",
        "run_id":               run_id,
        "instances":            instances_list,
    }

    _call_upsert(schema, "upsert_oracle_discovery", base_payload)
    messages.append(f"Server '{hostname}' upserted ({len(instances_list)} instance(s)).")

    # Extended: PDBs + NUP users
    pdbs_list = []
    for row in parsed.get("pdbs", []):
        pdbs_list.append({
            "pdb_name":   row.get("pdb_name", "").strip(),
            "con_id":     _safe_int(row.get("con_id")),
            "open_mode":  row.get("open_mode", "").strip(),
            "restricted": row.get("restricted", "NO").strip(),
        })

    nup_rows   = parsed.get("users", [])
    nup_total  = 0
    nup_active = 0
    for r in nup_rows:
        cat = r.get("category", "").lower()
        cnt = _safe_int(r.get("user_count"))
        if "total" in cat:
            nup_total = cnt
        elif "open" in cat or "active" in cat:
            nup_active = cnt

    extended_payload = {
        "hostname": hostname,
        "run_id":   run_id,
        "pdbs":     pdbs_list,
        "nup_users": {"total": nup_total, "active": nup_active},
    }
    _call_upsert(schema, "upsert_oracle_extended_discovery", extended_payload)
    if pdbs_list:
        messages.append(f"PDB topology upserted ({len(pdbs_list)} PDB(s)).")
    if nup_total:
        messages.append(f"NUP user counts upserted (total={nup_total}, active={nup_active}).")

    # Feature usage — build migration-21 payload
    def _feat_row_to_dict(row):
        return {
            "feature_name":    row.get("feature_name", "").strip().strip('"'),
            "db_version":      row.get("db_version", "").strip(),
            "detected_usages": _safe_int(row.get("detected_usages")),
            "total_samples":   _safe_int(row.get("total_samples")),
            "currently_used":  row.get("currently_used", "FALSE").strip().upper() == "TRUE",
            "first_usage_date": row.get("first_usage_date") or None,
            "last_usage_date":  row.get("last_usage_date")  or None,
        }

    if instances_list and ("feature_usage" in parsed or "pdb_feature_usage" in parsed):
        # Group PDB features by pdb_name
        pdb_feats: dict = {}
        for row in parsed.get("pdb_feature_usage", []):
            pdb = row.get("pdb_name", "").strip()
            pdb_feats.setdefault(pdb, []).append(_feat_row_to_dict(row))

        # Attach features to the first (and usually only) instance at CDB level
        sid = instances_list[0]["sid"]
        feat_instances = [{
            "sid":           sid,
            "feature_usage": [_feat_row_to_dict(r) for r in parsed.get("feature_usage", [])],
            "pdbs": [
                {"pdb_name": pdb_name, "feature_usage": feats}
                for pdb_name, feats in pdb_feats.items()
            ],
        }]
        _call_upsert(schema, "upsert_oracle_feature_usage",
                     {"run_id": run_id, "instances": feat_instances})
        n_cdb = len(feat_instances[0]["feature_usage"])
        n_pdb = sum(len(v) for v in pdb_feats.values())
        messages.append(f"Feature usage upserted ({n_cdb} CDB features, {n_pdb} PDB features).")

    # Store Diagnostics Pack / Tuning Pack from _mgmt_packs.csv and
    # Advanced Security from feature_usage keywords as oracle_options rows.
    pack_options = []
    mgmt_rows = parsed.get("mgmt_packs", [])
    if any(r.get("diagnostics_licensed", "").strip().upper() == "YES" for r in mgmt_rows):
        pack_options.append("Diagnostics Pack")
    if any(r.get("tuning_licensed", "").strip().upper() == "YES" for r in mgmt_rows):
        pack_options.append("Tuning Pack")

    _aso_keywords = (
        "transparent data encryption", "encrypted tablespace",
        "data redaction", "securefile encryption", "backup encryption",
        "network encryption", "advanced security", "rman encryption",
        "tde", "securefile", "label security",
    )
    cdb_feature_names = [
        _feat_row_to_dict(r).get("feature_name", "").lower()
        for r in parsed.get("feature_usage", [])
    ]
    if any(kw in fn for fn in cdb_feature_names for kw in _aso_keywords):
        pack_options.append("Advanced Security")

    if pack_options and instances_list:
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    sid = instances_list[0]["sid"]
                    cur.execute(
                        f"""SELECT i.instance_id
                            FROM {schema}.oracle_instances i
                            JOIN {schema}.oracle_servers   s ON s.server_id = i.server_id
                            WHERE s.hostname = %s AND i.oracle_sid = %s
                            LIMIT 1""",
                        (hostname, sid)
                    )
                    row = cur.fetchone()
                    if row:
                        instance_id = row[0]
                        for pack in pack_options:
                            cur.execute(
                                f"""UPDATE {schema}.oracle_options
                                    SET status = 'TRUE', discovery_run_id = %s
                                    WHERE instance_id = %s AND option_name = %s""",
                                (run_id, instance_id, pack)
                            )
                            if cur.rowcount == 0:
                                cur.execute(
                                    f"""INSERT INTO {schema}.oracle_options
                                          (instance_id, option_name, status, discovery_run_id)
                                        VALUES (%s, %s, 'TRUE', %s)""",
                                    (instance_id, pack, run_id)
                                )
                conn.commit()
            messages.append(f"Management pack access stored ({', '.join(pack_options)}).")
        except Exception as e:
            messages.append(f"Warning: could not store management pack options: {e}")

    return {
        "success":  True,
        "messages": messages,
        "hostname": hostname,
        "db":       instances_list[0]["db_name"] if instances_list else "",
    }


def _process_wls_json_upload(schema: str, file_obj) -> dict:
    """Process a WebLogic discovery JSON file (from run_wls_discovery.sh)."""
    raw = file_obj.read().decode("utf-8", errors="replace")
    # Strip any shell banner / non-JSON preamble lines
    lines = raw.splitlines()
    json_start = next((i for i, l in enumerate(lines) if l.strip().startswith("{")), None)
    if json_start is None:
        raise ValueError("No JSON object found in the uploaded file.")
    doc = json.loads("\n".join(lines[json_start:]))

    hostname = doc.get("hostname") or doc.get("fqdn", "")
    if not hostname:
        raise ValueError("JSON is missing 'hostname' field.")

    messages = []

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {schema}.upsert_wls_discovery(%s::jsonb)",
                (json.dumps(doc),)
            )
        conn.commit()

    domains = doc.get("domains", [])
    total_ms = sum(len(d.get("managed_servers", [])) for d in domains)
    messages.append(f"Server upserted: {hostname}")
    messages.append(f"Domains loaded: {len(domains)}")
    messages.append(f"Managed servers loaded: {total_ms}")

    return {"success": True, "hostname": hostname, "db": "", "messages": messages}


def _process_wls_csv_upload(schema: str, files) -> dict:
    """Process CSV files from wls_discovery_csv.sh."""
    import io as _io

    named = {}
    for f in files:
        if f and f.filename:
            named[f.filename.lower()] = f

    def find_csv(suffix):
        for name, fobj in named.items():
            if name.endswith(suffix):
                return fobj
        return None

    server_file  = find_csv("_wls_server.csv")
    domains_file = find_csv("_wls_domains.csv")

    if not server_file:
        raise ValueError("Missing *_wls_server.csv — cannot proceed.")
    if not domains_file:
        raise ValueError("Missing *_wls_domains.csv — cannot proceed.")

    def read_csv(fobj):
        text = fobj.read().decode("utf-8", errors="replace")
        reader = csv.DictReader(_io.StringIO(text))
        return [dict(r) for r in reader]

    server_rows  = read_csv(server_file)
    domains_rows = read_csv(domains_file)
    ms_file      = find_csv("_wls_managed_servers.csv")
    prods_file   = find_csv("_wls_products.csv")
    ms_rows      = read_csv(ms_file)   if ms_file   else []
    prod_rows    = read_csv(prods_file) if prods_file else []

    if not server_rows:
        raise ValueError("*_wls_server.csv is empty.")

    srv = server_rows[0]
    hostname = srv.get("hostname") or srv.get("fqdn", "")
    if not hostname:
        raise ValueError("*_wls_server.csv is missing 'hostname'.")

    run_id       = srv.get("run_id", f"csv-{hostname}")
    discovered_at = srv.get("discovered_at", "")

    # Build domains list
    domain_map = {}
    for dr in domains_rows:
        domain_map[dr["domain_name"]] = {
            "domain_name":       dr.get("domain_name", ""),
            "domain_home":       dr.get("domain_home", ""),
            "wls_version":       dr.get("wls_version") or None,
            "wls_edition":       dr.get("wls_edition") or None,
            "admin_server_host": dr.get("admin_server_host", ""),
            "admin_server_port": _safe_int(dr.get("admin_server_port")),
            "managed_servers":   [],
            "installed_products": [],
        }

    for ms in ms_rows:
        dn = ms.get("domain_name", "")
        if dn in domain_map:
            domain_map[dn]["managed_servers"].append({
                "name":        ms.get("managed_server_name", ""),
                "listen_port": _safe_int(ms.get("listen_port")),
                "ssl_port":    _safe_int(ms.get("ssl_port")) or None,
                "cluster":     ms.get("cluster_name") or None,
                "machine":     ms.get("machine_name") or None,
                "state":       "UNKNOWN",
            })

    for pr in prod_rows:
        dn = pr.get("domain_name", "")
        if dn in domain_map:
            domain_map[dn]["installed_products"].append({
                "name":    pr.get("product_name", ""),
                "version": pr.get("product_version", ""),
                "home":    pr.get("home_path", ""),
            })

    payload = {
        "run_id":             run_id,
        "hostname":           hostname,
        "fqdn":               srv.get("fqdn", hostname),
        "ip_address":         srv.get("ip_address", ""),
        "os_family":          srv.get("os_family", ""),
        "os_distribution":    srv.get("os_distribution", ""),
        "os_version":         srv.get("os_version", ""),
        "environment":        srv.get("environment", "unknown"),
        "criticality":        srv.get("criticality", "unknown"),
        "datacenter":         srv.get("datacenter", ""),
        "cpu_sockets":        _safe_int(srv.get("cpu_sockets")),
        "cpu_cores_per_socket": _safe_int(srv.get("cores_per_socket")),
        "cpu_threads_per_core": _safe_int(srv.get("threads_per_core")),
        "cpu_model":          srv.get("cpu_model", ""),
        "cpu_architecture":   srv.get("cpu_architecture", ""),
        "virt_type":          srv.get("virt_type", "physical"),
        "is_vmware":          srv.get("is_vmware", "false").lower() == "true",
        "total_ram_mb":       _safe_int(srv.get("total_ram_mb")),
        "discovered_at":      discovered_at,
        "domains":            list(domain_map.values()),
    }

    messages = []
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT {schema}.upsert_wls_discovery(%s::jsonb)",
                (json.dumps(payload),)
            )
        conn.commit()

    messages.append(f"Server upserted: {hostname}")
    messages.append(f"Domains loaded: {len(domain_map)}")
    total_ms = sum(len(d["managed_servers"]) for d in domain_map.values())
    messages.append(f"Managed servers loaded: {total_ms}")
    if ms_file:
        messages.append(f"Managed servers file processed ({len(ms_rows)} rows).")
    if prods_file:
        messages.append(f"Products file processed ({len(prod_rows)} rows).")

    return {"success": True, "hostname": hostname, "db": "", "messages": messages}


@app.route("/servers/upload-discovery", methods=["GET", "POST"])
@login_required
def upload_discovery():
    schema  = get_schema()
    clients = get_clients()
    result  = None

    if request.method == "POST":
        target_schema = request.form.get("target_schema", schema)
        upload_type   = request.form.get("upload_type", "json")
        try:
            if upload_type == "json":
                f = request.files.get("json_file")
                if not f or not f.filename:
                    raise ValueError("No JSON file selected.")
                result = _process_json_upload(target_schema, f)
            elif upload_type == "csv":
                files = request.files.getlist("csv_files")
                if not files or not any(f.filename for f in files):
                    raise ValueError("No CSV files selected.")
                result = _process_csv_upload(target_schema, files)
            elif upload_type == "wls_json":
                f = request.files.get("wls_json_file")
                if not f or not f.filename:
                    raise ValueError("No WebLogic JSON file selected.")
                result = _process_wls_json_upload(target_schema, f)
            elif upload_type == "wls_csv":
                files = request.files.getlist("wls_csv_files")
                if not files or not any(f.filename for f in files):
                    raise ValueError("No WebLogic CSV files selected.")
                result = _process_wls_csv_upload(target_schema, files)
            else:
                raise ValueError(f"Unknown upload type: {upload_type}")
        except Exception as exc:
            result = {"success": False, "error": str(exc), "messages": []}

    return render_template("upload_discovery.html",
                           clients=clients,
                           result=result,
                           active_schema=schema)


# ---------------------------------------------------------------------------
# Discovery run history
# ---------------------------------------------------------------------------

@app.route("/servers/discovery-history")
@login_required
def discovery_history():
    schema = get_schema()

    try:
        if schema == "__all__":
            runs = query("""
                SELECT r.*, c.client_name
                FROM sam_admin.discovery_runs r
                JOIN sam_admin.clients c ON c.client_id = r.client_id
                ORDER BY r.started_at DESC
                LIMIT 500
            """)
        else:
            client = query(
                "SELECT client_id, client_name FROM sam_admin.clients WHERE schema_name = %s",
                (schema,), fetchall=False
            )
            runs = query("""
                SELECT r.*, c.client_name
                FROM sam_admin.discovery_runs r
                JOIN sam_admin.clients c ON c.client_id = r.client_id
                WHERE r.client_schema = %s
                ORDER BY r.started_at DESC
                LIMIT 500
            """, (schema,)) if client else []
    except Exception as e:
        flash(f"Could not load discovery history: {e}", "warning")
        runs = []

    # Per-server last-seen timeline for the active schema
    server_timeline = []
    if schema != "__all__":
        try:
            server_timeline = query(f"""
                SELECT server_id, hostname, fqdn, ip_address::TEXT,
                       discovery_source, last_discovery_run,
                       first_seen, last_seen,
                       discovery_conflict
                FROM {schema}.oracle_servers
                ORDER BY last_seen DESC NULLS LAST
            """)
        except Exception:
            pass

    return render_template("discovery_history.html",
                           runs=runs,
                           server_timeline=server_timeline,
                           schema=schema)


# ---------------------------------------------------------------------------
# Multi-client switcher
# ---------------------------------------------------------------------------
@app.route("/switch-client", methods=["POST"])
@login_required
def switch_client():
    schema = request.form.get("schema", "")
    if schema == "__all__":
        session["client_schema"] = "__all__"
        flash("Viewing all clients.", "success")
        return redirect(request.referrer or url_for("servers"))
    row = query(
        "SELECT schema_name, client_name FROM sam_admin.clients "
        "WHERE schema_name = %s AND is_active",
        (schema,), fetchall=False
    )
    if row:
        session["client_schema"] = schema
        flash(f"Switched to {row['client_name']}.", "success")
    else:
        flash("Unknown client schema.", "danger")
    return redirect(request.referrer or url_for("servers"))


# ---------------------------------------------------------------------------
# Client management
# ---------------------------------------------------------------------------
@app.route("/clients", methods=["GET", "POST"])
@login_required
def clients():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "provision":
            code    = request.form.get("client_code", "").strip().lower()
            name    = request.form.get("client_name", "").strip()
            contact = request.form.get("contact_email", "").strip() or None
            if not code or not name:
                flash("Client code and name are required.", "danger")
            else:
                try:
                    result = query(
                        "SELECT sam_admin.provision_client(%s, %s, %s) AS msg",
                        (code, name, contact), fetchall=False
                    )
                    flash(result["msg"], "success")
                except Exception as e:
                    flash(f"Error provisioning client: {e}", "danger")

        elif action == "deactivate":
            client_id = request.form.get("client_id")
            execute(
                "UPDATE sam_admin.clients SET is_active = FALSE WHERE client_id = %s",
                (client_id,)
            )
            flash("Client deactivated.", "warning")

        elif action == "reactivate":
            client_id = request.form.get("client_id")
            execute(
                "UPDATE sam_admin.clients SET is_active = TRUE WHERE client_id = %s",
                (client_id,)
            )
            flash("Client reactivated.", "success")

        return redirect(url_for("clients"))

    all_clients = query("""
        SELECT c.client_id, c.client_code, c.client_name, c.schema_name,
               c.contact_email, c.is_active, c.created_at::DATE AS created_at,
               (SELECT COUNT(*) FROM pg_tables
                WHERE schemaname = c.schema_name)           AS table_count,
               (SELECT COUNT(*) FROM sam_admin.discovery_runs dr
                WHERE dr.client_id = c.client_id)           AS discovery_runs,
               (SELECT MAX(started_at)::DATE
                FROM sam_admin.discovery_runs dr
                WHERE dr.client_id = c.client_id)           AS last_discovery
        FROM sam_admin.clients c
        ORDER BY c.is_active DESC, c.client_name
    """)
    return render_template("clients.html", clients=all_clients)


# ---------------------------------------------------------------------------
# Servers list
# ---------------------------------------------------------------------------
@app.route("/servers/weblogic")
@login_required
def weblogic_servers():
    schema = get_schema()
    rows = []

    def _wls_query(s, client_code=None, client_name=None):
        results = query(f"""
            WITH ula_servers AS (
                SELECT DISTINCT m.server_id
                FROM {s}.server_csi_map m
                JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id
                WHERE cs.is_ula
            )
            SELECT
                sv.server_id,
                sv.hostname,
                sv.environment::TEXT,
                sv.datacenter,
                sv.last_seen::DATE AS last_seen,
                d.domain_id,
                d.domain_name,
                d.wls_edition,
                d.wls_version,
                d.admin_server_host,
                d.admin_server_port,
                COUNT(DISTINCT ms.managed_server_id) AS managed_server_count,
                COUNT(DISTINCT ms.cluster_name)
                  FILTER (WHERE ms.cluster_name IS NOT NULL) AS cluster_count,
                STRING_AGG(DISTINCT ms.cluster_name, ', '
                  ORDER BY ms.cluster_name)
                  FILTER (WHERE ms.cluster_name IS NOT NULL) AS clusters,
                (SELECT COUNT(*) FROM {s}.wls_installed_products ip
                 WHERE ip.domain_id = d.domain_id)           AS product_count,
                lp.licences_required,
                lp.total_licensed,
                lp.compliance_status,
                (ula_servers.server_id IS NOT NULL)          AS has_ula
            FROM {s}.oracle_servers sv
            JOIN {s}.wls_domains d ON d.server_id = sv.server_id AND d.is_active
            LEFT JOIN {s}.wls_managed_servers ms ON ms.domain_id = d.domain_id
            LEFT JOIN {s}.license_position lp
                   ON lp.server_id = sv.server_id
                  AND lp.product_family = 'oracle_weblogic'
            LEFT JOIN ula_servers ON ula_servers.server_id = sv.server_id
            WHERE sv.is_active
            GROUP BY sv.server_id, sv.hostname, sv.environment, sv.datacenter,
                     sv.last_seen, d.domain_id, d.domain_name, d.wls_edition,
                     d.wls_version, d.admin_server_host, d.admin_server_port,
                     lp.licences_required, lp.total_licensed, lp.compliance_status,
                     ula_servers.server_id
            ORDER BY sv.hostname, d.domain_name
        """)
        for r in results:
            r = dict(r)
            r["_client_code"] = client_code or s
            r["_client_name"] = client_name or s
            r["_schema"]      = s
            rows.append(r)

    if schema == "__all__":
        clients_list = query(
            "SELECT schema_name, client_name, client_code FROM sam_admin.clients "
            "WHERE is_active ORDER BY client_name"
        )
        for c in clients_list:
            try:
                _wls_query(c["schema_name"], c["client_code"], c["client_name"])
            except Exception:
                pass
    else:
        try:
            _wls_query(schema)
        except Exception:
            pass

    return render_template("weblogic_servers.html", servers=rows, schema=schema)


@app.route("/")
@app.route("/servers")
@login_required
def servers():
    schema = get_schema()

    if schema == "__all__":
        active_clients = query(
            "SELECT schema_name, client_name, client_code FROM sam_admin.clients "
            "WHERE is_active ORDER BY client_name"
        )
        rows = []
        for c in active_clients:
            s = c["schema_name"]
            try:
                client_rows = query(f"""
                    WITH lp AS (
                        SELECT server_id,
                            JSONB_AGG(JSONB_BUILD_OBJECT(
                                'product_family',    product_family,
                                'product_detail',    product_detail,
                                'licences_required', licences_required,
                                'total_licensed',    total_licensed,
                                'compliance_status', compliance_status
                            ) ORDER BY product_family) AS licence_rows,
                            SUM(licences_required)                        AS total_licences_required,
                            BOOL_OR(compliance_status = 'under_licensed') AS any_under_licensed,
                            STRING_AGG(product_family||': '||licences_required::TEXT,
                                       ' | ' ORDER BY product_family)     AS licence_summary
                        FROM {s}.license_position
                        WHERE product_family = 'oracle_database'
                        GROUP BY server_id
                    ),
                    cpu_issues AS (
                        SELECT server_id, factor_unknown
                        FROM   {s}.cpu_validation_report WHERE factor_unknown
                    ),
                    ula_servers AS (
                        SELECT DISTINCT m.server_id
                        FROM {s}.server_csi_map m
                        JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id
                        WHERE cs.is_ula
                    ),
                    exadata AS (
                        SELECT DISTINCT server_id
                        FROM   {s}.oracle_processors WHERE is_exadata = TRUE
                    )
                    SELECT
                        s.server_id, s.hostname, s.environment::TEXT, s.datacenter,
                        s.ip_address::TEXT, s.last_seen::DATE AS last_seen,
                        COALESCE(s.licence_metric_override,'processor_perpetual') AS licence_metric,
                        s.licence_metric_override IS NOT NULL AS metric_overridden,
                        COUNT(DISTINCT m.csi_id) AS csi_count,
                        lp.licence_rows,
                        COALESCE(lp.total_licences_required,0) AS total_licences_required,
                        COALESCE(lp.any_under_licensed,FALSE)  AS any_under_licensed,
                        lp.licence_summary,
                        COALESCE(cpu_issues.factor_unknown,FALSE) AS cpu_unvalidated,
                        (ula_servers.server_id IS NOT NULL)    AS has_ula,
                        (exadata.server_id IS NOT NULL)        AS is_exadata
                    FROM {s}.oracle_servers s
                    JOIN {s}.oracle_instances i ON i.server_id = s.server_id AND i.is_active
                    LEFT JOIN {s}.server_csi_map m ON m.server_id = s.server_id
                    LEFT JOIN lp                   ON lp.server_id = s.server_id
                    LEFT JOIN cpu_issues           ON cpu_issues.server_id = s.server_id
                    LEFT JOIN ula_servers          ON ula_servers.server_id = s.server_id
                    LEFT JOIN exadata              ON exadata.server_id = s.server_id
                    WHERE s.is_active
                    GROUP BY s.server_id, s.hostname, s.environment, s.datacenter,
                             s.ip_address, s.last_seen, s.licence_metric_override,
                             lp.licence_rows, lp.total_licences_required,
                             lp.any_under_licensed, lp.licence_summary, cpu_issues.factor_unknown,
                             ula_servers.server_id, exadata.server_id
                    ORDER BY s.hostname
                """)
                for r in client_rows:
                    r = dict(r)
                    r["_client_name"] = c["client_name"]
                    r["_client_code"] = c["client_code"]
                    r["_schema"]      = s
                    rows.append(r)
            except Exception:
                pass
        se2_count = 0
        return render_template("servers.html", servers=rows, schema="__all__",
                               se2_count=se2_count)

    rows = query(f"""
        WITH lp AS (
            SELECT
                server_id,
                JSONB_AGG(
                    JSONB_BUILD_OBJECT(
                        'product_family',    product_family,
                        'product_detail',    product_detail,
                        'licences_required', licences_required,
                        'total_licensed',    total_licensed,
                        'compliance_status', compliance_status
                    ) ORDER BY product_family
                ) AS licence_rows,
                SUM(licences_required)                        AS total_licences_required,
                BOOL_OR(compliance_status = 'under_licensed') AS any_under_licensed,
                STRING_AGG(
                    product_family || ': ' || licences_required::TEXT,
                    ' | ' ORDER BY product_family
                )                                             AS licence_summary
            FROM {schema}.license_position
            WHERE product_family = 'oracle_database'
            GROUP BY server_id
        ),
        cpu_issues AS (
            SELECT server_id, factor_unknown
            FROM   {schema}.cpu_validation_report
            WHERE  factor_unknown = TRUE
        ),
        ula_servers AS (
            SELECT DISTINCT m.server_id
            FROM {schema}.server_csi_map m
            JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id
            WHERE cs.is_ula
        ),
        exadata AS (
            SELECT DISTINCT server_id
            FROM   {schema}.oracle_processors
            WHERE  is_exadata = TRUE
        )
        SELECT
            s.server_id,
            s.hostname,
            s.environment::TEXT,
            s.datacenter,
            s.ip_address::TEXT,
            s.last_seen::DATE                                  AS last_seen,
            COALESCE(s.licence_metric_override, 'processor_perpetual') AS licence_metric,
            s.licence_metric_override IS NOT NULL              AS metric_overridden,
            COUNT(DISTINCT m.csi_id)                           AS csi_count,
            lp.licence_rows,
            COALESCE(lp.total_licences_required, 0)            AS total_licences_required,
            COALESCE(lp.any_under_licensed, FALSE)             AS any_under_licensed,
            lp.licence_summary,
            COALESCE(cpu_issues.factor_unknown, FALSE)         AS cpu_unvalidated,
            (ula_servers.server_id IS NOT NULL)                AS has_ula,
            (exadata.server_id IS NOT NULL)                    AS is_exadata
        FROM {schema}.oracle_servers s
        JOIN {schema}.oracle_instances i       ON i.server_id = s.server_id AND i.is_active
        LEFT JOIN {schema}.server_csi_map m    ON m.server_id = s.server_id
        LEFT JOIN lp                           ON lp.server_id = s.server_id
        LEFT JOIN cpu_issues                   ON cpu_issues.server_id = s.server_id
        LEFT JOIN ula_servers                  ON ula_servers.server_id = s.server_id
        LEFT JOIN exadata                      ON exadata.server_id = s.server_id
        WHERE s.is_active = TRUE
        GROUP BY s.server_id, s.hostname, s.environment, s.datacenter,
                 s.ip_address, s.last_seen, s.licence_metric_override,
                 lp.licence_rows, lp.total_licences_required,
                 lp.any_under_licensed, lp.licence_summary, cpu_issues.factor_unknown,
                 ula_servers.server_id, exadata.server_id
        ORDER BY s.hostname
    """)

    # Tag each row with the client name/code for the Owner column
    client_info = query(
        "SELECT client_name, client_code FROM sam_admin.clients WHERE schema_name = %s",
        (schema,), fetchall=False
    ) or {}
    rows = [dict(r,
                 _client_name=client_info.get("client_name", schema),
                 _client_code=client_info.get("client_code", schema))
            for r in rows]

    # SE2 violations for badge count in header
    try:
        se2_count = len(query(f"SELECT 1 FROM {schema}.se2_violations"))
    except Exception:
        se2_count = 0

    return render_template("servers.html", servers=rows, schema=schema,
                           se2_count=se2_count)


# ---------------------------------------------------------------------------
# Stale servers
# ---------------------------------------------------------------------------
STALE_THRESHOLD_DAYS = 14

@app.route("/servers/stale")
@login_required
def stale_servers():
    active_clients = query(
        "SELECT client_id, schema_name, client_name, client_code FROM sam_admin.clients "
        "WHERE is_active ORDER BY client_name"
    )
    rows = []
    for c in active_clients:
        s = c["schema_name"]
        try:
            client_rows = query(f"""
                SELECT
                    s.server_id,
                    s.hostname,
                    s.environment::TEXT AS environment,
                    s.datacenter,
                    s.ip_address::TEXT  AS ip_address,
                    s.last_seen::DATE   AS last_seen,
                    (CURRENT_DATE - s.last_seen::DATE) AS days_missing,
                    s.discovery_source
                FROM {s}.oracle_servers s
                WHERE s.is_active
                  AND s.last_seen < NOW() - INTERVAL '{STALE_THRESHOLD_DAYS} days'
                ORDER BY s.last_seen ASC
            """)
            for r in client_rows:
                r = dict(r)
                r["_client_name"]   = c["client_name"]
                r["_client_code"]   = c["client_code"]
                r["_client_schema"] = s
                rows.append(r)
        except Exception:
            pass

    # Attach any existing investigations
    if rows:
        inv_map = {}
        invs = query(
            "SELECT * FROM sam_admin.stale_server_investigations WHERE status NOT IN ('resolved','dismissed')"
        )
        for inv in invs:
            inv_map[(inv["client_schema"], inv["server_id"])] = inv

        for r in rows:
            r["investigation"] = inv_map.get((r["_client_schema"], r["server_id"]))

    app_users = query(
        "SELECT username, display_name FROM sam_admin.app_users WHERE is_active ORDER BY display_name"
    )

    # Decommissioned archive — show all clients' records
    schema = get_schema()
    if schema == "__all__":
        decomm_rows = query("""
            SELECT * FROM sam_admin.decommissioned_servers
            ORDER BY decommissioned_at DESC
        """)
    else:
        decomm_rows = query("""
            SELECT * FROM sam_admin.decommissioned_servers
            WHERE client_schema = %s ORDER BY decommissioned_at DESC
        """, (schema,))

    return render_template("stale_servers.html", rows=rows, app_users=app_users,
                           decomm_rows=decomm_rows, threshold=STALE_THRESHOLD_DAYS)


@app.route("/servers/stale/investigate", methods=["POST"])
@login_required
def stale_investigate_open():
    client_schema = request.form["client_schema"]
    server_id     = int(request.form["server_id"])
    hostname      = request.form["hostname"]
    assigned_to   = request.form.get("assigned_to") or None
    notes         = request.form.get("notes") or None
    username      = session.get("username", "system")

    execute("""
        INSERT INTO sam_admin.stale_server_investigations
            (client_schema, server_id, hostname, assigned_to, notes, opened_by, status)
        VALUES (%s, %s, %s, %s, %s, %s, 'open')
        ON CONFLICT (client_schema, server_id) DO UPDATE
            SET assigned_to = EXCLUDED.assigned_to,
                notes       = COALESCE(EXCLUDED.notes, stale_server_investigations.notes),
                status      = CASE WHEN stale_server_investigations.status IN ('resolved','dismissed')
                                   THEN 'open' ELSE stale_server_investigations.status END,
                opened_by   = EXCLUDED.opened_by,
                opened_at   = CASE WHEN stale_server_investigations.status IN ('resolved','dismissed')
                                   THEN NOW() ELSE stale_server_investigations.opened_at END
    """, (client_schema, server_id, hostname, assigned_to, notes, username))

    _audit("stale_server.open", entity_type="server", entity_name=hostname,
           client_schema=client_schema,
           new_values={"assigned_to": assigned_to, "notes": notes})

    flash(f"Investigation opened for {hostname}.", "success")
    return redirect(url_for("stale_servers"))


@app.route("/servers/stale/<int:investigation_id>/update", methods=["POST"])
@login_required
def stale_investigate_update(investigation_id):
    status      = request.form["status"]
    assigned_to = request.form.get("assigned_to") or None
    notes       = request.form.get("notes") or None
    username    = session.get("username", "system")

    inv = query(
        "SELECT * FROM sam_admin.stale_server_investigations WHERE investigation_id = %s",
        (investigation_id,), fetchall=False
    )
    if not inv:
        flash("Investigation not found.", "danger")
        return redirect(url_for("stale_servers"))

    resolved_at = "NOW()" if status in ("resolved", "dismissed") else "NULL"
    execute(f"""
        UPDATE sam_admin.stale_server_investigations
        SET status      = %s,
            assigned_to = %s,
            notes       = %s,
            resolved_by = %s,
            resolved_at = {resolved_at}
        WHERE investigation_id = %s
    """, (status, assigned_to, notes, username if status in ("resolved","dismissed") else None,
          investigation_id))

    _audit("stale_server.update", entity_type="server", entity_name=inv["hostname"],
           client_schema=inv["client_schema"],
           old_values={"status": inv["status"], "assigned_to": inv["assigned_to"]},
           new_values={"status": status, "assigned_to": assigned_to})

    flash(f"Investigation updated.", "success")
    return redirect(url_for("stale_servers"))


@app.route("/servers/stale/decommission", methods=["POST"])
@login_required
def stale_decommission():
    """Archive a server's licence snapshot, release its licences, and deactivate it."""
    import json as _json

    client_schema = request.form["client_schema"]
    server_id     = int(request.form["server_id"])
    notes         = request.form.get("notes") or None
    username      = session.get("username", "system")

    # Resolve the client name
    client_row = query(
        "SELECT client_name FROM sam_admin.clients WHERE schema_name = %s AND is_active",
        (client_schema,), fetchall=False
    )
    client_name = client_row["client_name"] if client_row else client_schema

    # Fetch server metadata
    server = query(
        f"SELECT * FROM {client_schema}.oracle_servers WHERE server_id = %s",
        (server_id,), fetchall=False
    )
    if not server:
        flash("Server not found.", "danger")
        return redirect(url_for("stale_servers"))

    # Build the licence snapshot from server_csi_map joined to contract info
    licence_rows = query(f"""
        SELECT
            m.map_id,
            m.csi_id,
            cs.csi_number,
            cs.product_family,
            m.product_detail,
            cs.contract_name,
            m.licences_consumed,
            m.effective_date::TEXT AS effective_date,
            m.notes
        FROM {client_schema}.server_csi_map m
        JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id
        WHERE m.server_id = %s
        ORDER BY cs.product_family, m.product_detail
    """, (server_id,))

    licence_snapshot = _json.dumps([dict(r) for r in licence_rows], default=str)

    # Insert into archive
    execute("""
        INSERT INTO sam_admin.decommissioned_servers
            (client_schema, client_name, server_id, hostname, fqdn, ip_address,
             environment, datacenter, os_family, first_seen, last_seen,
             licence_snapshot, decommissioned_by, notes)
        VALUES (%s, %s, %s, %s, %s, %s::TEXT, %s, %s, %s, %s, %s, %s::JSONB, %s, %s)
    """, (
        client_schema, client_name, server_id,
        server["hostname"], server.get("fqdn"),
        str(server["ip_address"]) if server.get("ip_address") else None,
        server.get("environment"), server.get("datacenter"), server.get("os_family"),
        server.get("first_seen"), server.get("last_seen"),
        licence_snapshot, username, notes
    ))

    # Release licences — delete all CSI map entries for this server
    execute(
        f"DELETE FROM {client_schema}.server_csi_map WHERE server_id = %s",
        (server_id,)
    )

    # Deactivate the server (removes it from the servers page)
    execute(
        f"UPDATE {client_schema}.oracle_servers SET is_active = FALSE WHERE server_id = %s",
        (server_id,)
    )

    # Close any open investigation for this server
    execute("""
        UPDATE sam_admin.stale_server_investigations
        SET status      = 'dismissed',
            notes       = COALESCE(notes || E'\n', '') || 'Server decommissioned.',
            resolved_by = %s,
            resolved_at = NOW()
        WHERE client_schema = %s AND server_id = %s
          AND status NOT IN ('resolved', 'dismissed')
    """, (username, client_schema, server_id))

    _audit("server.decommission", entity_type="server", entity_id=server_id,
           entity_name=server["hostname"], client_schema=client_schema,
           new_values={
               "hostname":          server["hostname"],
               "licences_released": len(licence_rows),
               "notes":             notes,
           })

    flash(f"{server['hostname']} has been decommissioned. "
          f"{len(licence_rows)} licence assignment(s) archived and released.", "success")
    return redirect(url_for("stale_servers"))


@app.route("/servers/decommissioned")
@login_required
def decommissioned_servers():
    schema = get_schema()
    if schema == "__all__":
        rows = query("""
            SELECT * FROM sam_admin.decommissioned_servers
            ORDER BY decommissioned_at DESC
        """)
    else:
        rows = query("""
            SELECT * FROM sam_admin.decommissioned_servers
            WHERE client_schema = %s
            ORDER BY decommissioned_at DESC
        """, (schema,))
    return render_template("decommissioned_servers.html", rows=rows)


# ---------------------------------------------------------------------------
# Edit server
# ---------------------------------------------------------------------------
@app.route("/servers/<int:server_id>", methods=["GET", "POST"])
@login_required
def edit_server(server_id):
    schema = get_schema()

    # When viewing all clients, resolve the actual schema for this server
    if schema == "__all__":
        active_clients = query(
            "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
        )
        for c in active_clients:
            s = c["schema_name"]
            row = query(
                f"SELECT 1 FROM {s}.oracle_servers WHERE server_id = %s",
                (server_id,), fetchall=False
            )
            if row:
                schema = s
                break
        if schema == "__all__":
            flash("Server not found.", "danger")
            return redirect(url_for("servers"))

    if request.method == "POST":
        action = request.form.get("action")

        if action == "set_metric":
            metric = request.form.get("metric")
            old_metric_row = query(
                f"SELECT COALESCE(licence_metric_override,'processor_perpetual') AS m "
                f"FROM {schema}.oracle_servers WHERE server_id = %s",
                (server_id,), fetchall=False
            )
            if metric == "processor_perpetual":
                execute(
                    f"UPDATE {schema}.oracle_servers "
                    f"SET licence_metric_override = NULL WHERE server_id = %s",
                    (server_id,)
                )
            else:
                execute(
                    f"UPDATE {schema}.oracle_servers "
                    f"SET licence_metric_override = %s WHERE server_id = %s",
                    (metric, server_id)
                )
            _audit("server.set_metric", entity_type="server", entity_id=server_id,
                   old_values={"metric": (old_metric_row or {}).get("m")},
                   new_values={"metric": metric}, client_schema=schema)
            flash("Licence metric updated.", "success")

        elif action == "assign_csi":
            # client and dba roles go through the approval queue
            if current_role() in ("client", "dba"):
                return redirect(url_for("assignment_propose"), 307)

            csi_id         = request.form.get("csi_id")
            family         = request.form.get("product_family")
            product_detail = request.form.get("product_detail") or None
            consumed_raw   = request.form.get("licences_consumed") or None
            notes          = request.form.get("notes") or None

            # Re-derive the server's actual requirement for this line server-side —
            # never trust the hidden form fields as authorization for the write.
            line_row = query(
                f"SELECT licences_required FROM {schema}.license_position "
                f"WHERE server_id = %s AND product_family = %s "
                f"AND product_detail IS NOT DISTINCT FROM %s",
                (server_id, family, product_detail), fetchall=False
            )
            if not line_row:
                flash("That licence line no longer exists for this server.", "danger")
                return redirect(url_for("edit_server", server_id=server_id))

            # Check whether this CSI covers the assigned product (standard contracts only;
            # ULAs are handled separately via the assign_ula action)
            # Fetch server metric to validate CSI metric compatibility
            _sv_row = query(
                f"SELECT COALESCE(licence_metric_override,'processor_perpetual') AS licence_metric "
                f"FROM {schema}.oracle_servers WHERE server_id = %s",
                (server_id,), fetchall=False
            )
            _sv_metric = (_sv_row or {}).get("licence_metric", "processor_perpetual")

            entitlement_lines = query(
                "SELECT product_name, product_family::TEXT AS product_family, "
                "license_metric::TEXT AS license_metric "
                "FROM shared.license_entitlement_lines WHERE csi_id = %s AND is_active",
                (csi_id,)
            )
            compatible_lines = [
                l for l in entitlement_lines
                if _is_compatible_product(family, product_detail,
                                          l["product_family"], l["product_name"])
            ]
            if not compatible_lines:
                flash("That CSI contract doesn't cover this product/edition — "
                      "pick a contract that matches.", "danger")
                return redirect(url_for("edit_server", server_id=server_id))

            # Reject metric mismatch: NUP server must use NUP CSI lines and vice-versa
            if _sv_metric == "named_user_plus":
                if not any((l.get("license_metric") or "processor") == "named_user_plus"
                           for l in compatible_lines):
                    flash("This server uses Named User Plus licensing — "
                          "only NUP CSI contracts can be assigned.", "danger")
                    return redirect(url_for("edit_server", server_id=server_id))
            else:
                if any((l.get("license_metric") or "processor") == "named_user_plus"
                       for l in compatible_lines):
                    flash("This server uses processor licensing — "
                          "NUP CSI contracts cannot be assigned.", "danger")
                    return redirect(url_for("edit_server", server_id=server_id))

            # Enforce client_locked CSIs — cannot be assigned to a different client's server
            csi_policy = query(
                """SELECT cs.sharing_policy, c.client_code AS owning_client
                   FROM shared.csi_contracts cs
                   LEFT JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id
                   WHERE cs.csi_id = %s""",
                (csi_id,), fetchall=False
            )
            if csi_policy and csi_policy["sharing_policy"] == "client_locked":
                server_client = query(
                    "SELECT client_code FROM sam_admin.clients WHERE schema_name = %s",
                    (schema,), fetchall=False
                )
                server_code = server_client["client_code"] if server_client else None
                if csi_policy["owning_client"] != server_code:
                    flash(
                        f"CSI is locked to {csi_policy['owning_client']} and cannot be "
                        f"assigned to a {server_code} server.",
                        "danger"
                    )
                    return redirect(url_for("edit_server", server_id=server_id))

            consumed = None
            if consumed_raw:
                try:
                    consumed = Decimal(consumed_raw)
                except InvalidOperation:
                    flash("Licences consumed must be a number.", "danger")
                    return redirect(url_for("edit_server", server_id=server_id))

                required = line_row["licences_required"]
                if required is not None:
                    other_consumed = query(
                        f"SELECT COALESCE(SUM(licences_consumed), 0) AS total "
                        f"FROM {schema}.server_csi_map "
                        f"WHERE server_id = %s AND product_family = %s "
                        f"AND product_detail IS NOT DISTINCT FROM %s AND csi_id != %s",
                        (server_id, family, product_detail, csi_id), fetchall=False
                    )["total"]
                    if other_consumed + consumed > required:
                        remaining = required - other_consumed
                        flash(
                            f"This line only needs {required} licence(s) — "
                            f"{other_consumed} already assigned from other contracts, "
                            f"leaving {remaining} available. Reduce the quantity.",
                            "danger"
                        )
                        return redirect(url_for("edit_server", server_id=server_id))

            # Check CSI capacity — total licences in the CSI vs already consumed
            # across ALL client schemas (excluding the current assignment being replaced)
            if consumed is not None:
                csi_capacity = query(
                    "SELECT COALESCE(SUM(quantity), 0) AS total_qty "
                    "FROM shared.license_entitlement_lines "
                    "WHERE csi_id = %s AND is_active",
                    (csi_id,), fetchall=False
                )["total_qty"]

                if csi_capacity > 0:
                    all_schemas = [
                        r["schema_name"] for r in query(
                            "SELECT schema_name FROM sam_admin.clients WHERE is_active"
                        )
                    ]
                    # Sum consumed for this CSI across all schemas, excluding the row
                    # we are about to replace (same server+csi+family+detail)
                    already_consumed = 0
                    for s in all_schemas:
                        try:
                            row = query(
                                f"SELECT COALESCE(SUM(licences_consumed), 0) AS total "
                                f"FROM {s}.server_csi_map "
                                f"WHERE csi_id = %s "
                                f"AND NOT (server_id = %s AND product_family = %s "
                                f"         AND product_detail IS NOT DISTINCT FROM %s)",
                                (csi_id, server_id, family, product_detail),
                                fetchall=False
                            )
                            already_consumed += row["total"] or 0
                        except Exception:
                            pass

                    if already_consumed + consumed > csi_capacity:
                        remaining = max(csi_capacity - already_consumed, 0)
                        flash(
                            f"This CSI only has {csi_capacity} licences in total — "
                            f"{already_consumed} are already assigned to other servers, "
                            f"leaving {remaining} available. You entered {consumed}.",
                            "danger"
                        )
                        return redirect(url_for("edit_server", server_id=server_id))

            # Remove any existing assignment for this server+csi+family+detail first
            execute(f"""
                DELETE FROM {schema}.server_csi_map
                WHERE server_id = %s AND csi_id = %s AND product_family = %s
                  AND (product_detail IS NOT DISTINCT FROM %s)
            """, (server_id, csi_id, family, product_detail))
            execute(f"""
                INSERT INTO {schema}.server_csi_map
                  (server_id, csi_id, product_family, product_detail,
                   licences_consumed, notes, assigned_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (server_id, csi_id, family, product_detail, consumed, notes, ADMIN_USER))
            _audit("csi.assign", entity_type="csi_assignment", entity_id=server_id,
                   entity_name=f"server {server_id} / CSI {csi_id}",
                   new_values={"csi_id": csi_id, "product_family": family,
                               "product_detail": product_detail,
                               "licences_consumed": str(consumed) if consumed else None},
                   client_schema=schema)
            flash("CSI assignment saved.", "success")

        elif action == "remove_csi":
            # client and dba roles go through the approval queue
            if current_role() in ("client", "dba"):
                return redirect(url_for("assignment_propose"), 307)

            map_id = request.form.get("map_id")
            old_csi = query(
                f"SELECT csi_id, product_family::TEXT, product_detail, licences_consumed "
                f"FROM {schema}.server_csi_map WHERE map_id = %s", (map_id,), fetchall=False
            )
            execute(f"DELETE FROM {schema}.server_csi_map WHERE map_id = %s", (map_id,))
            _audit("csi.remove", entity_type="csi_assignment", entity_id=server_id,
                   entity_name=f"server {server_id} / map {map_id}",
                   old_values=dict(old_csi) if old_csi else None,
                   client_schema=schema)
            flash("CSI assignment removed.", "success")

        elif action == "assign_ula":
            csi_id = request.form.get("csi_id")
            # Validate this is a client-locked ULA belonging to this client
            server_client = query(
                "SELECT client_code FROM sam_admin.clients WHERE schema_name = %s",
                (schema,), fetchall=False
            )
            server_code = server_client["client_code"] if server_client else None
            ula_row = query(
                """SELECT cs.is_ula, cs.sharing_policy, c.client_code
                   FROM shared.csi_contracts cs
                   JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id
                   WHERE cs.csi_id = %s""",
                (csi_id,), fetchall=False
            )
            if (not ula_row or not ula_row["is_ula"]
                    or ula_row["sharing_policy"] != "client_locked"
                    or ula_row["client_code"] != server_code):
                flash("Invalid ULA selection.", "danger")
                return redirect(url_for("edit_server", server_id=server_id))

            # --- ULA coverage check ---
            # Verify every product the server requires is listed in the ULA's
            # covered products.  If ula_covered_products has no rows we warn
            # (the ULA has no scope defined) but still allow assignment so
            # legacy contracts without a product list aren't hard-blocked.
            uncovered = _check_ula_coverage(csi_id, server_id, schema)
            if uncovered is None:
                # No covered-products configured — warn but don't block
                flash(
                    "Warning: this ULA has no covered products configured. "
                    "Add products via the contract page to enable coverage validation.",
                    "warning"
                )
            elif uncovered:
                flash(
                    "Cannot assign to this ULA — the following products are required "
                    "on this server but are not listed in the ULA's covered products: "
                    + ", ".join(f'"{p}"' for p in uncovered)
                    + ". Add the missing products to the ULA contract first.",
                    "danger"
                )
                return redirect(url_for("edit_server", server_id=server_id))

            # Get all product families present on this server
            lp_rows = query(
                f"SELECT DISTINCT product_family::TEXT FROM {schema}.license_position "
                f"WHERE server_id = %s",
                (server_id,)
            )
            families = [r["product_family"] for r in lp_rows]
            # Remove ALL existing CSI assignments for this server
            execute(f"DELETE FROM {schema}.server_csi_map WHERE server_id = %s", (server_id,))
            # Insert one ULA row per product family
            for fam in families:
                execute(f"""
                    INSERT INTO {schema}.server_csi_map
                      (server_id, csi_id, product_family, product_detail,
                       licences_consumed, notes, assigned_by)
                    VALUES (%s, %s, %s, NULL, NULL, NULL, %s)
                """, (server_id, csi_id, fam, ADMIN_USER))
            flash("Server assigned to ULA — all individual CSI assignments removed.", "success")

        elif action == "remove_ula":
            execute(
                f"DELETE FROM {schema}.server_csi_map "
                f"WHERE server_id = %s AND csi_id IN "
                f"(SELECT csi_id FROM shared.csi_contracts WHERE is_ula)",
                (server_id,)
            )
            flash("ULA assignment removed.", "success")

        elif action == "save_contacts":
            client_row = query(
                "SELECT client_id FROM sam_admin.clients WHERE schema_name = %s",
                (schema,), fetchall=False
            )
            if client_row:
                client_id = client_row["client_id"]
                execute(
                    "DELETE FROM shared.client_contacts WHERE client_id = %s",
                    (client_id,)
                )
                for i in range(1, 4):
                    name  = (request.form.get(f"contact_name_{i}") or "").strip()
                    email = (request.form.get(f"contact_email_{i}") or "").strip()
                    phone = (request.form.get(f"contact_phone_{i}") or "").strip()
                    if name or email or phone:
                        execute(
                            """INSERT INTO shared.client_contacts
                               (client_id, full_name, email, phone, sort_order)
                               VALUES (%s, %s, %s, %s, %s)""",
                            (client_id, name, email, phone, i)
                        )
                flash("Client contacts saved.", "success")

        elif action == "set_java_exempt":
            java_id = request.form.get("java_id")
            exempt  = request.form.get("exempt") == "1"
            reason  = request.form.get("exempt_reason") or None
            notes   = request.form.get("exempt_notes") or None
            if exempt:
                execute(
                    f"UPDATE {schema}.java_installations "
                    f"SET licence_exempt = TRUE, exempt_reason = %s, "
                    f"    exempt_notes = %s, exempt_set_by = %s, exempt_set_at = NOW() "
                    f"WHERE java_id = %s",
                    (reason, notes, ADMIN_USER, java_id)
                )
                flash("Java installation marked as exempt from licensing.", "success")
            else:
                execute(
                    f"UPDATE {schema}.java_installations "
                    f"SET licence_exempt = FALSE, exempt_reason = NULL, "
                    f"    exempt_notes = NULL, exempt_set_by = NULL, exempt_set_at = NULL "
                    f"WHERE java_id = %s",
                    (java_id,)
                )
                flash("Java licence exemption cleared.", "success")

        elif action == "update_server_details":
            if not can_write_licences():
                abort(403)
            new_env = request.form.get("environment", "").strip() or None
            new_dc  = request.form.get("datacenter", "").strip() or None
            new_ip  = request.form.get("ip_address", "").strip() or None

            old_server = query(
                f"SELECT environment::TEXT, datacenter, ip_address::TEXT "
                f"FROM {schema}.oracle_servers WHERE server_id = %s",
                (server_id,), fetchall=False
            )
            execute(
                f"""UPDATE {schema}.oracle_servers
                    SET environment = COALESCE(%s::environment_type, environment),
                        datacenter  = %s,
                        ip_address  = COALESCE(%s::INET, ip_address)
                    WHERE server_id = %s""",
                (new_env, new_dc, new_ip, server_id)
            )
            _audit("server.update_details", entity_type="server", entity_id=server_id,
                   old_values=dict(old_server) if old_server else None,
                   new_values={"environment": new_env, "datacenter": new_dc,
                               "ip_address": new_ip},
                   client_schema=schema)
            flash("Server details updated.", "success")

        elif action == "deactivate_server":
            if not can_write_licences():
                abort(403)
            execute(
                f"UPDATE {schema}.oracle_servers SET is_active = FALSE WHERE server_id = %s",
                (server_id,)
            )
            _audit("server.deactivate", entity_type="server", entity_id=server_id,
                   new_values={"is_active": False}, client_schema=schema)
            flash("Server removed from inventory.", "warning")
            return redirect(url_for("servers"))

        elif action == "reactivate_server":
            if not can_write_licences():
                abort(403)
            execute(
                f"UPDATE {schema}.oracle_servers SET is_active = TRUE WHERE server_id = %s",
                (server_id,)
            )
            _audit("server.reactivate", entity_type="server", entity_id=server_id,
                   new_values={"is_active": True}, client_schema=schema)
            flash("Server reactivated.", "success")

        return redirect(url_for("edit_server", server_id=server_id))

    # GET
    server = query(
        f"""SELECT s.server_id, s.hostname, s.environment::TEXT, s.datacenter,
                   s.ip_address::TEXT, s.last_seen::DATE,
                   s.is_active,
                   COALESCE(s.licence_metric_override, 'processor_perpetual') AS licence_metric,
                   s.licence_metric_override IS NOT NULL AS metric_overridden
            FROM {schema}.oracle_servers s WHERE s.server_id = %s""",
        (server_id,), fetchall=False
    )
    if not server:
        flash("Server not found.", "danger")
        return redirect(url_for("servers"))

    instances = query(
        f"SELECT oracle_sid, edition, db_version FROM {schema}.oracle_instances "
        f"WHERE server_id = %s AND is_active ORDER BY oracle_sid",
        (server_id,)
    )

    assignments = query(f"""
        SELECT m.map_id, m.csi_id, m.product_family, m.product_detail,
               m.licences_consumed, m.notes, m.assigned_by, m.effective_date,
               cs.csi_number, cs.contract_name, cs.support_expiry,
               cs.status AS contract_status, cs.is_ula
        FROM {schema}.server_csi_map m
        JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id
        WHERE m.server_id = %s
        ORDER BY m.product_family, m.product_detail, cs.csi_number
    """, (server_id,))

    # Fetch the client code for this server so we can filter out locked CSIs
    server_client_row = query(
        "SELECT client_code FROM sam_admin.clients WHERE schema_name = %s",
        (schema,), fetchall=False
    )
    server_client_code = server_client_row["client_code"] if server_client_row else None

    entitlement_lines = query("""
        SELECT l.csi_id, l.product_name, l.product_family::TEXT AS product_family,
               l.license_metric::TEXT AS license_metric,
               l.quantity, l.unit_price, cs.csi_number, cs.contract_name, cs.support_expiry,
               cs.sharing_policy, c.client_code AS owning_client
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        LEFT JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id
        WHERE cs.status = 'active'
          AND (cs.sharing_policy != 'client_locked' OR c.client_code = %s)
        ORDER BY cs.csi_number
    """, (server_client_code,))

    # Detect NUP columns (added by migration 05)
    _nup_col_check = query(
        """SELECT COUNT(*) AS n FROM information_schema.columns
           WHERE table_schema = %s AND table_name = 'license_position'
             AND column_name IN ('licence_metric','nup_minimum','nup_active_users')""",
        (schema,), fetchall=False
    )
    _lp_has_nup = (_nup_col_check or {}).get("n", 0) == 3

    if _lp_has_nup:
        licence_position = query(
            f"SELECT *, licence_metric, nup_minimum, nup_active_users "
            f"FROM {schema}.license_position WHERE server_id = %s", (server_id,)
        )
    else:
        licence_position = query(
            f"SELECT * FROM {schema}.license_position WHERE server_id = %s", (server_id,)
        )
        for row in licence_position:
            row["licence_metric"]    = "processor_perpetual"
            row["nup_minimum"]       = None
            row["nup_active_users"]  = None

    def _licence_sort_key(row):
        d = (row.get("product_detail") or "").lower()
        if "enterprise" in d or "standard" in d:
            return (0, d)
        if "diagnostic" in d:
            return (1, d)
        if "tuning" in d:
            return (2, d)
        return (3, d)

    licence_position = sorted(licence_position, key=_licence_sort_key)

    # Per-line: which active CSIs actually cover this product/edition, and how
    # much licence headroom is left once existing assignments are subtracted.
    consumed_by_line = {}
    for a in assignments:
        if a["licences_consumed"] is not None:
            key = (a["product_family"], a["product_detail"])
            consumed_by_line[key] = consumed_by_line.get(key, 0) + a["licences_consumed"]

    # Quantity per (csi_id, product_name) from entitlement lines
    line_qty = {}
    for l in entitlement_lines:
        line_qty[(l["csi_id"], l["product_name"])] = l.get("quantity") or 0

    # Consumed per (csi_id, product_detail) across ALL active client schemas
    active_schemas = [
        r["schema_name"] for r in query(
            "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
        )
    ]
    consumed_by_csi_detail = {}   # (csi_id, product_detail) -> total consumed
    if active_schemas:
        union_sql = " UNION ALL ".join(
            f"SELECT csi_id, product_family::TEXT, COALESCE(product_detail,'') AS product_detail, "
            f"COALESCE(SUM(licences_consumed), 0) AS consumed "
            f"FROM {s}.server_csi_map GROUP BY csi_id, product_family, product_detail"
            for s in active_schemas
        )
        for r in query(
            f"SELECT csi_id, product_family, product_detail, SUM(consumed) AS total "
            f"FROM ({union_sql}) t GROUP BY csi_id, product_family, product_detail"
        ):
            key = (r["csi_id"], r["product_detail"])
            consumed_by_csi_detail[key] = consumed_by_csi_detail.get(key, 0) + r["total"]

    # Index consumed entries by csi_id for fast lookup
    consumed_entries_by_csi = {}
    for (csi_id, detail), amt in consumed_by_csi_detail.items():
        consumed_entries_by_csi.setdefault(csi_id, []).append((detail, amt))

    # Client-locked ULAs owned by this client, grouped by applicable licence line
    ula_contracts = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.ula_expiry
        FROM shared.csi_contracts cs
        JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id
        WHERE cs.is_ula AND cs.status = 'active'
          AND cs.sharing_policy = 'client_locked'
          AND c.client_code = %s
        ORDER BY cs.csi_number
    """, (server_client_code,))
    _ula_covered = {}
    for u in ula_contracts:
        rows = query(
            "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s",
            (u["csi_id"],)
        )
        _ula_covered[u["csi_id"]] = [r["product_name"] for r in rows]

    _server_metric = server.get("licence_metric", "processor_perpetual")
    compatible_csis_by_line = {}
    ula_by_line = {}   # line_key -> list of applicable ULA dicts
    for row in licence_position:
        key = f"{row['product_family']}|{row['product_detail'] or ''}"
        matches = {}
        for l in entitlement_lines:
            # When the server uses NUP, only show NUP-metric lines; otherwise exclude NUP lines
            line_metric = l.get("license_metric") or "processor"
            if _server_metric == "named_user_plus" and line_metric != "named_user_plus":
                continue
            if _server_metric != "named_user_plus" and line_metric == "named_user_plus":
                continue
            if _is_compatible_product(row["product_family"], row["product_detail"],
                                       l["product_family"], l["product_name"]):
                if l["csi_id"] not in matches:
                    qty = line_qty.get((l["csi_id"], l["product_name"]), 0)
                    consumed = sum(
                        amt for det, amt in consumed_entries_by_csi.get(l["csi_id"], [])
                        if _is_compatible_product(
                            l["product_family"], det or None,
                            l["product_family"], l["product_name"])
                    )
                    matches[l["csi_id"]] = dict(l,
                        csi_total_qty=qty,
                        csi_consumed=consumed,
                        csi_available=max(qty - consumed, 0)
                    )
        compatible_csis_by_line[key] = list(matches.values())

        # Applicable ULAs for this licence line (shown separately, not in the CSI dropdown)
        applicable_ulas = []
        for u in ula_contracts:
            covered = _ula_covered[u["csi_id"]]
            if not covered or any(
                _ula_product_matches_family(p, row["product_family"]) for p in covered
            ):
                applicable_ulas.append(u)
        ula_by_line[key] = applicable_ulas

        already = consumed_by_line.get((row["product_family"], row["product_detail"]), 0)
        row["already_consumed"] = already
        row["remaining_capacity"] = (
            row["licences_required"] - already if row["licences_required"] is not None else None
        )

    # Deduplicated ULAs applicable to any line on this server (for the single top-level toggle)
    _seen_ula_ids = set()
    all_server_ulas = []
    for _ulas in ula_by_line.values():
        for _u in _ulas:
            if _u["csi_id"] not in _seen_ula_ids:
                _seen_ula_ids.add(_u["csi_id"])
                all_server_ulas.append(_u)

    # WebLogic domains discovered on this server
    try:
        wls_domains = query(
            f"""SELECT domain_id, domain_name, domain_home, wls_version, wls_edition,
                       admin_server_host, admin_server_port, last_seen::DATE AS last_seen
                FROM {schema}.wls_domains
                WHERE server_id = %s AND is_active
                ORDER BY domain_name""",
            (server_id,)
        )
        for d in wls_domains:
            d = dict(d)
            d["managed_servers"] = query(
                f"""SELECT managed_server_name, listen_port, ssl_port,
                           cluster_name, machine_name, state
                    FROM {schema}.wls_managed_servers
                    WHERE domain_id = %s ORDER BY managed_server_name""",
                (d["domain_id"],)
            )
            d["installed_products"] = query(
                f"""SELECT DISTINCT product_name, product_version
                    FROM {schema}.wls_installed_products
                    WHERE domain_id = %s ORDER BY product_name""",
                (d["domain_id"],)
            )
    except Exception:
        wls_domains = []

    java_installations = query(
        f"""SELECT java_id, java_home, java_vendor, java_version,
                   java_major_version, java_edition, is_oracle_jdk,
                   requires_licence, licence_metric,
                   licence_exempt, exempt_reason, exempt_notes,
                   exempt_set_by, exempt_set_at::DATE AS exempt_set_at
            FROM {schema}.java_installations
            WHERE server_id = %s
            ORDER BY java_major_version DESC, java_home""",
        (server_id,)
    )

    # SE2 violations for this server
    try:
        se2_violations = query(
            f"SELECT * FROM {schema}.se2_violations WHERE server_id = %s", (server_id,)
        )
    except Exception:
        se2_violations = []

    # CPU validation
    try:
        cpu_validation = query(
            f"""SELECT v.*, p.cpu_architecture, p.virt_type::TEXT, p.is_vmware, p.is_exadata
                FROM {schema}.cpu_validation_report v
                LEFT JOIN LATERAL (
                    SELECT cpu_architecture, virt_type, is_vmware, is_exadata
                    FROM {schema}.oracle_processors
                    WHERE server_id = v.server_id
                    ORDER BY recorded_at DESC LIMIT 1
                ) p ON TRUE
                WHERE v.server_id = %s""",
            (server_id,), fetchall=False
        )
    except Exception:
        cpu_validation = None

    # Client contacts
    client_row = query(
        "SELECT client_id FROM sam_admin.clients WHERE schema_name = %s",
        (schema,), fetchall=False
    )
    client_contacts = []
    if client_row:
        try:
            client_contacts = query(
                "SELECT full_name, email, phone FROM shared.client_contacts "
                "WHERE client_id = %s ORDER BY sort_order",
                (client_row["client_id"],)
            )
        except Exception:
            client_contacts = []

    # Oracle options (from v$option / Ansible discovery)
    try:
        oracle_options = query(
            f"""SELECT o.option_name, o.option_version, o.status,
                       i.oracle_sid
                FROM {schema}.oracle_options o
                JOIN {schema}.oracle_instances i ON i.instance_id = o.instance_id
                WHERE i.server_id = %s AND o.status = 'TRUE'
                ORDER BY i.oracle_sid, o.option_name""",
            (server_id,)
        )
    except Exception:
        oracle_options = []

    # Detected licensed products derived from feature usage.
    # Feature names matched against MOS Doc ID 1317265.1 MAP CTE, filtered to
    # names that oracle_discovery.sql actually stores (its cursor excludes several).
    try:
        detected_products = query(
            f"""SELECT product, MAX(last_usage_date) AS last_usage_date,
                       SUM(detected_usages) AS detected_usages
                FROM (
                  SELECT
                    CASE
                      WHEN f.feature_name = ANY(ARRAY[
                             'SQL Profile',
                             'SQL Monitoring and Tuning pages',
                             'SQL Tuning Advisor',
                             'SQL Access Advisor',
                             'Tuning Pack'
                           ])
                        THEN 'Tuning Pack'
                      WHEN f.feature_name = ANY(ARRAY[
                             'ADDM',
                             'AWR Baseline',
                             'AWR Baseline Template',
                             'AWR Report',
                             'Automatic Workload Repository',
                             'Baseline Adaptive Thresholds',
                             'Baseline Static Computations',
                             'Diagnostic Pack',
                             'Diagnostics Pack',
                             'Active Session History'
                           ])
                        THEN 'Diagnostics Pack'
                      WHEN f.feature_name = ANY(ARRAY[
                             'Transparent Data Encryption',
                             'Encrypted Tablespaces',
                             'Data Redaction',
                             'SecureFile Encryption (user)',
                             'Backup Encryption',
                             'Advanced Security'
                           ])
                        THEN 'Advanced Security (ASO)'
                      WHEN f.feature_name ILIKE '%Database Vault%'
                        THEN 'Database Vault'
                      WHEN f.feature_name = ANY(ARRAY[
                             'Partitioning',
                             'Interval Partitioning'
                           ])
                           OR f.feature_name ILIKE 'Partitioning%'
                        THEN 'Partitioning'
                      WHEN f.feature_name ILIKE '%Real Application Clusters%'
                        OR f.feature_name = 'RAC'
                        THEN 'Real Application Clusters'
                      WHEN f.feature_name ILIKE '%Active Data Guard%'
                        OR f.feature_name = 'Active Data Guard - Real-Time Query on Physical Standby'
                        THEN 'Active Data Guard'
                      WHEN f.feature_name = 'Oracle Multitenant'
                        OR (f.feature_name ILIKE '%Multitenant%'
                            AND f.feature_name NOT ILIKE '%Non-CDB%')
                        THEN 'Multitenant'
                      WHEN f.feature_name ILIKE '%Label Security%'
                        THEN 'Label Security'
                      ELSE NULL
                    END AS product,
                    f.last_usage_date,
                    f.detected_usages
                  FROM {schema}.oracle_feature_usage f
                  JOIN {schema}.oracle_instances i ON i.instance_id = f.instance_id
                  WHERE i.server_id = %s AND f.detected_usages > 0
                ) mapped
                WHERE product IS NOT NULL
                GROUP BY product
                ORDER BY product""",
            (server_id,)
        )
    except Exception:
        detected_products = []

    return render_template("edit_server.html",
                           server=server,
                           instances=instances,
                           assignments=assignments,
                           compatible_csis_by_line=compatible_csis_by_line,
                           ula_by_line=ula_by_line,
                           all_server_ulas=all_server_ulas,
                           licence_position=licence_position,
                           wls_domains=wls_domains,
                           java_installations=java_installations,
                           se2_violations=se2_violations,
                           cpu_validation=cpu_validation,
                           client_contacts=client_contacts,
                           oracle_options=oracle_options,
                           detected_products=detected_products)


# ---------------------------------------------------------------------------
# Server change history
# ---------------------------------------------------------------------------
@app.route("/servers/<int:server_id>/history")
@login_required
def server_history(server_id):
    schema = get_schema()
    server = query(
        f"SELECT server_id, hostname FROM {schema}.oracle_servers WHERE server_id = %s",
        (server_id,), fetchall=False
    )
    if not server:
        flash("Server not found.", "danger")
        return redirect(url_for("servers"))

    history = query(f"""
        SELECT change_id, change_category, object_name, object_type,
               severity, old_value, new_value, change_description,
               detected_at::TIMESTAMPTZ, acknowledged, acknowledged_by,
               acknowledged_at::DATE
        FROM {schema}.discovery_changelog
        WHERE server_id = %s
        ORDER BY detected_at DESC
        LIMIT 500
    """, (server_id,))

    return render_template("history.html", server=server, history=history)


@app.route("/servers/<int:server_id>/history/<int:change_id>/acknowledge",
           methods=["POST"])
@login_required
def acknowledge_change(server_id, change_id):
    schema = get_schema()
    execute(
        f"UPDATE {schema}.discovery_changelog "
        f"SET acknowledged = TRUE, acknowledged_by = %s, acknowledged_at = NOW() "
        f"WHERE change_id = %s",
        (ADMIN_USER, change_id)
    )
    flash("Change acknowledged.", "success")
    return redirect(url_for("server_history", server_id=server_id))


# ---------------------------------------------------------------------------
# CSI contracts browser
# ---------------------------------------------------------------------------
@app.route("/licence-summary")
@login_required
def licence_summary():
    # Per-client totals from client_locked CSIs
    client_rows = query("""
        SELECT c.client_code, c.client_name,
               l.product_name, l.product_family::TEXT AS product_family,
               COALESCE(SUM(l.quantity), 0) AS total_qty
        FROM sam_admin.clients c
        JOIN shared.csi_contracts cs ON cs.owning_client_id = c.client_id
            AND cs.status = 'active'
            AND cs.sharing_policy = 'client_locked'
        JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        GROUP BY c.client_code, c.client_name, l.product_name, l.product_family
        ORDER BY c.client_code, l.product_family, l.product_name
    """)

    # Shared/pooled CSI totals (sharing_policy != 'client_locked')
    shared_rows = query("""
        SELECT l.product_name, l.product_family::TEXT AS product_family,
               COALESCE(SUM(l.quantity), 0) AS total_qty
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.status = 'active'
          AND cs.sharing_policy != 'client_locked'
        GROUP BY l.product_name, l.product_family
        ORDER BY l.product_family, l.product_name
    """)

    def _line_sort(name):
        n = (name or "").lower()
        if "enterprise" in n or "standard" in n:
            return (0, n)
        if "diagnostic" in n:
            return (1, n)
        if "tuning" in n:
            return (2, n)
        return (3, n)

    # Group client rows by client
    clients_map = {}
    for r in client_rows:
        key = r["client_code"]
        if key not in clients_map:
            clients_map[key] = {"client_code": r["client_code"],
                                "client_name": r["client_name"],
                                "lines": []}
        clients_map[key]["lines"].append(r)
    for v in clients_map.values():
        v["lines"].sort(key=lambda r: _line_sort(r["product_name"]))

    shared_lines = sorted(shared_rows, key=lambda r: _line_sort(r["product_name"]))

    return render_template("licence_summary.html",
                           clients=list(clients_map.values()),
                           shared_lines=shared_lines)


@app.route("/licence-summary/server-costs")
@login_required
def licence_summary_server_costs():
    """Per-server licence cost breakdown across all active client schemas."""
    schema = get_schema()
    if schema and schema != "__all__":
        active_clients = query(
            "SELECT client_id, client_code, client_name, schema_name "
            "FROM sam_admin.clients WHERE is_active AND schema_name = %s",
            (schema,)
        )
    else:
        active_clients = query(
            "SELECT client_id, client_code, client_name, schema_name "
            "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
        )

    # Build a lookup: (csi_id, product_name) -> line data, for fallback when
    # server_csi_map.line_id is NULL but product_detail is set.
    line_lookup_rows = query("""
        SELECT csi_id, product_name, unit_price, quantity, annual_support_cost
        FROM shared.license_entitlement_lines
        WHERE is_active
    """)
    # key: (csi_id, lowercase product_name)
    line_lookup = {}
    for lr in line_lookup_rows:
        key = (lr["csi_id"], (lr["product_name"] or "").lower())
        line_lookup[key] = lr

    servers = []
    for c in active_clients:
        s = c["schema_name"]
        try:
            rows = query(f"""
                SELECT
                    sv.server_id,
                    sv.hostname,
                    sv.environment::TEXT AS environment,
                    scm.map_id,
                    scm.csi_id,
                    scm.product_family::TEXT AS product_family,
                    COALESCE(scm.product_detail, '') AS product_detail,
                    COALESCE(scm.licences_consumed, 0) AS licences_consumed,
                    cs.csi_number,
                    cs.contract_name,
                    cs.sharing_policy::TEXT AS sharing_policy,
                    l.product_name,
                    l.unit_price,
                    l.quantity       AS line_quantity,
                    l.annual_support_cost
                FROM {s}.server_csi_map scm
                JOIN {s}.oracle_servers sv ON sv.server_id = scm.server_id
                JOIN shared.csi_contracts cs ON cs.csi_id = scm.csi_id
                LEFT JOIN shared.license_entitlement_lines l ON l.line_id = scm.line_id
                ORDER BY sv.hostname, scm.product_family, l.product_name
            """)
        except Exception as e:
            app.logger.error("server_costs query failed for schema %s: %s", s, e)
            rows = []

        # Group by server
        server_map = {}
        for r in rows:
            key = (c["schema_name"], r["server_id"])
            if key not in server_map:
                server_map[key] = {
                    "hostname":    r["hostname"],
                    "environment": r["environment"],
                    "client_name": c["client_name"] or c["client_code"],
                    "client_code": c["client_code"],
                    "assignments": [],
                    "total_licence_cost": 0.0,
                    "total_support_cost": 0.0,
                }
            sv = server_map[key]
            # If the direct line join returned nothing, try the lookup by product_detail
            fallback = None
            if not r["product_name"] and r["product_detail"]:
                fallback = line_lookup.get((r["csi_id"], r["product_detail"].lower()))
            consumed  = float(r["licences_consumed"])
            if fallback:
                qty      = float(fallback["quantity"] or 0)
                unit_p   = float(fallback["unit_price"] or 0)
                ann_supp = float(fallback["annual_support_cost"] or 0)
                pname    = fallback["product_name"]
            else:
                qty      = float(r["line_quantity"] or 0)
                unit_p   = float(r["unit_price"] or 0)
                ann_supp = float(r["annual_support_cost"] or 0)
                pname    = r["product_name"] or r["product_detail"] or r["product_family"]
            line_lic  = unit_p * consumed
            line_supp = (ann_supp / qty * consumed) if qty else 0
            sv["assignments"].append({
                "product_name":      pname,
                "product_family":    r["product_family"],
                "csi_id":            r["csi_id"],
                "csi_number":        r["csi_number"],
                "contract_name":     r["contract_name"],
                "sharing_policy":    r["sharing_policy"],
                "licences_consumed": consumed,
                "unit_price":        unit_p,
                "line_licence_cost": line_lic,
                "line_support_cost": line_supp,
            })
            sv["total_licence_cost"] += line_lic
            sv["total_support_cost"] += line_supp

        for sv in server_map.values():
            sv["total_cost"] = sv["total_licence_cost"] + sv["total_support_cost"]
            servers.append(sv)

    servers.sort(key=lambda s: (s["client_name"], s["hostname"]))

    return render_template("licence_summary_server_costs.html", servers=servers)


def _build_licence_detail(entitlement_rows):
    """Return list of dicts with total_qty, assigned_qty, unassigned_qty, servers per product."""
    active_schemas = [
        r["schema_name"] for r in query(
            "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
        )
    ]

    # Per-server assignments: (csi_id, product_detail) -> [{hostname, licences_consumed}]
    server_assignments = {}   # (csi_id, product_detail) -> list of {hostname, consumed}
    if active_schemas:
        union_sql = " UNION ALL ".join(
            f"SELECT m.csi_id, COALESCE(m.product_detail,'') AS product_detail, "
            f"s.hostname, COALESCE(m.licences_consumed, 0) AS consumed "
            f"FROM {s}.server_csi_map m "
            f"JOIN {s}.oracle_servers s ON s.server_id = m.server_id"
            for s in active_schemas
        )
        for r in query(f"SELECT csi_id, product_detail, hostname, SUM(consumed) AS consumed "
                       f"FROM ({union_sql}) t GROUP BY csi_id, product_detail, hostname "
                       f"ORDER BY hostname"):
            key = (r["csi_id"], r["product_detail"])
            server_assignments.setdefault(key, []).append(
                {"hostname": r["hostname"], "consumed": int(r["consumed"])}
            )

    # Aggregated consumed by csi_id for totals
    consumed_entries_by_csi = {}
    for (csi_id, detail), entries in server_assignments.items():
        total = sum(e["consumed"] for e in entries)
        consumed_entries_by_csi.setdefault(csi_id, []).append((detail, total))

    def _line_sort(name):
        n = (name or "").lower()
        if "enterprise" in n or "standard" in n:
            return (0, n)
        if "diagnostic" in n:
            return (1, n)
        if "tuning" in n:
            return (2, n)
        return (3, n)

    # Roll up by product_name across all CSIs in the set
    product_totals = {}   # product_name -> {total, assigned, servers, family}
    for r in entitlement_rows:
        pname = r["product_name"]
        if pname not in product_totals:
            product_totals[pname] = {"product_name": pname,
                                     "product_family": r["product_family"],
                                     "total_qty": 0, "assigned_qty": 0,
                                     "servers": {}}   # hostname -> consumed
        product_totals[pname]["total_qty"] += int(r["quantity"] or 0)

        # Match server assignments for this CSI line
        for (csi_id, detail), entries in server_assignments.items():
            if csi_id != r["csi_id"]:
                continue
            if _is_compatible_product(r["product_family"], detail or None,
                                      r["product_family"], r["product_name"]):
                for e in entries:
                    if not e["consumed"]:
                        continue
                    h = e["hostname"]
                    product_totals[pname]["servers"][h] = (
                        product_totals[pname]["servers"].get(h, 0) + e["consumed"]
                    )

        consumed = sum(
            amt for det, amt in consumed_entries_by_csi.get(r["csi_id"], [])
            if _is_compatible_product(r["product_family"], det or None,
                                      r["product_family"], r["product_name"])
        )
        product_totals[pname]["assigned_qty"] += consumed

    # Fetch license_position data: required + NUP info per (hostname, product_detail)
    lp_data = {}  # (hostname, product_detail) -> {required, nup_minimum, nup_users, licence_metric}
    if active_schemas:
        # Check if the NUP columns exist (added by migration 05)
        first_schema = active_schemas[0]
        _nup_check = query(
            """
            SELECT COUNT(*) AS n FROM information_schema.columns
            WHERE table_schema = %s AND table_name = 'license_position'
              AND column_name IN ('licence_metric','nup_minimum','nup_active_users')
            """,
            (first_schema,), fetchall=False
        )
        _has_nup = (_nup_check or {}).get("n", 0) == 3

        if _has_nup:
            _nup_select = "nup_minimum, nup_active_users, licence_metric"
        else:
            _nup_select = "NULL::NUMERIC AS nup_minimum, NULL::NUMERIC AS nup_active_users, 'processor_perpetual'::TEXT AS licence_metric"

        lp_union = " UNION ALL ".join(
            f"SELECT hostname, product_detail, licences_required, {_nup_select} "
            f"FROM {s}.license_position"
            for s in active_schemas
        )
        for r in query(f"SELECT * FROM ({lp_union}) t"):
            key = (r["hostname"], (r["product_detail"] or "").lower())
            lp_data[key] = {
                "required":        float(r["licences_required"] or 0),
                "nup_minimum":     float(r["nup_minimum"]) if r["nup_minimum"] is not None else None,
                "nup_users":       int(r["nup_active_users"]) if r["nup_active_users"] is not None else None,
                "licence_metric":  r["licence_metric"] or "processor_perpetual",
            }

    lines = sorted(product_totals.values(), key=lambda x: _line_sort(x["product_name"]))
    for ln in lines:
        ln["unassigned_qty"] = max(ln["total_qty"] - ln["assigned_qty"], 0)
        pname_lower = (ln["product_name"] or "").lower()
        # Convert servers dict to sorted list, enriched with license_position data
        server_list = []
        any_nup = False
        for h, c in ln["servers"].items():
            lp = lp_data.get((h, pname_lower)) or lp_data.get((h, ""))
            is_nup = lp and lp["licence_metric"] == "named_user_plus"
            if is_nup:
                any_nup = True
            server_list.append({
                "hostname":       h,
                "consumed":       c,
                "required":       lp["required"] if lp else None,
                "nup_minimum":    lp["nup_minimum"] if lp else None,
                "nup_users":      lp["nup_users"] if lp else None,
                "licence_metric": lp["licence_metric"] if lp else None,
                "is_nup":         is_nup,
            })
        ln["servers"] = sorted(server_list, key=lambda x: x["hostname"])
        ln["any_nup"] = any_nup
    return lines


@app.route("/licence-summary/client/<client_code>")
@login_required
def licence_summary_client(client_code):
    client = query(
        "SELECT client_id, client_code, client_name FROM sam_admin.clients WHERE client_code = %s",
        (client_code,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("licence_summary"))

    entitlement_rows = query("""
        SELECT l.csi_id, l.product_name, l.product_family::TEXT AS product_family,
               COALESCE(l.quantity, 0) AS quantity
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.status = 'active'
          AND cs.sharing_policy = 'client_locked'
          AND cs.owning_client_id = %s
    """, (client["client_id"],))

    lines = _build_licence_detail(entitlement_rows)
    return render_template("licence_summary_client.html",
                           client=client,
                           lines=lines)


@app.route("/licence-summary/shared")
@login_required
def licence_summary_shared():
    # Per-client usage breakdown
    shareable_csis = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name,
               lel.product_name, lel.product_family::TEXT AS product_family,
               lel.quantity AS total_qty, lel.unit_price
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines lel ON lel.csi_id = cs.csi_id AND lel.is_active
        WHERE cs.sharing_policy = 'shareable' AND cs.status = 'active'
        ORDER BY lel.product_name, cs.contract_name
    """)

    shareable_csi_ids = set(r["csi_id"] for r in shareable_csis)
    entitlement_map = {}
    for r in shareable_csis:
        key = (r["csi_id"], r["product_name"])
        entitlement_map[key] = {
            "csi_id":        r["csi_id"],
            "csi_number":    r["csi_number"],
            "contract_name": r["contract_name"],
            "product_name":  r["product_name"],
            "product_family":r["product_family"],
            "total_qty":     int(r["total_qty"] or 0),
            "unit_price":    float(r["unit_price"] or 0),
            "clients":       [],
            "used_qty":      0,
        }

    clients_list = query(
        "SELECT client_id, client_code, client_name, schema_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    for c in clients_list:
        s = c["schema_name"]
        try:
            rows = query(f"""
                SELECT csi_id, COALESCE(SUM(licences_consumed), 0)::int AS licences_used
                FROM {s}.server_csi_map
                WHERE csi_id = ANY(%s)
                GROUP BY csi_id
            """, (list(shareable_csi_ids),))
        except Exception:
            rows = []

        for r in rows:
            licences = int(r["licences_used"] or 0)
            if licences <= 0:
                continue
            csi_id = r["csi_id"]
            # Match to the entitlement key for this CSI (take first matching key)
            key = next((k for k in entitlement_map if k[0] == csi_id), None)
            if key is None:
                continue
            entitlement_map[key]["clients"].append({
                "client_name": c["client_name"] or c["client_code"],
                "client_code": c["client_code"],
                "licences":    licences,
            })
            entitlement_map[key]["used_qty"] += licences

    pool_products = sorted(entitlement_map.values(), key=lambda x: x["product_name"])
    for p in pool_products:
        p["clients"]         = sorted(p["clients"], key=lambda x: -x["licences"])
        p["available_qty"]   = max(p["total_qty"] - p["used_qty"], 0)
        p["utilisation_pct"] = round(p["used_qty"] / p["total_qty"] * 100, 1) if p["total_qty"] > 0 else 0

    return render_template("licence_summary_shared.html",
                           pool_products=pool_products)


import math as _math

_FINOPS_PALETTE = ["#2a78d6","#1baf7a","#eda100","#008300",
                   "#4a3aa7","#e34948","#e87ba4","#eb6834"]

def _finops_line_sort(name):
    n = (name or "").lower()
    if "enterprise" in n or "standard" in n:
        return (0, n)
    if "diagnostic" in n:
        return (1, n)
    if "tuning" in n:
        return (2, n)
    return (3, n)

def _pie_slices(items, cx=100, cy=100, r=80, gap_deg=1.5):
    total = sum(i["value"] for i in items)
    if total == 0:
        return []
    slices, angle = [], 0.0
    for i, item in enumerate(items):
        sweep = item["value"] / total * 360
        a0 = _math.radians(angle + gap_deg / 2)
        a1 = _math.radians(angle + sweep - gap_deg / 2)
        x1, y1 = cx + r * _math.cos(a0), cy + r * _math.sin(a0)
        x2, y2 = cx + r * _math.cos(a1), cy + r * _math.sin(a1)
        slices.append({
            "path": (f"M {cx} {cy} L {x1:.2f} {y1:.2f} "
                     f"A {r} {r} 0 {1 if sweep >= 180 else 0} 1 {x2:.2f} {y2:.2f} Z"),
            "colour": _FINOPS_PALETTE[i % len(_FINOPS_PALETTE)],
            "label": item["label"],
            "value": item["value"],
            "pct": round(item["value"] / total * 100, 1),
        })
        angle += sweep
    return slices

def _build_client_finops(client_id):
    rows = query("""
        SELECT l.csi_id,
               l.product_name,
               l.product_family::TEXT AS product_family,
               cs.sharing_policy::TEXT AS sharing_policy,
               COALESCE(SUM(l.quantity), 0)            AS qty,
               COALESCE(SUM(l.total_price), 0)         AS licence_cost,
               COALESCE(SUM(l.annual_support_cost), 0) AS support_cost,
               CASE WHEN SUM(l.quantity) > 0
                    THEN SUM(l.total_price) / SUM(l.quantity)
                    ELSE 0 END                          AS unit_cost
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l
             ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.status = 'active'
          AND (
            cs.owning_client_id = %s
            OR cs.csi_id IN (
              SELECT csi_id FROM shared.csi_client_map WHERE client_id = %s
            )
          )
        GROUP BY l.csi_id, l.product_name, l.product_family, cs.sharing_policy
    """, (client_id, client_id))
    if not rows:
        return None

    # Consumed quantities per (csi_id, product_detail) across all schemas
    active_schemas = [r["schema_name"] for r in query(
        "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
    )]
    consumed_entries_by_csi = {}
    if active_schemas:
        union_sql = " UNION ALL ".join(
            f"SELECT csi_id, COALESCE(product_detail,'') AS product_detail, "
            f"COALESCE(SUM(licences_consumed),0) AS consumed "
            f"FROM {s}.server_csi_map GROUP BY csi_id, product_detail"
            for s in active_schemas
        )
        for r in query(f"SELECT csi_id, product_detail, SUM(consumed) AS total "
                       f"FROM ({union_sql}) t GROUP BY csi_id, product_detail"):
            consumed_entries_by_csi.setdefault(r["csi_id"], []).append(
                (r["product_detail"], float(r["total"]))
            )

    # Roll up by (product_name, sharing_policy) so locked vs shared stay separate
    product_map = {}
    for r in rows:
        pname  = r["product_name"]
        source = r["sharing_policy"]  # 'client_locked' or 'shareable'
        key    = (pname, source)
        if key not in product_map:
            product_map[key] = {
                "product_name": pname,
                "product_family": r["product_family"],
                "source": source,
                "qty": 0, "licence_cost": 0.0,
                "support_cost": 0.0, "unit_cost": float(r["unit_cost"] or 0),
                "assigned_qty": 0,
            }
        product_map[key]["qty"]          += int(r["qty"] or 0)
        product_map[key]["licence_cost"] += float(r["licence_cost"] or 0)
        product_map[key]["support_cost"] += float(r["support_cost"] or 0)

        consumed = sum(
            amt for det, amt in consumed_entries_by_csi.get(r["csi_id"], [])
            if _is_compatible_product(r["product_family"], det or None,
                                      r["product_family"], r["product_name"])
        )
        product_map[key]["assigned_qty"] += consumed

    lines = sorted(product_map.values(), key=lambda r: (r["source"] != "client_locked", _finops_line_sort(r["product_name"])))
    for i, ln in enumerate(lines):
        ln["colour"] = _FINOPS_PALETTE[i % len(_FINOPS_PALETTE)]
        qty   = ln["qty"]
        unit  = ln["unit_cost"]
        assigned = min(int(ln["assigned_qty"]), qty)
        unassigned = max(qty - assigned, 0)
        ln["assigned_qty"]      = assigned
        ln["unassigned_qty"]    = unassigned
        ln["assigned_cost"]     = round(assigned   * unit, 2)
        ln["unassigned_cost"]   = round(unassigned * unit, 2)

    # Support cost only applies to client-locked lines; shared pool contributes its in-use assigned cost
    total_support        = sum(ln["support_cost"]   for ln in lines if ln["source"] == "client_locked")
    total_shared_inuse   = sum(ln["assigned_cost"]  for ln in lines if ln["source"] == "shareable")
    total_assigned       = sum(ln["assigned_cost"]  for ln in lines)
    total_unassigned     = sum(ln["unassigned_cost"] for ln in lines if ln["source"] == "client_locked")

    # -----------------------------------------------------------------------
    # FY shared pool cost: sum actual monthly snapshot costs (Apr 1 – Mar 31)
    # Only count months that have a snapshot — so 6 months used = 6 × monthly cost
    # -----------------------------------------------------------------------
    today    = date.today()
    fy_year  = today.year if today.month >= 4 else today.year - 1
    fy_start = date(fy_year, 4, 1)
    fy_end   = date(fy_year + 1, 3, 31)

    # Build shareable CSI pricing map: csi_number → list of entitlement lines
    shareable_ent = query("""
        SELECT cs.csi_number, l.product_family::TEXT AS product_family,
               l.product_name, COALESCE(l.unit_price, 0) AS unit_price
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.sharing_policy = 'shareable'
          AND cs.status = 'active'
          AND (
            cs.owning_client_id = %(cid)s
            OR cs.csi_id IN (SELECT csi_id FROM shared.csi_client_map WHERE client_id = %(cid)s)
          )
    """, {"cid": client_id})

    shareable_csi_nums = set(r["csi_number"] for r in shareable_ent if r["csi_number"])
    _ent_by_csi = {}
    for r in shareable_ent:
        _ent_by_csi.setdefault(r["csi_number"], []).append(r)

    def _snap_unit_price(csi_num, product_family, product_detail):
        lines_e = _ent_by_csi.get(csi_num, [])
        pf = [l for l in lines_e if l["product_family"] == product_family] or lines_e
        if product_detail:
            det = (product_detail or "").lower()
            for l in pf:
                if l["product_name"].lower() in det or det in l["product_name"].lower():
                    return float(l["unit_price"])
        return float(pf[0]["unit_price"]) if pf else 0.0

    snap_rows = query("""
        SELECT s.snapshot_month,
               sl.product_family,
               sl.product_detail,
               sl.csi_number          AS raw_csi_numbers,
               SUM(sl.licences_required) AS licences
        FROM sam_admin.licence_snapshots s
        JOIN sam_admin.licence_snapshot_lines sl ON sl.snapshot_id = s.snapshot_id
        WHERE s.client_id = %s
          AND s.snapshot_month >= %s
          AND s.snapshot_month <= %s
          AND sl.csi_number IS NOT NULL
        GROUP BY s.snapshot_month, sl.product_family, sl.product_detail, sl.csi_number
    """, (client_id, fy_start, fy_end))

    shared_pool_fy_cost = 0.0
    shared_pool_months  = 0
    _months_seen = set()
    for r in snap_rows:
        csi_nums = [c.strip() for c in (r["raw_csi_numbers"] or "").split(";") if c.strip()]
        shared = [n for n in csi_nums if n in shareable_csi_nums]
        if not shared:
            continue
        csi_num  = shared[0]
        up       = _snap_unit_price(csi_num, r["product_family"], r["product_detail"])
        licences = float(r["licences"] or 0)
        shared_pool_fy_cost += licences * up / 12.0
        _months_seen.add(r["snapshot_month"])

    shared_pool_months = len(_months_seen)
    shared_pool_fy_cost = round(shared_pool_fy_cost, 2)

    # If no snapshot data exists yet, fall back to live server_csi_map usage for the current month
    # (total_shared_inuse is licences x unit_price = annual value; divide by 12 for one month)
    shared_pool_estimated = False
    if shared_pool_fy_cost == 0 and total_shared_inuse > 0:
        shared_pool_fy_cost  = round(total_shared_inuse / 12.0, 2)
        shared_pool_months   = 1
        shared_pool_estimated = True

    # Total client cost = annual support on locked licences + actual FY shared pool spend
    total_client_cost = round(total_support + shared_pool_fy_cost, 2)

    # Pie chart: support cost for locked lines; FY shared pool cost spread across shared lines
    shared_pie_total = shared_pool_fy_cost
    shared_lines = [ln for ln in lines if ln["source"] == "shareable"]
    pie_items = []
    for ln in lines:
        if ln["source"] == "client_locked":
            value = ln["support_cost"]
        else:
            # Apportion FY shared pool cost proportionally to assigned_cost weights
            denom = total_shared_inuse if total_shared_inuse else 1
            value = round(shared_pie_total * (ln["assigned_cost"] / denom), 2) if denom else 0
        if value > 0:
            pie_items.append({"label": ln["product_name"], "value": value, "colour": ln["colour"]})
    pie_slices = _pie_slices(pie_items)
    for sl, item in zip(pie_slices, pie_items):
        sl["colour"] = item["colour"]

    return {
        "lines":                lines,
        "total_support":        total_support,
        "total_shared_inuse":   total_shared_inuse,   # current point-in-time (kept for detail table)
        "shared_pool_fy_cost":   shared_pool_fy_cost,
        "shared_pool_months":    shared_pool_months,
        "shared_pool_estimated": shared_pool_estimated,
        "fy_label":             f"FY {fy_year}/{str(fy_year+1)[2:]}",
        "total_client_cost":    total_client_cost,
        "total_assigned":       total_assigned,
        "total_unassigned":     total_unassigned,
        "pie_slices":           pie_slices,
    }


def _build_shared_pool_live():
    """Return list of clients with per-CSI shared pool usage for the monthly overview."""
    shareable_csis = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name,
               lel.product_name, lel.unit_price
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines lel ON lel.csi_id = cs.csi_id AND lel.is_active
        WHERE cs.sharing_policy = 'shareable' AND cs.status = 'active'
        ORDER BY lel.product_name
    """) or []

    if not shareable_csis:
        return []

    csi_meta = {}   # csi_id -> {csi_number, contract_name, product_name, unit_price}
    for r in shareable_csis:
        csi_meta[r["csi_id"]] = {
            "csi_number":    r["csi_number"] or str(r["csi_id"]),
            "contract_name": r["contract_name"],
            "product_name":  r["product_name"],
            "unit_price":    float(r["unit_price"] or 0),
        }
    csi_id_list = list(csi_meta.keys())

    clients_list = query(
        "SELECT client_id, client_code, client_name, schema_name FROM sam_admin.clients "
        "WHERE is_active ORDER BY client_name, client_code"
    ) or []

    client_rows = []
    for c in clients_list:
        s = c["schema_name"]
        try:
            usage = query(f"""
                SELECT csi_id, COALESCE(SUM(licences_consumed), 0)::numeric AS licences_used
                FROM {s}.server_csi_map
                WHERE csi_id = ANY(%s::integer[])
                GROUP BY csi_id
            """, (csi_id_list,)) or []
        except Exception:
            usage = []

        csi_lines = []
        for u in usage:
            licences = float(u["licences_used"] or 0)
            if licences <= 0:
                continue
            meta = csi_meta.get(u["csi_id"])
            if not meta:
                continue
            monthly_cost = round(licences * meta["unit_price"] / 12.0, 2)
            csi_lines.append({
                "csi_number":    meta["csi_number"],
                "contract_name": meta["contract_name"],
                "product_name":  meta["product_name"],
                "unit_price":    meta["unit_price"],
                "licences_used": licences,
                "monthly_cost":  monthly_cost,
            })

        if not csi_lines:
            continue

        csi_lines.sort(key=lambda x: -x["monthly_cost"])
        client_rows.append({
            "client_id":      c["client_id"],
            "client_code":    c["client_code"],
            "client_name":    c["client_name"] or c["client_code"],
            "csi_lines":      csi_lines,
            "total_licences": sum(l["licences_used"] for l in csi_lines),
            "monthly_cost":   round(sum(l["monthly_cost"] for l in csi_lines), 2),
        })

    client_rows.sort(key=lambda x: -x["monthly_cost"])
    return client_rows


def _fy_label(m):
    y = m.year if m.month >= 4 else m.year - 1
    return f"FY {y}/{str(y+1)[-2:]}"


@app.route("/finops/annual-overview")
def finops_annual_overview():
    return redirect(url_for("finops"), 301)


@app.route("/finops/monthly-overview")
@login_required
def finops_monthly_overview():
    if current_role() not in ("superadmin", "contracting"):
        flash("Access restricted.", "danger")
        return redirect(url_for("finops"))

    try:
        live_rows = _build_shared_pool_live()
    except Exception as e:
        app.logger.exception("finops_monthly_overview: _build_shared_pool_live() failed")
        flash(f"Error loading live data: {e}", "danger")
        live_rows = []

    # Historical snapshots — gracefully handle missing/old table
    try:
        snap_rows = query("""
            SELECT s.snapshot_id, s.snapshot_month, s.taken_at, s.taken_by,
                   l.client_id, l.client_name, l.csi_number, l.contract_name,
                   l.product_name, l.licences_used, l.unit_price, l.monthly_cost
            FROM sam_admin.finops_pool_snapshots s
            JOIN sam_admin.finops_pool_snapshot_lines l ON l.snapshot_id = s.snapshot_id
            ORDER BY s.snapshot_month DESC, l.client_name, l.monthly_cost DESC
        """) or []
    except Exception:
        snap_rows = []

    # Group snap rows: month -> {meta, clients: {client_name -> [csi_lines]}}
    snap_map = {}
    try:
        for r in snap_rows:
            m = r["snapshot_month"]
            if m not in snap_map:
                snap_map[m] = {"snapshot_month": m, "taken_at": r["taken_at"],
                               "taken_by": r["taken_by"], "clients": {}}
            clients = snap_map[m]["clients"]
            cname = r["client_name"]
            clients.setdefault(cname, []).append({
                "csi_number":    r["csi_number"],
                "contract_name": r["contract_name"],
                "product_name":  r["product_name"],
                "licences_used": float(r["licences_used"] or 0),
                "unit_price":    float(r["unit_price"] or 0),
                "monthly_cost":  float(r["monthly_cost"] or 0),
            })

        for snap in snap_map.values():
            snap["clients"] = sorted(
                [{"client_name": k, "csi_lines": v,
                  "total_licences": sum(l["licences_used"] for l in v),
                  "monthly_cost": sum(l["monthly_cost"] for l in v)}
                 for k, v in snap["clients"].items()],
                key=lambda x: -x["monthly_cost"]
            )
    except Exception as e:
        app.logger.exception("finops_monthly_overview: snapshot processing failed")
        snap_map = {}

    fy_map = {}
    for m, snap in sorted(snap_map.items(), reverse=True):
        fy_map.setdefault(_fy_label(m), []).append(snap)

    today      = date.today()
    this_month = date(today.year, today.month, 1)
    snapshot_exists = this_month in snap_map

    return render_template("finops_monthly_overview.html",
                           live_rows=live_rows,
                           fy_history=list(fy_map.items()),
                           snapshot_exists=snapshot_exists,
                           this_month=this_month)


@app.route("/finops/monthly-overview/snapshot", methods=["POST"])
@login_required
def finops_pool_snapshot_take():
    if current_role() not in ("superadmin", "contracting"):
        flash("Access restricted.", "danger")
        return redirect(url_for("finops_monthly_overview"))

    today      = date.today()
    snap_month = date(today.year, today.month, 1)
    taken_by   = current_user().get("username", "admin")

    client_rows = _build_shared_pool_live()
    line_count  = sum(len(c["csi_lines"]) for c in client_rows)

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO sam_admin.finops_pool_snapshots (snapshot_month, taken_by)
                    VALUES (%s, %s)
                    ON CONFLICT (snapshot_month) DO UPDATE
                      SET taken_at = NOW(), taken_by = EXCLUDED.taken_by
                    RETURNING snapshot_id
                """, (snap_month, taken_by))
                snap_id = cur.fetchone()[0]
                cur.execute(
                    "DELETE FROM sam_admin.finops_pool_snapshot_lines WHERE snapshot_id = %s",
                    (snap_id,)
                )
                for c in client_rows:
                    for ln in c["csi_lines"]:
                        cur.execute("""
                            INSERT INTO sam_admin.finops_pool_snapshot_lines
                              (snapshot_id, client_id, client_name, csi_number,
                               contract_name, product_name, licences_used, unit_price, monthly_cost)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (snap_id, c["client_id"], c["client_name"],
                              ln["csi_number"], ln["contract_name"], ln["product_name"],
                              ln["licences_used"], ln["unit_price"], ln["monthly_cost"]))
            conn.commit()
        flash(f"Snapshot taken for {snap_month.strftime('%B %Y')} — "
              f"{len(client_rows)} client(s), {line_count} CSI line(s) recorded.", "success")
    except Exception as e:
        flash(f"Snapshot failed: {e}", "danger")

    return redirect(url_for("finops_monthly_overview"))


def _client_has_ulas(schema=None):
    """Return True if the given schema's client (or any client if None/'__all__') has ULA contracts."""
    try:
        if schema and schema != "__all__":
            row = query("""
                SELECT 1 FROM shared.csi_contracts cs
                WHERE cs.is_ula
                  AND (
                    cs.owning_client_id = (
                        SELECT client_id FROM sam_admin.clients WHERE schema_name = %s
                    )
                    OR cs.csi_id IN (
                        SELECT cm.csi_id FROM shared.csi_client_map cm
                        JOIN sam_admin.clients cl ON cl.client_id = cm.client_id
                        WHERE cl.schema_name = %s
                    )
                  )
                LIMIT 1
            """, (schema, schema), fetchall=False)
        else:
            row = query(
                "SELECT 1 FROM shared.csi_contracts WHERE is_ula LIMIT 1",
                fetchall=False
            )
        return bool(row)
    except Exception:
        return True  # fail-open: show the tab rather than hide it on error


@app.route("/finops/current")
@login_required
def finops_current():
    """Redirect to the selected client's cost detail, or the overview if all-clients."""
    schema = get_schema()
    if schema and schema != "__all__":
        client = query(
            "SELECT client_code FROM sam_admin.clients WHERE schema_name = %s AND is_active",
            (schema,), fetchall=False
        )
        if client:
            return redirect(url_for("finops_client", client_code=client["client_code"]))
    return redirect(url_for("finops"))


@app.route("/finops")
@login_required
def finops():
    clients_list = query(
        "SELECT client_id, client_code, client_name FROM sam_admin.clients "
        "WHERE is_active ORDER BY client_name, client_code"
    )
    summary = []
    for c in clients_list:
        data = _build_client_finops(c["client_id"])
        if data:
            summary.append({
                "client_code":        c["client_code"],
                "client_name":        c["client_name"] or c["client_code"],
                "total_support":      data["total_support"],
                "total_shared_inuse": data["total_shared_inuse"],
                "shared_pool_fy_cost":   data["shared_pool_fy_cost"],
                "shared_pool_months":    data["shared_pool_months"],
                "shared_pool_estimated": data["shared_pool_estimated"],
                "total_client_cost":  data["total_client_cost"],
                "product_count":      len(data["lines"]),
            })
    has_ulas = _client_has_ulas(get_schema())
    return render_template("finops_summary.html", summary=summary, has_ulas=has_ulas)


@app.route("/finops/<client_code>")
@login_required
def finops_client(client_code):
    schema = get_schema()
    if schema and schema != "__all__":
        session_client = query(
            "SELECT client_code FROM sam_admin.clients WHERE schema_name = %s AND is_active",
            (schema,), fetchall=False
        )
        if session_client and session_client["client_code"] != client_code:
            return redirect(url_for("finops_client", client_code=session_client["client_code"]))

    client = query(
        "SELECT client_id, client_code, client_name FROM sam_admin.clients "
        "WHERE client_code = %s", (client_code,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("finops"))
    data = _build_client_finops(client["client_id"])
    if not data:
        return render_template("finops.html", client=None,
                               client_name=client["client_name"] or client["client_code"],
                               client_code=client["client_code"])
    data["client_code"] = client["client_code"]
    data["client_name"] = client["client_name"] or client["client_code"]
    return render_template("finops.html", client=data)


@app.route("/finops/ulas")
@login_required
def finops_ulas():
    selected_csi = request.args.get("csi_id", type=int)

    ulas = query("""
        SELECT csi_id, csi_number, contract_name, ula_expiry, status
        FROM shared.csi_contracts
        WHERE is_ula
        ORDER BY contract_name
    """)

    clients_list = query(
        "SELECT client_code, client_name, schema_name FROM sam_admin.clients WHERE is_active"
    )

    ula_map = {}
    for u in ulas:
        ula_map[u["csi_id"]] = {
            "csi_id":        u["csi_id"],
            "csi_number":    u["csi_number"],
            "contract_name": u["contract_name"],
            "ula_expiry":    u["ula_expiry"],
            "status":        u["status"],
            "server_count":  0,
            "servers":       [],
            "product_totals": {},
        }

    for c in clients_list:
        s = c["schema_name"]
        try:
            rows = query(f"""
                SELECT DISTINCT ON (m.csi_id, sv.server_id, lp.product_family, lp.product_detail)
                    m.csi_id,
                    sv.server_id,
                    sv.hostname,
                    lp.product_family::TEXT AS product_family,
                    lp.product_detail::TEXT AS product_detail,
                    lp.licences_required
                FROM {s}.server_csi_map m
                JOIN shared.csi_contracts cs ON cs.csi_id = m.csi_id AND cs.is_ula
                JOIN {s}.oracle_servers sv ON sv.server_id = m.server_id AND sv.is_active
                LEFT JOIN {s}.license_position lp ON lp.server_id = sv.server_id
            """)
            seen_servers = set()
            for r in rows:
                csi_id = r["csi_id"]
                if csi_id not in ula_map:
                    continue
                entry = ula_map[csi_id]
                srv_key = (c["client_code"], r["server_id"])
                if srv_key not in seen_servers:
                    seen_servers.add(srv_key)
                    entry["server_count"] += 1
                    entry["servers"].append({
                        "hostname":    r["hostname"],
                        "client_name": c["client_name"] or c["client_code"],
                        "client_code": c["client_code"],
                        "server_id":   r["server_id"],
                    })
                if r["product_detail"] and r["licences_required"]:
                    prod_key = r["product_detail"]
                    entry["product_totals"][prod_key] = (
                        entry["product_totals"].get(prod_key, 0) + r["licences_required"]
                    )
        except Exception:
            pass

    ula_list = list(ula_map.values())
    selected_ula = ula_map.get(selected_csi) if selected_csi else None

    return render_template(
        "finops_ula.html",
        ulas=ula_list,
        selected_csi=selected_csi,
        selected_ula=selected_ula,
        has_ulas=True,
    )


@app.route("/finops/shared-pool-usage")
@login_required
def shared_pool_usage():
    return redirect(url_for("licence_summary_shared"))


@app.route("/finops/shared-pool-monthly")
@login_required
def shared_pool_monthly():
    from calendar import month_abbr

    schema = get_schema()
    if schema == "__all__":
        flash("Please select a specific client to view shared pool monthly costs.", "warning")
        return redirect(url_for("finops"))

    client = query(
        "SELECT client_id, client_name, client_code FROM sam_admin.clients "
        "WHERE schema_name = %s AND is_active",
        (schema,), fetchall=False
    )
    if not client:
        return redirect(url_for("finops"))

    client_id = client["client_id"]

    # Fiscal year: April 1 – March 31
    today = date.today()
    fy_year = today.year if today.month >= 4 else today.year - 1
    fy_start = date(fy_year, 4, 1)
    fy_end   = date(fy_year + 1, 3, 31)
    fy_label = f"FY {fy_year}/{str(fy_year + 1)[2:]}"

    # Build ordered list of FY months (Apr … Mar)
    fy_months = []
    y, m = fy_year, 4
    for _ in range(12):
        fy_months.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1

    # Shareable CSIs accessible to this client + their entitlement lines
    ent_rows = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.currency,
               l.product_name, l.product_family::TEXT AS product_family,
               COALESCE(l.unit_price, 0)              AS unit_price,
               l.quantity                              AS entitled_qty
        FROM shared.csi_contracts cs
        JOIN shared.license_entitlement_lines l
             ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.sharing_policy = 'shareable'
          AND cs.status = 'active'
          AND (
            cs.owning_client_id = %(cid)s
            OR cs.csi_id IN (
                SELECT csi_id FROM shared.csi_client_map WHERE client_id = %(cid)s
            )
          )
        ORDER BY cs.csi_number, l.product_name
    """, {"cid": client_id})

    # Index: csi_number → list of entitlement lines
    csi_num_to_lines = {}
    for r in ent_rows:
        csi_num_to_lines.setdefault(r["csi_number"], []).append(r)

    # Set of shareable CSI numbers for quick lookup
    shareable_csi_numbers = set(csi_num_to_lines.keys())
    currency = ent_rows[0]["currency"] if ent_rows else "USD"

    # Helper: given a product_family + product_detail, find best unit_price from an
    # entitlement line by partial-name match, falling back to first line in that CSI
    def _unit_price(csi_number, product_family, product_detail):
        lines = csi_num_to_lines.get(csi_number, [])
        if not lines:
            return 0.0
        pf_lines = [l for l in lines if l["product_family"] == product_family]
        if not pf_lines:
            pf_lines = lines
        if product_detail:
            det = product_detail.lower()
            for l in pf_lines:
                if l["product_name"].lower() in det or det in l["product_name"].lower():
                    return float(l["unit_price"])
        return float(pf_lines[0]["unit_price"])

    # -----------------------------------------------------------------------
    # Snapshots for this client in the fiscal year
    # -----------------------------------------------------------------------
    snap_rows = query("""
        SELECT s.snapshot_month,
               sl.product_family,
               sl.product_detail,
               sl.csi_number          AS raw_csi_numbers,
               SUM(sl.licences_required) AS licences
        FROM sam_admin.licence_snapshots s
        JOIN sam_admin.licence_snapshot_lines sl ON sl.snapshot_id = s.snapshot_id
        WHERE s.client_id = %s
          AND s.snapshot_month >= %s
          AND s.snapshot_month <= %s
          AND sl.csi_number IS NOT NULL
        GROUP BY s.snapshot_month, sl.product_family, sl.product_detail, sl.csi_number
        ORDER BY s.snapshot_month, sl.product_family, sl.product_detail
    """, (client_id, fy_start, fy_end))

    # Group snapshot lines by month, filtering to shared-pool CSIs only
    snap_by_month = {}
    for r in snap_rows:
        # raw_csi_numbers is e.g. "12345; 67890"
        csi_nums = [c.strip() for c in (r["raw_csi_numbers"] or "").split(";") if c.strip()]
        shared_nums = [n for n in csi_nums if n in shareable_csi_numbers]
        if not shared_nums:
            continue
        m = r["snapshot_month"].replace(day=1)
        snap_by_month.setdefault(m, []).append({
            "csi_number":    shared_nums[0],   # attribute to first matching shared CSI
            "product_family": r["product_family"],
            "product_detail": r["product_detail"],
            "licences":       float(r["licences"] or 0),
        })

    # -----------------------------------------------------------------------
    # Live current-month data from server_csi_map (used when no snapshot yet)
    # -----------------------------------------------------------------------
    cur_month = today.replace(day=1)
    live_lines = []
    cur_month_is_live = False
    if cur_month not in snap_by_month and cur_month >= fy_start and cur_month <= fy_end:
        try:
            live_rows = query(f"""
                SELECT cs.csi_number,
                       scm.product_family::TEXT AS product_family,
                       COALESCE(scm.product_detail, '') AS product_detail,
                       SUM(scm.licences_consumed) AS licences
                FROM {schema}.server_csi_map scm
                JOIN shared.csi_contracts cs ON cs.csi_id = scm.csi_id
                WHERE cs.sharing_policy = 'shareable'
                  AND cs.status = 'active'
                GROUP BY cs.csi_number, scm.product_family, scm.product_detail
            """)
            for r in live_rows:
                if r["csi_number"] in shareable_csi_numbers:
                    live_lines.append({
                        "csi_number":    r["csi_number"],
                        "product_family": r["product_family"],
                        "product_detail": r["product_detail"],
                        "licences":       float(r["licences"] or 0),
                    })
        except Exception:
            pass
    if live_lines:
        snap_by_month[cur_month] = live_lines
        cur_month_is_live = True

    # -----------------------------------------------------------------------
    # Build per-month summary
    # -----------------------------------------------------------------------
    monthly_data = []
    fy_total = 0.0

    for mo in fy_months:
        is_future = mo > cur_month
        is_live   = (mo == cur_month) and cur_month_is_live
        has_data  = mo in snap_by_month and not is_future

        lines_out = []
        month_total = 0.0

        if has_data:
            # Group by (csi_number, product_detail) to avoid duplicate rows
            grouped = {}
            for ln in snap_by_month[mo]:
                key = (ln["csi_number"], ln["product_detail"] or ln["product_family"])
                if key not in grouped:
                    grouped[key] = {
                        "csi_number":     ln["csi_number"],
                        "contract_name":  (csi_num_to_lines.get(ln["csi_number"]) or [{}])[0].get("contract_name", "—"),
                        "product":        ln["product_detail"] or ln["product_family"],
                        "licences":       0.0,
                        "unit_price":     _unit_price(ln["csi_number"], ln["product_family"], ln["product_detail"]),
                    }
                grouped[key]["licences"] += ln["licences"]

            for item in sorted(grouped.values(), key=lambda x: (x["csi_number"], x["product"])):
                monthly_cost = round(item["licences"] * item["unit_price"] / 12, 2)
                item["monthly_cost"] = monthly_cost
                month_total += monthly_cost
                lines_out.append(item)

        fy_total += month_total
        monthly_data.append({
            "month":       mo,
            "label":       f"{month_abbr[mo.month]} {mo.year}",
            "is_future":   is_future,
            "is_live":     is_live,
            "has_data":    has_data,
            "lines":       lines_out,
            "month_total": round(month_total, 2),
        })

    # Bar chart data (only months with data)
    bar_months  = [d["label"]       for d in monthly_data if not d["is_future"]]
    bar_totals  = [d["month_total"] for d in monthly_data if not d["is_future"]]

    return render_template(
        "finops_shared_pool_monthly.html",
        client=client,
        fy_label=fy_label,
        fy_start=fy_start,
        fy_end=fy_end,
        monthly_data=monthly_data,
        fy_total=round(fy_total, 2),
        currency=currency,
        bar_months=bar_months,
        bar_totals=bar_totals,
        ent_rows=ent_rows,
    )


@app.route("/renewal-calendar")
@login_required
def renewal_calendar():
    today = date.today()
    renewals = query("""
        SELECT
            cs.csi_id, cs.csi_number, cs.contract_name, cs.is_ula,
            cs.status, cs.currency,
            cs.support_expiry, cs.ula_expiry,
            CASE WHEN cs.is_ula THEN cs.ula_expiry ELSE cs.support_expiry END AS expiry_date,
            oc.client_name AS owning_client,
            COALESCE(SUM(l.annual_support_cost), 0) AS annual_support_cost
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        LEFT JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        WHERE (cs.is_ula AND cs.ula_expiry IS NOT NULL)
           OR (NOT cs.is_ula AND cs.support_expiry IS NOT NULL)
        GROUP BY cs.csi_id, cs.csi_number, cs.contract_name, cs.is_ula,
                 cs.status, cs.currency, cs.support_expiry, cs.ula_expiry,
                 oc.client_name
        ORDER BY expiry_date
    """)

    # Bucket into time bands
    past, within_30, within_90, within_12m, beyond = [], [], [], [], []
    for r in renewals:
        exp = r['expiry_date']
        if exp is None:
            continue
        delta = (exp - today).days
        r = dict(r)
        r['days'] = delta
        if delta < 0:
            past.append(r)
        elif delta <= 30:
            within_30.append(r)
        elif delta <= 90:
            within_90.append(r)
        elif delta <= 365:
            within_12m.append(r)
        else:
            beyond.append(r)

    return render_template("renewal_calendar.html",
                           today=today,
                           past=past,
                           within_30=within_30,
                           within_90=within_90,
                           within_12m=within_12m,
                           beyond=beyond)


@app.route("/cost-optimisation")
@login_required
def cost_optimisation():
    today = date.today()

    # All active standard (non-ULA) entitlement lines with cost info
    line_rows = query("""
        SELECT l.line_id, l.csi_id, l.line_number, l.product_name,
               l.product_family::TEXT AS product_family,
               l.license_metric::TEXT AS license_metric,
               l.quantity, l.unit_price, l.total_price, l.annual_support_cost,
               cs.csi_number, cs.contract_name, cs.currency, cs.status,
               cs.support_expiry, cs.is_ula,
               oc.client_name AS owning_client
        FROM shared.license_entitlement_lines l
        JOIN shared.csi_contracts cs ON cs.csi_id = l.csi_id
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        WHERE l.is_active AND NOT cs.is_ula AND cs.status = 'active'
        ORDER BY l.annual_support_cost DESC NULLS LAST, l.total_price DESC NULLS LAST
    """)

    # Build consumed map (same pattern as contracts())
    active_schemas = [
        r["schema_name"] for r in query(
            "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
        )
    ]
    consumed_entries_by_csi = {}
    if active_schemas:
        union_sql = " UNION ALL ".join(
            f"SELECT csi_id, product_family::TEXT, COALESCE(product_detail,'') AS product_detail, "
            f"COALESCE(SUM(licences_consumed), 0) AS consumed "
            f"FROM {s}.server_csi_map GROUP BY csi_id, product_family, product_detail"
            for s in active_schemas
        )
        for r in query(
            f"SELECT csi_id, product_family, product_detail, SUM(consumed) AS total "
            f"FROM ({union_sql}) t GROUP BY csi_id, product_family, product_detail"
        ):
            consumed_entries_by_csi.setdefault(r["csi_id"], []).append(
                (r["product_family"], r["product_detail"], r["total"])
            )

    # Annotate each line with consumed/available/utilisation
    annotated = []
    for l in line_rows:
        consumed = float(sum(
            float(amt) for fam, det, amt in consumed_entries_by_csi.get(l["csi_id"], [])
            if _is_compatible_product(l["product_family"], det or None,
                                      l["product_family"], l["product_name"])
        ))
        qty = float(l["quantity"] or 0)
        util_pct = int(consumed / qty * 100) if qty else 0
        waste_licences = max(qty - consumed, 0)
        # Estimate wasted annual cost (pro-rate annual_support_cost by unused fraction)
        wasted_cost = None
        if qty and l["annual_support_cost"]:
            wasted_cost = float(l["annual_support_cost"]) * (waste_licences / qty)
        elif qty and l["total_price"]:
            wasted_cost = float(l["total_price"]) * (waste_licences / qty)
        annotated.append(dict(l,
                              consumed=consumed,
                              available=waste_licences,
                              util_pct=util_pct,
                              wasted_cost=wasted_cost))

    # Bucket findings
    unused   = [r for r in annotated if r["quantity"] and r["consumed"] == 0]
    low_util = [r for r in annotated if r["quantity"] and 0 < r["util_pct"] < 50]
    # Sort each bucket: most costly waste first
    unused.sort(key=lambda r: r["wasted_cost"] or 0, reverse=True)
    low_util.sort(key=lambda r: r["wasted_cost"] or 0, reverse=True)

    # Contracts with no entitlement lines (active, non-ULA)
    empty_contracts = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.currency,
               oc.client_name AS owning_client
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        LEFT JOIN shared.license_entitlement_lines l
               ON l.csi_id = cs.csi_id AND l.is_active
        WHERE cs.status = 'active' AND NOT cs.is_ula
        GROUP BY cs.csi_id, cs.csi_number, cs.contract_name, cs.currency, oc.client_name
        HAVING COUNT(l.line_id) = 0
        ORDER BY cs.contract_name
    """)

    # ULAs expiring within 12 months — check server assignment counts
    ula_risks = []
    ula_contracts = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.ula_expiry, cs.currency,
               oc.client_name AS owning_client, oc.schema_name
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        WHERE cs.is_ula AND cs.status = 'active'
          AND cs.ula_expiry BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '12 months'
        ORDER BY cs.ula_expiry
    """)
    for ula in ula_contracts:
        server_count = 0
        if ula["schema_name"]:
            s = ula["schema_name"]
            row = query(
                f"SELECT COUNT(DISTINCT server_id) AS cnt FROM {s}.server_csi_map WHERE csi_id = %s",
                (ula["csi_id"],), fetchall=False
            )
            server_count = row["cnt"] if row else 0
        days = (ula["ula_expiry"] - today).days
        ula_risks.append(dict(ula, server_count=server_count, days=days))

    # Summary totals
    total_wasted_cost = sum((r["wasted_cost"] or 0) for r in unused + low_util)
    total_unused_licences = sum(r["available"] for r in unused)

    return render_template("cost_optimisation.html",
                           today=today,
                           unused=unused,
                           low_util=low_util,
                           empty_contracts=empty_contracts,
                           ula_risks=ula_risks,
                           total_wasted_cost=total_wasted_cost,
                           total_unused_licences=total_unused_licences)


@app.route("/api/licence-analysis/servers")
@login_required
def licence_analysis_servers_api():
    """JSON endpoint — returns servers for a client, for the analysis form dropdown."""
    client_id = request.args.get("client_id", "")
    if not client_id:
        return jsonify([])
    client = query(
        "SELECT schema_name FROM sam_admin.clients WHERE client_id=%s AND is_active",
        (int(client_id),), fetchall=False
    )
    if not client:
        return jsonify([])
    schema = client["schema_name"]
    rows = query(f"""
        SELECT sv.server_id,
               sv.hostname,
               sv.environment::TEXT                   AS environment,
               COALESCE(op.cpu_sockets, 1)            AS cpu_sockets,
               COALESCE(op.cores_per_socket, 1)       AS cores_per_socket,
               COALESCE(op.total_physical_cores,
                        op.cpu_sockets * op.cores_per_socket, 1) AS total_physical_cores
        FROM   {schema}.oracle_servers sv
        LEFT JOIN {schema}.oracle_processors op ON op.server_id = sv.server_id
        WHERE  sv.is_active IS NOT FALSE
        ORDER  BY sv.hostname
    """)
    return jsonify([dict(r) for r in rows])


def _handle_price_import(req, sess):
    """
    Parse an uploaded Oracle Technology Global Price List Excel file and
    upsert matching rows into shared.oracle_product_list_prices.

    The Oracle price list Excel has varied column layouts across versions.
    We scan the header row for recognisable column names rather than using
    fixed column indices.

    Returns a Flask redirect with a session flash-style query param so the
    template can show a result banner.
    """
    import openpyxl
    from openpyxl import load_workbook

    file = req.files.get("price_file")
    if not file or not file.filename:
        return redirect(url_for("oracle_price_list", import_err="no_file"))

    eff_date = req.form.get("import_date") or None
    mark_current = req.form.get("mark_current") == "1"
    updated_by   = sess.get("username", "import")

    try:
        file_bytes = io.BytesIO(file.read())
        wb = load_workbook(filename=file_bytes, read_only=True, data_only=True)
    except Exception as e:
        return redirect(url_for("oracle_price_list", import_err=str(e)))

    # Oracle price list may have multiple sheets; find the one with product data.
    # Typically the main sheet is called "Technology" or "Database" or Sheet1.
    target_sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if any(k in nl for k in ("technology", "database", "product", "price")):
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    # Scan rows to find the header row (contains "Part" or "Product" etc.)
    # We allow the header to appear anywhere in the first 20 rows.
    HEADER_KEYWORDS = {"part", "product", "license", "licence", "metric", "processor", "named user"}
    header_row_idx  = None
    col_map         = {}   # col name → 0-based column index

    rows_iter = target_sheet.iter_rows(values_only=True)
    raw_rows  = []
    for i, row in enumerate(rows_iter):
        raw_rows.append(row)
        if i > 19:  # only search first 20 rows for header
            break

    for i, row in enumerate(raw_rows):
        cells = [str(c).lower().strip() if c is not None else "" for c in row]
        hits  = sum(1 for c in cells if any(k in c for k in HEADER_KEYWORDS))
        if hits >= 2:
            header_row_idx = i
            for j, c in enumerate(cells):
                col_map[c] = j
            break

    if header_row_idx is None:
        return redirect(url_for("oracle_price_list", import_err="no_header"))

    # Find key columns by fuzzy name matching
    def find_col(*candidates):
        for cand in candidates:
            for key, idx in col_map.items():
                if cand in key:
                    return idx
        return None

    col_product   = find_col("product", "description", "name")
    col_processor = find_col("processor license", "processor lic", "full use", "processor")
    col_nup       = find_col("named user plus", "named user", "nup")
    col_metric    = find_col("license metric", "licence metric", "metric")
    col_part      = find_col("part number", "part #", "part no", "ordering")

    if col_product is None or col_processor is None:
        return redirect(url_for("oracle_price_list", import_err="no_cols"))

    # Reopen from the in-memory buffer (file.stream is already consumed)
    file_bytes.seek(0)
    wb2 = load_workbook(filename=file_bytes, read_only=True, data_only=True)
    sheet2 = wb2[target_sheet.title]

    imported = 0
    skipped  = 0

    for row_i, row in enumerate(sheet2.iter_rows(values_only=True)):
        if row_i <= header_row_idx:
            continue  # skip header and anything before it

        product_name = str(row[col_product]).strip() if row[col_product] is not None else ""
        if not product_name or product_name.lower() in ("none", "nan", ""):
            continue

        # Skip section headers / totals / blank rows
        proc_raw = row[col_processor] if col_processor < len(row) else None
        nup_raw  = row[col_nup]       if col_nup is not None and col_nup < len(row) else None

        def to_price(val):
            if val is None:
                return None
            try:
                return float(str(val).replace(",", "").replace("$", "").strip())
            except (ValueError, TypeError):
                return None

        proc_price = to_price(proc_raw)
        nup_price  = to_price(nup_raw)

        if proc_price is None and nup_price is None:
            skipped += 1
            continue

        rows_to_upsert = []
        if proc_price is not None:
            rows_to_upsert.append((product_name, "processor", proc_price))
        if nup_price is not None:
            rows_to_upsert.append((product_name, "named user plus", nup_price))

        for pname, metric, price in rows_to_upsert:
            try:
                execute(
                    "INSERT INTO shared.oracle_product_list_prices "
                    "(product_name, metric, list_price, currency, effective_date, is_current, notes, updated_by) "
                    "VALUES (%s,%s,%s,'USD',COALESCE(%s::DATE,CURRENT_DATE),%s,'Imported from Oracle price list',%s) "
                    "ON CONFLICT (product_name, metric, effective_date) DO UPDATE SET "
                    "  list_price=EXCLUDED.list_price, is_current=EXCLUDED.is_current, "
                    "  notes=EXCLUDED.notes, updated_by=EXCLUDED.updated_by",
                    (pname, metric, price, eff_date, mark_current, updated_by)
                )
                imported += 1
            except Exception:
                skipped += 1

    return redirect(url_for("oracle_price_list", imported=imported, skipped=skipped))


# Oracle Technology Global Price List — April 2026 published prices (USD)
# Source: oracle.com/us/corporate/pricing/technology-price-list-070617.pdf
ORACLE_PUBLISHED_PRICES = [
    # product_name, metric, list_price
    # ── Database Editions ────────────────────────────────────────────────
    ("Oracle Database Enterprise Edition",          "processor",       47500.00),
    ("Oracle Database Enterprise Edition",          "named_user_plus",   950.00),
    # ── Management Packs ─────────────────────────────────────────────────
    ("Oracle Database Diagnostics Pack",            "processor",        7500.00),
    ("Oracle Database Diagnostics Pack",            "named_user_plus",   150.00),
    ("Oracle Database Tuning Pack",                 "processor",        5000.00),
    ("Oracle Database Tuning Pack",                 "named_user_plus",   100.00),
    ("Oracle Database Lifecycle Management Pack",   "processor",        5000.00),
    ("Oracle Database Lifecycle Management Pack",   "named_user_plus",   100.00),
    ("Oracle Configuration Management Pack",        "processor",        5000.00),
    ("Oracle Configuration Management Pack",        "named_user_plus",   100.00),
    ("Oracle Provisioning and Patch Automation",    "processor",        5000.00),
    ("Oracle Provisioning and Patch Automation",    "named_user_plus",   100.00),
    # ── Database Editions (continued) ────────────────────────────────────
    ("Oracle Database Standard Edition 2",          "processor",       17500.00),
    ("Oracle Database Standard Edition 2",          "named_user_plus",   350.00),
    ("Oracle Database Personal Edition",            "named_user_plus",   460.00),
    # ── Database Options ─────────────────────────────────────────────────
    ("Oracle Partitioning",                         "processor",       11500.00),
    ("Oracle Partitioning",                         "named_user_plus",   230.00),
    ("Oracle Real Application Clusters",            "processor",       23000.00),
    ("Oracle Real Application Clusters",            "named_user_plus",   460.00),
    ("Oracle Multitenant",                          "processor",       17500.00),
    ("Oracle Multitenant",                          "named_user_plus",   350.00),
    ("Oracle Active Data Guard",                    "processor",       11500.00),
    ("Oracle Active Data Guard",                    "named_user_plus",   230.00),
    ("Oracle Advanced Security",                    "processor",       15000.00),
    ("Oracle Advanced Security",                    "named_user_plus",   300.00),
    ("Oracle Label Security",                       "processor",        5000.00),
    ("Oracle Label Security",                       "named_user_plus",   100.00),
    ("Oracle Database Vault",                       "processor",       10000.00),
    ("Oracle Database Vault",                       "named_user_plus",   200.00),
    ("Oracle OLAP",                                 "processor",       11500.00),
    ("Oracle OLAP",                                 "named_user_plus",   230.00),
    ("Oracle Spatial and Graph",                    "processor",       17500.00),
    ("Oracle Spatial and Graph",                    "named_user_plus",   350.00),
    ("Oracle GoldenGate",                           "processor",       17500.00),
    ("Oracle GoldenGate",                           "named_user_plus",   350.00),
    ("Oracle RAC One Node",                         "processor",       10000.00),
    ("Oracle RAC One Node",                         "named_user_plus",   200.00),
    ("Oracle In-Memory",                            "processor",       23000.00),
    ("Oracle In-Memory",                            "named_user_plus",   460.00),
    # ── WebLogic ─────────────────────────────────────────────────────────
    ("Oracle WebLogic Server Enterprise Edition",   "processor",       45000.00),
    ("Oracle WebLogic Server Enterprise Edition",   "named_user_plus",   900.00),
    ("Oracle WebLogic Server Standard Edition",     "processor",       15000.00),
    ("Oracle WebLogic Server Standard Edition",     "named_user_plus",   300.00),
    ("Oracle WebLogic Suite",                       "processor",       90000.00),
    ("Oracle WebLogic Suite",                       "named_user_plus",  1800.00),
]


@app.route("/oracle-price-list/seed", methods=["POST"])
@login_required
def oracle_price_list_seed():
    role = current_role()
    if role not in ("superadmin", "contracting", "dba"):
        return abort(403)
    eff_date = "2026-04-16"
    updated_by = session.get("username", "seed")
    try:
        conn = get_db()
        try:
            with conn.cursor() as cur:
                for product_name, metric, list_price in ORACLE_PUBLISHED_PRICES:
                    cur.execute(
                        "INSERT INTO shared.oracle_product_list_prices "
                        "(product_name, metric, list_price, currency, effective_date, is_current, notes, updated_by) "
                        "VALUES (%s,%s,%s,'USD',%s::DATE,true,"
                        "'Oracle Technology Global Price List Apr-2026',%s) "
                        "ON CONFLICT (product_name, metric, effective_date) DO UPDATE SET "
                        "  list_price=EXCLUDED.list_price, is_current=true, "
                        "  notes=EXCLUDED.notes, updated_by=EXCLUDED.updated_by",
                        (product_name, metric, list_price, eff_date, updated_by)
                    )
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        app.logger.error("oracle_price_list_seed error: %s", e)
        return redirect(url_for("oracle_price_list", seed_err=str(e)))
    return redirect(url_for("oracle_price_list", seeded=len(ORACLE_PUBLISHED_PRICES)))


@app.route("/oracle-price-list", methods=["GET", "POST"])
@login_required
def oracle_price_list():
    role = current_role()
    if role not in ("superadmin", "contracting", "dba"):
        return abort(403)

    try:
        prices = query(
            "SELECT price_id, product_name, metric, list_price, currency, "
            "       effective_date, is_current, notes "
            "FROM shared.oracle_product_list_prices "
            "ORDER BY effective_date DESC, product_name, metric"
        )
        _seen, _ordered = set(), []
        for pn, _, _ in ORACLE_PUBLISHED_PRICES:
            if pn not in _seen:
                _ordered.append(pn)
                _seen.add(pn)
        prices = sorted(prices, key=lambda p: _ordered.index(p["product_name"]) if p["product_name"] in _seen else len(_ordered))
    except Exception as e:
        app.logger.error("oracle_price_list DB error: %s", e, exc_info=True)
        prices = []
        return render_template(
            "oracle_price_list.html",
            prices=prices,
            db_error=str(e),
            imported=None, skipped=None, import_err=None,
            seeded=None, seed_err=None,
            oracle_price_count=len(ORACLE_PUBLISHED_PRICES),
        )

    if request.method == "POST":
        action = request.form.get("action", "")

        if action == "save_price":
            pid      = request.form.get("price_id")
            pname    = request.form.get("product_name", "").strip()
            metric   = request.form.get("metric", "").strip()
            lprice   = request.form.get("list_price", "").strip()
            currency = request.form.get("currency", "USD").strip().upper() or "USD"
            eff_date = request.form.get("effective_date") or None
            notes    = request.form.get("notes", "").strip() or None
            is_current = request.form.get("is_current") == "1"
            if pname and metric and lprice:
                if pid:
                    execute(
                        "UPDATE shared.oracle_product_list_prices "
                        "SET product_name=%s, metric=%s, list_price=%s, currency=%s, "
                        "    effective_date=COALESCE(%s::DATE, CURRENT_DATE), is_current=%s, "
                        "    notes=%s, updated_by=%s "
                        "WHERE price_id=%s",
                        (pname, metric, float(lprice), currency,
                         eff_date, is_current, notes,
                         session.get("username", "system"), int(pid))
                    )
                else:
                    execute(
                        "INSERT INTO shared.oracle_product_list_prices "
                        "(product_name, metric, list_price, currency, effective_date, is_current, notes, updated_by) "
                        "VALUES (%s,%s,%s,%s,COALESCE(%s::DATE,CURRENT_DATE),%s,%s,%s) "
                        "ON CONFLICT (product_name, metric, effective_date) DO UPDATE SET "
                        "  list_price=EXCLUDED.list_price, is_current=EXCLUDED.is_current, "
                        "  notes=EXCLUDED.notes, updated_by=EXCLUDED.updated_by",
                        (pname, metric, float(lprice), currency,
                         eff_date, is_current, notes,
                         session.get("username", "system"))
                    )
            return redirect(url_for("oracle_price_list"))

        if action == "delete_price":
            pid = request.form.get("price_id")
            if pid:
                execute("DELETE FROM shared.oracle_product_list_prices WHERE price_id=%s", (int(pid),))
            return redirect(url_for("oracle_price_list"))

        if action == "import_prices":
            return _handle_price_import(request, session)

    return render_template(
        "oracle_price_list.html",
        prices=prices,
        db_error=None,
        imported=request.args.get("imported"),
        skipped=request.args.get("skipped"),
        import_err=request.args.get("import_err"),
        seeded=request.args.get("seeded"),
        seed_err=request.args.get("seed_err"),
        oracle_price_count=len(ORACLE_PUBLISHED_PRICES),
    )


@app.route("/licence-analysis", methods=["GET", "POST"])
@login_required
def licence_analysis():
    role = current_role()
    if role not in ("superadmin", "contracting", "dba"):
        return abort(403)

    SUPPORT_RATE = 0.22  # Oracle standard annual support rate — locked

    clients = query(
        "SELECT client_id, client_code, client_name FROM sam_admin.clients "
        "WHERE is_active ORDER BY client_name"
    )
    prices = query(
        "SELECT product_name, metric, list_price, is_current "
        "FROM shared.oracle_product_list_prices"
    )
    # ------------------------------------------------------------------
    # Analysis form values
    # ------------------------------------------------------------------
    mode = request.args.get("mode", "server")   # "server" or "manual"

    def _adj(name):
        """Parse an optional adjustment query param as float, default 0."""
        try:
            v = float(request.args.get(name, "") or 0)
            return max(v, 0)
        except (ValueError, TypeError):
            return 0.0

    _DEFAULT_MANAGED_PER_CORE = 2175.58
    _onprem_raw = request.args.get("adj_onprem", None)
    onprem_per_core = max(float(_onprem_raw), 0) if _onprem_raw not in (None, "") else _DEFAULT_MANAGED_PER_CORE
    adj_oci            = _adj("adj_oci")
    _managed_raw = request.args.get("adj_exacc", None)
    managed_per_core = max(float(_managed_raw), 0) if _managed_raw not in (None, "") else _DEFAULT_MANAGED_PER_CORE
    adj_azure          = _adj("adj_azure")
    adj_onprem_upfront = 0.0  # removed from UI; kept at zero for downstream compat
    adj_exacc_upfront  = _adj("adj_exacc_upfront")
    adj_onprem = 0.0   # resolved per-analysis once physical_cores is known
    adj_exacc  = 0.0   # resolved per-analysis once physical_cores is known
    adj_any = any([onprem_per_core, adj_oci, managed_per_core, adj_azure])

    try:
        _ld = float(request.args.get("licence_discount", "") or 0)
        licence_discount_pct = max(0.0, min(100.0, _ld))
    except (ValueError, TypeError):
        licence_discount_pct = 0.0

    form_vals = {
        "mode":                mode,
        "client_id":           request.args.get("client_id", ""),
        "server_id":           request.args.get("server_id", ""),
        # manual-entry fields
        "m_client_id":         request.args.get("m_client_id", ""),
        "m_hostname":          request.args.get("m_hostname", ""),
        "m_cores":             request.args.get("m_cores", ""),
        "m_sockets":           request.args.get("m_sockets", "1"),
        "m_ram_gb":            request.args.get("m_ram_gb", ""),
        "m_edition":           request.args.get("m_edition", "Enterprise Edition"),
        "horizon_years":       request.args.get("horizon_years", "5"),
        # vendor quote adjustments
        "adj_onprem":          onprem_per_core    if onprem_per_core    else "",
        "adj_oci":             int(adj_oci)       if adj_oci            else "",
        "adj_exacc":           managed_per_core   if managed_per_core   else "",
        "adj_azure":           int(adj_azure)          if adj_azure          else "",
        "adj_onprem_upfront":  int(adj_onprem_upfront) if adj_onprem_upfront else "",
        "adj_exacc_upfront":   int(adj_exacc_upfront)  if adj_exacc_upfront  else "",
        "adj_any":             adj_any,
        "licence_discount":    licence_discount_pct if licence_discount_pct else "",
        # manual mode: which options are selected
        "m_options": [p for p, _, _, _ in MANUAL_DB_OPTIONS
                      if request.args.get(p)],
    }

    result      = None
    server_list = []  # pre-populate server dropdown if client already selected

    if form_vals["client_id"]:
        try:
            client_id_int = int(form_vals["client_id"])
            cl = query(
                "SELECT schema_name FROM sam_admin.clients WHERE client_id=%s AND is_active",
                (client_id_int,), fetchall=False
            )
            if cl:
                schema = cl["schema_name"]
                server_list = query(f"""
                    SELECT sv.server_id,
                           sv.hostname,
                           sv.environment::TEXT                         AS environment,
                           COALESCE(op.cpu_sockets, 1)                 AS cpu_sockets,
                           COALESCE(op.cores_per_socket, 1)            AS cores_per_socket,
                           COALESCE(op.total_physical_cores,
                                    op.cpu_sockets * op.cores_per_socket, 1) AS total_physical_cores
                    FROM   {schema}.oracle_servers sv
                    LEFT JOIN {schema}.oracle_processors op ON op.server_id = sv.server_id
                    WHERE  sv.is_active IS NOT FALSE
                    ORDER  BY sv.hostname
                """)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Run analysis when enough parameters are present
    # ------------------------------------------------------------------
    ready_server = (mode == "server" and form_vals["client_id"] and form_vals["server_id"])
    ready_manual = (mode == "manual" and form_vals["m_cores"])

    if ready_server or ready_manual:
        try:
            horizon = min(max(int(form_vals["horizon_years"]), 1), 20)

            # Build a price lookup map from catalogue
            price_map: dict = {}
            for p in prices:
                if p["is_current"]:
                    price_map[(p["product_name"].lower(), p["metric"])] = float(p["list_price"])

            def find_price(product_label, metric):
                pl_lower = (product_label or "").lower()
                if (pl_lower, metric) in price_map:
                    return price_map[(pl_lower, metric)]
                for (pn, m), pr in price_map.items():
                    if m == metric and (pn in pl_lower or pl_lower in pn):
                        return pr
                return None

            # ----------------------------------------------------------
            # MODE A — existing server
            # ----------------------------------------------------------
            if ready_server:
                client_id_int = int(form_vals["client_id"])
                server_id_int = int(form_vals["server_id"])

                client = query(
                    "SELECT schema_name, client_name, client_code "
                    "FROM sam_admin.clients WHERE client_id=%s",
                    (client_id_int,), fetchall=False
                )
                if not client:
                    raise ValueError("Client not found")
                schema = client["schema_name"]

                # Server hardware
                server = query(f"""
                    SELECT sv.server_id, sv.hostname, sv.environment::TEXT AS environment,
                           sv.os_family, sv.datacenter,
                           COALESCE(op.cpu_sockets, 1)           AS cpu_sockets,
                           COALESCE(op.cores_per_socket, 1)      AS cores_per_socket,
                           COALESCE(op.total_physical_cores,
                                    op.cpu_sockets * op.cores_per_socket, 1) AS total_physical_cores,
                           COALESCE(sv.total_ram_mb, 0)          AS ram_mb
                    FROM   {schema}.oracle_servers sv
                    LEFT JOIN {schema}.oracle_processors op ON op.server_id = sv.server_id
                    WHERE  sv.server_id = %s
                """, (server_id_int,), fetchall=False)
                if not server:
                    raise ValueError("Server not found")

                physical_cores = float(server["total_physical_cores"] or 1)

                # Licence requirements for this server from license_position
                reqs = query(f"""
                    SELECT product_family::TEXT                           AS product_family,
                           COALESCE(product_detail, product_family::TEXT) AS product_label,
                           licence_metric::TEXT                           AS metric,
                           CEIL(licences_required)::NUMERIC              AS units_required
                    FROM   {schema}.license_position
                    WHERE  server_id = %s
                      AND  licences_required > 0
                    ORDER  BY product_family, product_detail, licence_metric
                """, (server_id_int,))

                # Current licence assignments with cost context — one row per
                # product/CSI combination.  The LEFT JOIN on entitlement lines is
                # narrowed to a single best-matching line via DISTINCT ON so that
                # a CSI with multiple lines under the same product_family doesn't
                # fan the map row out into duplicates.
                assignments = query(f"""
                    SELECT m.product_family::TEXT                          AS product_family,
                           COALESCE(m.product_detail,
                                    m.product_family::TEXT)               AS product_label,
                           SUM(m.licences_consumed)                       AS licences_consumed,
                           cs.csi_number,
                           cs.contract_name,
                           (SELECT l2.unit_price
                            FROM   shared.license_entitlement_lines l2
                            WHERE  l2.csi_id = m.csi_id
                              AND  l2.product_family::TEXT = m.product_family::TEXT
                              AND  l2.is_active
                            ORDER BY
                              CASE WHEN LOWER(COALESCE(l2.product_name,'')) LIKE
                                        '%%' || LOWER(COALESCE(m.product_detail, m.product_family::TEXT)) || '%%'
                                   THEN 0 ELSE 1 END,
                              l2.unit_price DESC NULLS LAST
                            LIMIT 1)                                       AS list_price_per_unit,
                           (SELECT l2.license_metric::TEXT
                            FROM   shared.license_entitlement_lines l2
                            WHERE  l2.csi_id = m.csi_id
                              AND  l2.product_family::TEXT = m.product_family::TEXT
                              AND  l2.is_active
                            LIMIT 1)                                       AS metric
                    FROM   {schema}.server_csi_map m
                    JOIN   shared.csi_contracts cs ON cs.csi_id = m.csi_id
                    WHERE  m.server_id = %s
                      AND  cs.status = 'active'
                    GROUP  BY m.product_family, m.product_detail, m.csi_id,
                              cs.csi_number, cs.contract_name
                    ORDER  BY m.product_family, m.product_detail
                """, (server_id_int,))

                # Compute assigned_cost and assigned_annual_support per row
                enriched = []
                for a in assignments:
                    consumed  = float(a["licences_consumed"] or 0)
                    unit_price = float(a["list_price_per_unit"]) if a.get("list_price_per_unit") is not None else None
                    assigned_cost = round(consumed * unit_price, 2) if unit_price is not None else None
                    assigned_annual = round(assigned_cost * SUPPORT_RATE, 2) if assigned_cost is not None else None
                    enriched.append({**dict(a),
                                     "assigned_cost":   assigned_cost,
                                     "assigned_annual": assigned_annual})
                assignments = enriched

                # Build per-product-family rollup for the requirements table
                assign_by_family: dict = {}
                for a in assignments:
                    pf = a["product_family"]
                    if pf not in assign_by_family:
                        assign_by_family[pf] = {"consumed": 0.0, "initial": 0.0, "annual": 0.0}
                    assign_by_family[pf]["consumed"] += float(a["licences_consumed"] or 0)
                    if a["assigned_cost"] is not None:
                        assign_by_family[pf]["initial"] += a["assigned_cost"]
                    if a["assigned_annual"] is not None:
                        assign_by_family[pf]["annual"]  += a["assigned_annual"]

                input_label   = server["hostname"]
                input_env     = server["environment"] or ""
                input_sockets = int(server["cpu_sockets"] or 1)
                input_cps     = int(server["cores_per_socket"] or 1)
                input_ram_gb  = round(float(server["ram_mb"] or 0) / 1024, 1)

            # ----------------------------------------------------------
            # MODE B — manual entry
            # ----------------------------------------------------------
            else:
                client        = None
                server        = None
                assignments   = []
                assign_by_family = {}
                physical_cores = float(form_vals["m_cores"])
                edition       = form_vals["m_edition"]
                input_label   = form_vals["m_hostname"] or "Manual Entry"
                input_env     = ""
                input_sockets = 1
                input_cps     = int(physical_cores)
                input_ram_gb  = float(form_vals["m_ram_gb"]) if form_vals["m_ram_gb"] else None

                # Synthetic requirements: base edition + any selected options
                reqs = [{
                    "product_family": edition,
                    "product_label":  edition,
                    "metric":         "processor",
                    "units_required": physical_cores,
                }]
                selected_params = {p for p, _, _, _ in MANUAL_DB_OPTIONS
                                   if request.args.get(p)}
                for param, label, prod_label, metric in MANUAL_DB_OPTIONS:
                    if param in selected_params:
                        reqs.append({
                            "product_family": prod_label,
                            "product_label":  prod_label,
                            "metric":         metric,
                            "units_required": physical_cores,
                        })

            # ----------------------------------------------------------
            # Pool availability (manual mode only)
            # ----------------------------------------------------------
            pool_availability = []   # populated only in manual mode
            pool_allocations  = {}   # product_label -> units allocated from pool

            if not ready_server:
                m_client_id = form_vals.get("m_client_id", "")

                # Fetch all active, non-ULA entitlement lines with their CSI metadata
                pool_rows = query("""
                    SELECT
                        l.line_id,
                        l.csi_id,
                        l.product_name,
                        l.product_family::TEXT AS product_family,
                        l.license_metric::TEXT AS license_metric,
                        l.quantity,
                        l.unit_price,
                        l.annual_support_cost,
                        cs.contract_name,
                        cs.csi_number,
                        cs.sharing_policy::TEXT AS sharing_policy,
                        cs.owning_client_id
                    FROM shared.license_entitlement_lines l
                    JOIN shared.csi_contracts cs ON cs.csi_id = l.csi_id
                    WHERE l.is_active
                      AND cs.status NOT IN ('expired','terminated')
                      AND NOT cs.is_ula
                      AND l.license_metric = 'processor'
                    ORDER BY cs.sharing_policy, l.product_name
                """)

                # Compute consumed counts across all active client schemas
                consumed_map = {}  # (csi_id, product_family) -> total consumed
                active_schemas = query(
                    "SELECT schema_name FROM sam_admin.clients WHERE is_active"
                )
                for sc in active_schemas:
                    s = sc["schema_name"]
                    try:
                        rows = query(f"""
                            SELECT csi_id, product_family::TEXT AS product_family,
                                   SUM(licences_consumed) AS consumed
                            FROM {s}.server_csi_map
                            GROUP BY csi_id, product_family
                        """)
                        for r in rows:
                            key = (r["csi_id"], r["product_family"])
                            consumed_map[key] = consumed_map.get(key, 0) + float(r["consumed"] or 0)
                    except Exception:
                        pass

                for row in pool_rows:
                    qty = float(row["quantity"] or 0)
                    consumed = consumed_map.get((row["csi_id"], row["product_family"]), 0.0)
                    available = max(qty - consumed, 0.0)
                    if available <= 0:
                        continue

                    is_client_locked = row["sharing_policy"] == "client_locked"
                    # Include shared pool always; client-locked only if client matches
                    if is_client_locked and str(row["owning_client_id"]) != str(m_client_id):
                        continue

                    # Read allocation input (carried as alloc_{line_id})
                    alloc_key = f"alloc_{row['line_id']}"
                    raw_alloc = request.args.get(alloc_key, "")
                    allocated = min(float(raw_alloc), available) if raw_alloc else 0.0

                    # Per-unit annual support from the line (total / qty)
                    if row["annual_support_cost"] and qty > 0:
                        support_per_unit = float(row["annual_support_cost"]) / qty
                    elif row["unit_price"]:
                        support_per_unit = float(row["unit_price"]) * SUPPORT_RATE
                    else:
                        support_per_unit = 0.0

                    pool_availability.append({
                        "line_id":          row["line_id"],
                        "csi_id":           row["csi_id"],
                        "product_name":     row["product_name"],
                        "product_family":   row["product_family"],
                        "contract_name":    row["contract_name"],
                        "csi_number":       row["csi_number"],
                        "sharing_policy":   row["sharing_policy"],
                        "quantity":         qty,
                        "consumed":         consumed,
                        "available":        available,
                        "allocated":        allocated,
                        "unit_price":       float(row["unit_price"] or 0),
                        "support_per_unit": support_per_unit,
                        "alloc_key":        alloc_key,
                    })

                # Filter pool to only products actually needed for this analysis.
                # Match on the most distinctive part of the product label — strip
                # the generic "Oracle Database " / "Oracle " prefix before comparing.
                def _distinctive(name):
                    n = name.lower()
                    for prefix in ("oracle database ", "oracle "):
                        if n.startswith(prefix):
                            n = n[len(prefix):]
                            break
                    # normalise plural/singular per word: "diagnostics" == "diagnostic"
                    return " ".join(w.rstrip("s") for w in n.split())

                needed_distinctive = [_distinctive(r["product_label"]) for r in reqs]
                # Products that are never DB server licences — always exclude
                _EXCLUDE_FAMILIES = ("weblogic",)

                def _pool_matches_req(pname):
                    pn_lower = pname.lower()
                    if any(ex in pn_lower for ex in _EXCLUDE_FAMILIES):
                        return False
                    pn = _distinctive(pname)
                    return any(nd in pn or pn in nd for nd in needed_distinctive)

                pool_availability = [pa for pa in pool_availability
                                     if _pool_matches_req(pa["product_name"])]

                # Map each pool entry to the req it belongs to (by name match)
                # so cost model can look up allocations per req label.
                def _matches_label(pname, req_label):
                    pn_lower = pname.lower()
                    if any(ex in pn_lower for ex in _EXCLUDE_FAMILIES):
                        return False
                    pn = _distinctive(pname)
                    nd = _distinctive(req_label)
                    return nd in pn or pn in nd

                # req_label -> (total_allocated, total_annual_pool_support)
                for pa in pool_availability:
                    for r in reqs:
                        if _matches_label(pa["product_name"], r["product_label"]):
                            key = r["product_label"]
                            prev_alloc, prev_supp = pool_allocations.get(key, (0.0, 0.0))
                            pool_allocations[key] = (
                                prev_alloc + pa["allocated"],
                                prev_supp + pa["allocated"] * pa["support_per_unit"],
                            )
                            break

            # ----------------------------------------------------------
            # Cost model — perpetual on-prem
            # ----------------------------------------------------------
            lines = []
            total_licence_cost      = 0.0
            total_licence_cost_raw  = 0.0   # before discount, for display
            total_yr1_support       = 0.0
            total_yr2_annual        = 0.0

            # Build a quick lookup: product_family -> unit_price from assigned CSIs
            assignment_prices: dict = {}
            for a in assignments:
                pf = a.get("product_family")
                up = a.get("list_price_per_unit")
                if pf and up is not None and pf not in assignment_prices:
                    assignment_prices[pf] = float(up)

            # Pool support costs (from already-owned licences allocated from pool)
            pool_existing_support_yr1 = 0.0   # annual support on pool-allocated licences
            pool_existing_support_yr2 = 0.0

            for req in reqs:
                units_req  = float(req["units_required"] or 0)
                # Units already covered by pool allocation (keyed by product_label)
                _pool_entry = pool_allocations.get(req["product_label"], (0.0, 0.0))
                pool_alloc  = min(_pool_entry[0], units_req)
                pool_ann_support = round(_pool_entry[1], 2) if pool_alloc > 0 else 0.0
                units_new  = max(units_req - pool_alloc, 0.0)   # units needing purchase

                # Prefer unit price from assigned CSI lines; fall back to catalogue
                unit_price = (
                    assignment_prices.get(req["product_family"])
                    or find_price(req["product_label"], req["metric"])
                )

                # Cost of new licences only (discount applied to purchase price only)
                _raw_licence  = round(units_new * unit_price, 2) if unit_price is not None else None
                licence_cost  = round(_raw_licence * (1 - licence_discount_pct / 100), 2) if _raw_licence is not None else None
                yr1_sup      = round(licence_cost * SUPPORT_RATE, 2) if licence_cost is not None else None
                yr1_total    = round(licence_cost + yr1_sup, 2) if licence_cost is not None else None

                if licence_cost is not None:
                    total_licence_cost     += licence_cost
                    total_licence_cost_raw += _raw_licence
                    total_yr1_support      += yr1_sup
                    total_yr2_annual       += yr1_sup

                lines.append({
                    "product_label":    req["product_label"],
                    "product_family":   req["product_family"],
                    "metric":           req["metric"],
                    "units_required":   units_req,
                    "pool_alloc":       pool_alloc,
                    "pool_ann_support": pool_ann_support,
                    "units_new":        units_new,
                    "unit_price":       unit_price,
                    "licence_cost":     licence_cost,
                    "yr1_support":      yr1_sup,
                    "yr1_total":        yr1_total,
                    "yr2_annual":       yr1_sup,
                    "price_missing":    unit_price is None and units_new > 0,
                })

            # Support cost on pool-allocated licences (existing owned licences)
            for pa in pool_availability:
                if pa["allocated"] > 0:
                    ann_sup = round(pa["allocated"] * pa["support_per_unit"], 2)
                    pool_existing_support_yr1 += ann_sup
                    pool_existing_support_yr2 += ann_sup

            # Apply vendor quote adjustments to on-prem costs
            adj_onprem = round(onprem_per_core * physical_cores, 2) if onprem_per_core else 0.0
            # Pure licence costs (no platform adjustment) — reused by all BYOL options
            licence_yr1  = round(total_licence_cost + total_yr1_support + pool_existing_support_yr1, 2)
            licence_yr2  = round(total_yr2_annual + pool_existing_support_yr2, 2)
            onprem_yr1   = round(licence_yr1 + adj_onprem, 2)
            onprem_yr2   = round(licence_yr2 + adj_onprem, 2)

            # Year-by-year on-prem cumulative
            yearly_onprem = []
            for y in range(1, horizon + 1):
                ann  = onprem_yr1 if y == 1 else onprem_yr2
                cum  = onprem_yr1 if y == 1 else round(onprem_yr1 + (y - 1) * onprem_yr2, 2)
                yearly_onprem.append({"year": y, "annual": ann, "cum": cum})

            # ----------------------------------------------------------
            # OCI comparison
            # ----------------------------------------------------------
            # Oracle core factor for processor licensing: most x86 = 0.5
            # 1 physical core licensed = 0.5 processor licences on x86
            # OCI: 1 processor licence (perpetual) → 2 OCPUs (BYOL)
            # So total_physical_cores × 0.5 × 2 = total_physical_cores OCPUs
            ocpus = physical_cores  # net result: 1 physical core = 1 OCPU

            oci_skus          = get_oci_prices()
            oci_prices_static = (oci_skus is _OCI_STATIC_SKUS)
            oci_comparison    = build_oci_comparison(oci_skus, ocpus, horizon)
            if oci_comparison:
                oci_comparison["prices_static"] = oci_prices_static
                # Apply vendor quote adjustments to OCI and ExaCC annual costs
                if adj_oci:
                    for key in ("byol_annual", "li_annual"):
                        if oci_comparison.get(key) is not None:
                            oci_comparison[key] = round(oci_comparison[key] + adj_oci, 2)
                adj_exacc = round(managed_per_core * physical_cores, 2) if managed_per_core else 0.0
                if adj_exacc or adj_exacc_upfront:
                    for key in ("exacc_byol_annual", "exacc_li_annual"):
                        if oci_comparison.get(key) is not None:
                            oci_comparison[key] = round(oci_comparison[key] + adj_exacc, 2)
                    oci_comparison["exacc_upfront"] = adj_exacc_upfront

            azure_skus       = get_azure_prices()
            azure_comparison = build_azure_comparison(azure_skus, physical_cores, horizon)
            if azure_comparison and adj_azure:
                for key in ("byol_annual", "li_annual"):
                    if azure_comparison.get(key) is not None:
                        azure_comparison[key] = round(azure_comparison[key] + adj_azure, 2)

            any_missing_price = any(l["price_missing"] for l in lines)

            result = {
                "mode":              mode,
                "input_label":       input_label,
                "input_env":         input_env,
                "input_sockets":     input_sockets,
                "input_cps":         input_cps,
                "input_ram_gb":      input_ram_gb,
                "physical_cores":    physical_cores,
                "ocpus":             ocpus,
                "client":            client,
                "server":            server,
                "assignments":       assignments,
                "assign_by_family":  assign_by_family if ready_server else {},
                "lines":             lines,
                "support_rate_pct":  22,
                "total_licence_cost": round(total_licence_cost, 2),
                "licence_yr1":       licence_yr1,
                "licence_yr2":       licence_yr2,
                "onprem_yr1":        onprem_yr1,
                "onprem_yr2":        onprem_yr2,
                "yearly_onprem":     yearly_onprem,
                "horizon":           horizon,
                "any_missing_price": any_missing_price,
                "pool_availability": pool_availability,
                "pool_existing_support_yr1": round(pool_existing_support_yr1, 2),
                "pool_existing_support_yr2": round(pool_existing_support_yr2, 2),
                "new_licence_cost":  round(total_licence_cost, 2),
                "new_licence_support_yr1": round(total_yr1_support, 2),
                "licence_discount_pct": licence_discount_pct,
                "licence_discount_saving": round(total_licence_cost_raw - total_licence_cost, 2),
                "adjustments": {
                    "onprem":         adj_onprem,
                    "oci":            adj_oci,
                    "exacc":          adj_exacc,
                    "azure":          adj_azure,
                    "onprem_upfront": adj_onprem_upfront,
                    "exacc_upfront":  adj_exacc_upfront,
                    "any":            adj_any,
                },
                "oci":               oci_comparison,
                "azure":             azure_comparison,
            }

        except Exception as e:
            result = {"error": str(e)}

    return render_template(
        "licence_analysis.html",
        clients=clients,
        form_vals=form_vals,
        server_list=server_list,
        result=result,
        manual_db_options=MANUAL_DB_OPTIONS,
    )


@app.route("/contracts")
@login_required
def contracts():
    # One row per CSI header
    csi_rows = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name,
               cs.support_expiry, cs.sharing_policy, cs.status, cs.currency,
               cs.is_ula,
               oc.client_code AS owning_client
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        ORDER BY cs.csi_number
    """)

    # All active entitlement lines
    line_rows = query("""
        SELECT l.line_id, l.csi_id, l.product_name,
               l.product_family::TEXT AS product_family,
               l.license_metric::TEXT AS license_metric,
               l.quantity, l.unit_price, l.total_price
        FROM shared.license_entitlement_lines l
        WHERE l.is_active
        ORDER BY l.csi_id, l.line_number
    """)

    # Sum licences_consumed per (csi_id, product_detail) across all active schemas
    active_schemas = [
        r["schema_name"] for r in query(
            "SELECT schema_name FROM sam_admin.clients WHERE is_active ORDER BY schema_name"
        )
    ]
    consumed_entries_by_csi = {}   # csi_id -> [(product_family, product_detail, amount)]
    if active_schemas:
        union_sql = " UNION ALL ".join(
            f"SELECT csi_id, product_family::TEXT, COALESCE(product_detail,'') AS product_detail, "
            f"COALESCE(SUM(licences_consumed), 0) AS consumed "
            f"FROM {s}.server_csi_map GROUP BY csi_id, product_family, product_detail"
            for s in active_schemas
        )
        for r in query(
            f"SELECT csi_id, product_family, product_detail, SUM(consumed) AS total "
            f"FROM ({union_sql}) t GROUP BY csi_id, product_family, product_detail"
        ):
            consumed_entries_by_csi.setdefault(r["csi_id"], []).append(
                (r["product_family"], r["product_detail"], r["total"])
            )

    # Group lines under each CSI, attaching consumed/available per line
    # Match each line to assignments using product compatibility logic
    lines_by_csi = {}
    for l in line_rows:
        consumed = sum(
            amt for fam, det, amt in consumed_entries_by_csi.get(l["csi_id"], [])
            if _is_compatible_product(l["product_family"], det or None,
                                      l["product_family"], l["product_name"])
        )
        qty = l["quantity"] or 0
        line = dict(l,
                    consumed=consumed,
                    available=max(qty - consumed, 0))
        lines_by_csi.setdefault(l["csi_id"], []).append(line)

    contracts = [dict(r, lines=lines_by_csi.get(r["csi_id"], [])) for r in csi_rows]

    # Values for filter dropdowns
    all_clients = query(
        "SELECT client_code, client_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    sharing_policies = sorted({r["sharing_policy"] for r in csi_rows})

    return render_template("contracts.html", contracts=contracts,
                           all_clients=all_clients,
                           sharing_policies=sharing_policies)


@app.route("/contracts/<int:csi_id>")
@login_required
def contract_detail(csi_id):
    schema = get_schema()
    contract = query(
        "SELECT * FROM shared.csi_contract_summary WHERE csi_id = %s",
        (csi_id,), fetchall=False
    )
    if not contract:
        flash("Contract not found.", "danger")
        return redirect(url_for("contracts"))

    lines = query(
        "SELECT * FROM shared.license_entitlement_lines WHERE csi_id = %s ORDER BY line_number",
        (csi_id,)
    )
    # Gather assigned servers across all active client schemas
    all_schemas = query(
        "SELECT schema_name, client_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    assigned_servers = []
    for c in all_schemas:
        s = c["schema_name"]
        try:
            rows = query(f"""
                SELECT s.server_id, s.hostname, s.environment::TEXT,
                       m.product_family, m.product_detail,
                       m.licences_consumed, m.notes, m.effective_date, m.map_id,
                       %s AS client_name, %s AS schema_name
                FROM {s}.server_csi_map m
                JOIN {s}.oracle_servers s ON s.server_id = m.server_id
                WHERE m.csi_id = %s
                ORDER BY s.hostname
            """, (c["client_name"], s, csi_id))
            assigned_servers.extend(rows)
        except Exception:
            pass

    # Sum consumed per product_family across all schemas
    consumed_by_line = {}
    for row in assigned_servers:
        key = row["product_family"]
        consumed_by_line[key] = consumed_by_line.get(key, 0) + (row["licences_consumed"] or 0)

    # ULA extras
    ula_products = []
    ula_annual_support = None
    ula_available_servers = []
    ula_assigned_server_ids = set()
    if contract.get("is_ula"):
        ula_products = [r["product_name"] for r in query(
            "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s ORDER BY product_name",
            (csi_id,)
        )]

        # Sync entitlement lines from ula_covered_products so lines always reflect
        # what was selected, even for ULAs created before this was enforced.
        _ULA_PRODUCT_FAMILY = {
            "Enterprise Edition":        "oracle_database",
            "Standard Edition 2":        "oracle_database",
            "Tuning Pack":               "oracle_database",
            "Diagnostics Pack":          "oracle_database",
            "Real Application Clusters": "oracle_database",
            "Partitioning":              "oracle_database",
            "Advanced Security":         "oracle_database",
            "Label Security":            "oracle_database",
            "Database Vault":            "oracle_database",
            "OLAP":                      "oracle_database",
            "Spatial and Graph":         "oracle_database",
            "Active Data Guard":         "oracle_database",
            "Multitenant":               "oracle_database",
            "GoldenGate":                "oracle_database",
            "WebLogic Server":           "oracle_weblogic",
            "WebLogic Suite":            "oracle_weblogic",
            "Coherence":                 "oracle_coherence",
            "Java SE":                   "oracle_java",
            "Java SE Subscription":      "oracle_java",
        }
        existing_line_names = {r["product_name"] for r in lines}
        for line_no, p in enumerate(ula_products, start=1):
            if p not in existing_line_names:
                family = _ULA_PRODUCT_FAMILY.get(p, "oracle_database")
                execute(
                    "INSERT INTO shared.license_entitlement_lines "
                    "(csi_id, line_number, product_name, product_family, license_metric, quantity) "
                    "VALUES (%s, %s, %s, %s::shared.product_family, 'processor', 0) "
                    "ON CONFLICT (csi_id, line_number) DO NOTHING",
                    (csi_id, line_no, p, family)
                )
        # Reload lines after potential sync
        lines = query(
            "SELECT * FROM shared.license_entitlement_lines WHERE csi_id = %s ORDER BY line_number",
            (csi_id,)
        )

        # Get annual support cost (stored on any line that has it)
        support_row = query(
            "SELECT annual_support_cost FROM shared.license_entitlement_lines "
            "WHERE csi_id = %s AND annual_support_cost IS NOT NULL LIMIT 1",
            (csi_id,), fetchall=False
        )
        if support_row:
            ula_annual_support = support_row["annual_support_cost"]

        # Owning client schema for server list
        owner = query(
            "SELECT schema_name FROM sam_admin.clients WHERE client_id = %s",
            (contract["owning_client_id"],), fetchall=False
        ) if contract.get("owning_client_id") else None
        if owner:
            s = owner["schema_name"]
            ula_available_servers = query(
                f"SELECT server_id, hostname, environment::TEXT AS environment "
                f"FROM {s}.oracle_servers WHERE is_active ORDER BY hostname"
            )
            already = query(
                f"SELECT DISTINCT server_id FROM {s}.server_csi_map WHERE csi_id = %s",
                (csi_id,)
            )
            ula_assigned_server_ids = {r["server_id"] for r in already}

    return render_template("contract_detail.html",
                           contract=contract, lines=lines,
                           assigned_servers=assigned_servers,
                           consumed_by_line=consumed_by_line,
                           ula_products=ula_products,
                           ula_annual_support=ula_annual_support,
                           ula_available_servers=ula_available_servers,
                           ula_assigned_server_ids=ula_assigned_server_ids)


@app.route("/contracts/<int:csi_id>/lines/<int:line_id>/edit", methods=["POST"])
@login_required
def edit_entitlement_line(csi_id, line_id):
    """Update unit_price and/or annual_support_cost on a single entitlement line."""
    data = request.get_json(force=True, silent=True) or {}
    updates = {}
    errors = {}

    for field in ("unit_price", "annual_support_cost"):
        if field in data:
            raw = data[field]
            if raw == "" or raw is None:
                updates[field] = None
            else:
                try:
                    updates[field] = float(raw)
                    if updates[field] < 0:
                        raise ValueError
                except (ValueError, TypeError):
                    errors[field] = f"Invalid value for {field}"

    if errors:
        return jsonify({"ok": False, "errors": errors}), 400

    if not updates:
        return jsonify({"ok": True, "message": "Nothing to update"})

    set_clause = ", ".join(f"{k} = %s" for k in updates)
    values = list(updates.values()) + [line_id, csi_id]
    execute(
        f"UPDATE shared.license_entitlement_lines SET {set_clause} "
        f"WHERE line_id = %s AND csi_id = %s",
        values
    )

    # Return the refreshed row so the UI can update derived fields
    row = query(
        "SELECT line_id, unit_price, total_price, annual_support_cost, quantity "
        "FROM shared.license_entitlement_lines WHERE line_id = %s",
        (line_id,), fetchall=False
    )
    return jsonify({
        "ok": True,
        "unit_price":          float(row["unit_price"]) if row["unit_price"] is not None else None,
        "total_price":         float(row["total_price"]) if row["total_price"] is not None else None,
        "annual_support_cost": float(row["annual_support_cost"]) if row["annual_support_cost"] is not None else None,
        "quantity":            float(row["quantity"]) if row["quantity"] else None,
    })


@app.route("/contracts/<int:csi_id>/delete", methods=["POST"])
@login_required
def delete_contract(csi_id):
    # Remove server_csi_map rows in every client schema first (no cascade)
    all_schemas = query(
        "SELECT schema_name FROM sam_admin.clients WHERE is_active"
    )
    for c in all_schemas:
        s = c["schema_name"]
        try:
            execute(f"DELETE FROM {s}.server_csi_map WHERE csi_id = %s", (csi_id,))
        except Exception:
            pass

    # Also clean ula_certifications if referenced (no cascade defined)
    try:
        execute("DELETE FROM shared.ula_certifications WHERE csi_id = %s", (csi_id,))
    except Exception:
        pass

    # Delete the contract — cascades to lines, client_map, ula_covered_products
    execute("DELETE FROM shared.csi_contracts WHERE csi_id = %s", (csi_id,))
    flash("Contract deleted.", "success")
    return redirect(url_for("contracts"))


@app.route("/contracts/<int:csi_id>/notes", methods=["POST"])
@login_required
def contract_save_notes(csi_id):
    notes = request.form.get("notes", "").strip() or None
    execute(
        "UPDATE shared.csi_contracts SET notes = %s WHERE csi_id = %s",
        (notes, csi_id)
    )
    flash("Notes saved.", "success")
    return redirect(url_for("contract_detail", csi_id=csi_id))


@app.route("/contracts/<int:csi_id>/references", methods=["POST"])
@login_required
def contract_save_references(csi_id):
    br_number  = request.form.get("br_number",  "").strip() or None
    p2p_number = request.form.get("p2p_number", "").strip() or None
    old = query(
        "SELECT br_number, p2p_number FROM shared.csi_contracts WHERE csi_id = %s",
        (csi_id,), fetchall=False
    )
    execute(
        "UPDATE shared.csi_contracts SET br_number = %s, p2p_number = %s WHERE csi_id = %s",
        (br_number, p2p_number, csi_id)
    )
    _audit("contract.update_references", entity_type="contract", entity_id=csi_id,
           old_values={"br_number": old["br_number"] if old else None,
                       "p2p_number": old["p2p_number"] if old else None},
           new_values={"br_number": br_number, "p2p_number": p2p_number})
    flash("Reference numbers saved.", "success")
    return redirect(url_for("contract_detail", csi_id=csi_id))


@app.route("/contracts/<int:csi_id>/ula/products", methods=["POST"])
@login_required
def ula_save_products(csi_id):
    """Replace the covered-product list for a ULA contract."""
    products = [p.strip() for p in request.form.getlist("products") if p.strip()]
    execute("DELETE FROM shared.ula_covered_products WHERE csi_id = %s", (csi_id,))
    for p in products:
        execute(
            "INSERT INTO shared.ula_covered_products (csi_id, product_name) VALUES (%s, %s) ON CONFLICT DO NOTHING",
            (csi_id, p)
        )
    flash("ULA covered products updated.", "success")
    return redirect(url_for("contract_detail", csi_id=csi_id))


@app.route("/contracts/<int:csi_id>/ula/assign", methods=["POST"])
@login_required
def ula_assign_server(csi_id):
    """Assign a server to a ULA, validating that all its licence needs are covered."""
    server_id = request.form.get("server_id", type=int)
    if not server_id:
        flash("No server selected.", "danger")
        return redirect(url_for("contract_detail", csi_id=csi_id))

    # Resolve owning client schema
    contract = query(
        "SELECT cs.is_ula, c.schema_name, c.client_name "
        "FROM shared.csi_contracts cs "
        "JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id "
        "WHERE cs.csi_id = %s",
        (csi_id,), fetchall=False
    )
    if not contract or not contract["is_ula"]:
        flash("Contract is not a ULA.", "danger")
        return redirect(url_for("contract_detail", csi_id=csi_id))

    schema = contract["schema_name"]

    # Get what products the server needs
    needed = query(
        f"SELECT DISTINCT COALESCE(product_detail, product_family::TEXT) AS product "
        f"FROM {schema}.license_position WHERE server_id = %s",
        (server_id,)
    )
    needed_products = {r["product"] for r in needed}

    # Get what the ULA covers
    covered = query(
        "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s",
        (csi_id,)
    )
    covered_products = {r["product_name"] for r in covered}

    # Validate
    out_of_scope = needed_products - covered_products
    if out_of_scope:
        missing = ", ".join(sorted(out_of_scope))
        flash(
            f"Cannot assign server — the following licence requirements are outside this ULA's scope: {missing}",
            "danger"
        )
        return redirect(url_for("contract_detail", csi_id=csi_id))

    # Insert one server_csi_map row per covered product the server actually needs
    # (licences_consumed = NULL = unlimited under ULA)
    assigned_any = False
    for prod in needed_products:
        # Find product_family for this product_detail
        lp_row = query(
            f"SELECT product_family FROM {schema}.license_position "
            f"WHERE server_id = %s AND COALESCE(product_detail, product_family::TEXT) = %s LIMIT 1",
            (server_id, prod), fetchall=False
        )
        family = lp_row["product_family"] if lp_row else "oracle_database"
        try:
            exists = query(
                f"SELECT 1 FROM {schema}.server_csi_map "
                f"WHERE server_id = %s AND csi_id = %s AND product_family = %s::shared.product_family",
                (server_id, csi_id, family), fetchall=False
            )
            if not exists:
                execute(
                    f"INSERT INTO {schema}.server_csi_map "
                    f"(server_id, csi_id, product_family, product_detail, licences_consumed, notes, assigned_by) "
                    f"VALUES (%s, %s, %s::shared.product_family, %s, NULL, 'ULA', %s)",
                    (server_id, csi_id, family, prod, ADMIN_USER)
                )
                assigned_any = True
        except Exception as e:
            app.logger.warning("ULA assign insert failed: %s", e)

    if assigned_any:
        flash("Server assigned to ULA successfully.", "success")
    else:
        flash("Server was already fully assigned to this ULA.", "info")

    return redirect(url_for("contract_detail", csi_id=csi_id))


@app.route("/contracts/<int:csi_id>/ula/support", methods=["GET", "POST"])
@login_required
def ula_support_cost(csi_id):
    contract = query(
        "SELECT csi_id, contract_name, csi_number, currency FROM shared.csi_contracts WHERE csi_id = %s",
        (csi_id,), fetchall=False
    )
    if not contract:
        abort(404)
    ula_products = [r["product_name"] for r in query(
        "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s ORDER BY product_name",
        (csi_id,)
    )]
    if request.method == "POST":
        annual_cost = request.form.get("annual_support_cost", "").strip() or None
        if annual_cost:
            try:
                annual_cost = float(annual_cost.replace(",", ""))
            except ValueError:
                annual_cost = None
        # Store on the first entitlement line (all products share the ULA cost)
        execute(
            "UPDATE shared.license_entitlement_lines SET annual_support_cost = %s "
            "WHERE csi_id = %s AND line_number = 1",
            (annual_cost, csi_id)
        )
        flash("ULA contract saved.", "success")
        return redirect(url_for("contract_detail", csi_id=csi_id))
    return render_template("ula_support_cost.html", contract=contract, ula_products=ula_products)


@app.route("/contracts/<int:csi_id>/ula/remove", methods=["POST"])
@login_required
def ula_remove_server(csi_id):
    """Remove all ULA assignments for a server from this contract."""
    server_id = request.form.get("server_id", type=int)
    contract = query(
        "SELECT c.schema_name FROM shared.csi_contracts cs "
        "JOIN sam_admin.clients c ON c.client_id = cs.owning_client_id "
        "WHERE cs.csi_id = %s",
        (csi_id,), fetchall=False
    )
    if contract:
        schema = contract["schema_name"]
        execute(
            f"DELETE FROM {schema}.server_csi_map WHERE csi_id = %s AND server_id = %s",
            (csi_id, server_id)
        )
    flash("Server removed from ULA.", "success")
    return redirect(url_for("contract_detail", csi_id=csi_id))


# ---------------------------------------------------------------------------
# Add contract
# ---------------------------------------------------------------------------
@app.route("/contracts/new", methods=["GET", "POST"])
@login_required
def add_contract():
    all_clients = query(
        "SELECT client_id, client_code, client_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )

    if request.method == "POST":
        action = request.form.get("action")

        if action == "save_header":
            # ── Step 1: insert contract header ──────────────────────────────
            csi_number      = request.form.get("csi_number", "").strip() or None
            contract_name   = request.form.get("contract_name", "").strip()
            vendor_ref      = request.form.get("vendor_reference", "").strip() or None
            br_number       = request.form.get("br_number", "").strip() or None
            p2p_number      = request.form.get("p2p_number", "").strip() or None
            purchase_date   = request.form.get("purchase_date") or None
            support_start   = request.form.get("support_start") or None
            support_expiry  = request.form.get("support_expiry") or None
            ula_expiry      = request.form.get("ula_expiry") or None
            is_ula          = request.form.get("is_ula") == "1"
            currency        = request.form.get("currency", "USD").strip().upper() or "USD"
            sharing_policy  = request.form.get("sharing_policy", "unassigned")
            locked_client   = request.form.get("locked_client") or None
            notes           = request.form.get("notes", "").strip() or None
            status          = request.form.get("status", "active")
            ula_products    = [p.strip() for p in request.form.getlist("ula_products") if p.strip()]

            if not contract_name:
                flash("Contract name is required.", "danger")
                return render_template("add_contract.html", all_clients=all_clients, step=1)

            # Resolve owning_client_id
            owning_client_id = None
            if sharing_policy == "client_locked" and locked_client:
                row = query(
                    "SELECT client_id FROM sam_admin.clients WHERE client_code = %s",
                    (locked_client,), fetchall=False
                )
                if row:
                    owning_client_id = row["client_id"]

            try:
                new_row = query("""
                    INSERT INTO shared.csi_contracts
                      (csi_number, contract_name, vendor_reference, br_number, p2p_number,
                       purchase_date, support_start, support_expiry,
                       ula_expiry, is_ula,
                       currency, sharing_policy, owning_client_id, notes, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::shared.sharing_policy,%s,%s,%s::shared.license_status)
                    RETURNING csi_id
                """, (csi_number, contract_name, vendor_ref, br_number, p2p_number,
                      purchase_date, support_start, support_expiry,
                      ula_expiry, is_ula,
                      currency, sharing_policy, owning_client_id, notes, status),
                    fetchall=False)
                csi_id = new_row["csi_id"]

                # Save ULA covered products and auto-create entitlement lines, then go to support cost step
                if is_ula:
                    _ULA_PRODUCT_FAMILY = {
                        "Enterprise Edition":        "oracle_database",
                        "Standard Edition 2":        "oracle_database",
                        "Tuning Pack":               "oracle_database",
                        "Diagnostics Pack":          "oracle_database",
                        "Real Application Clusters": "oracle_database",
                        "Partitioning":              "oracle_database",
                        "Advanced Security":         "oracle_database",
                        "Label Security":            "oracle_database",
                        "Database Vault":            "oracle_database",
                        "OLAP":                      "oracle_database",
                        "Spatial and Graph":         "oracle_database",
                        "Active Data Guard":         "oracle_database",
                        "Multitenant":               "oracle_database",
                        "GoldenGate":                "oracle_database",
                        "WebLogic Server":           "oracle_weblogic",
                        "WebLogic Suite":            "oracle_weblogic",
                        "Coherence":                 "oracle_coherence",
                        "Java SE":                   "oracle_java",
                        "Java SE Subscription":      "oracle_java",
                    }
                    for line_no, p in enumerate(ula_products, start=1):
                        family = _ULA_PRODUCT_FAMILY.get(p, "oracle_database")
                        execute(
                            "INSERT INTO shared.ula_covered_products (csi_id, product_name) "
                            "VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            (csi_id, p)
                        )
                        execute(
                            "INSERT INTO shared.license_entitlement_lines "
                            "(csi_id, line_number, product_name, product_family, license_metric, quantity) "
                            "VALUES (%s, %s, %s, %s::shared.product_family, 'processor', 0)",
                            (csi_id, line_no, p, family)
                        )
                    return redirect(url_for("ula_support_cost", csi_id=csi_id))

                # If shareable and a client was selected, assign immediately
                if sharing_policy == "shareable" and locked_client:
                    row = query(
                        "SELECT client_id FROM sam_admin.clients WHERE client_code = %s",
                        (locked_client,), fetchall=False
                    )
                    if row:
                        execute(
                            "INSERT INTO shared.csi_client_map (csi_id, client_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            (csi_id, row["client_id"])
                        )

                flash("Contract created. Now add entitlement lines.", "success")
                return redirect(url_for("add_contract_lines", csi_id=csi_id))
            except Exception as e:
                flash(f"Error creating contract: {e}", "danger")
                return render_template("add_contract.html", all_clients=all_clients, step=1)

    return render_template("add_contract.html", all_clients=all_clients, step=1)


@app.route("/contracts/<int:csi_id>/lines", methods=["GET", "POST"])
@login_required
def add_contract_lines(csi_id):
    contract = query(
        "SELECT csi_id, contract_name, csi_number, is_ula FROM shared.csi_contracts WHERE csi_id = %s",
        (csi_id,), fetchall=False
    )
    if not contract:
        flash("Contract not found.", "danger")
        return redirect(url_for("contracts"))

    # ULA contracts use a different flow
    if contract.get("is_ula"):
        return redirect(url_for("ula_support_cost", csi_id=csi_id))

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_line":
            product_name   = request.form.get("product_name", "").strip()
            product_family = request.form.get("product_family", "oracle_database")
            metric         = request.form.get("license_metric", "processor")
            quantity       = request.form.get("quantity", "0").strip()
            unit_price     = request.form.get("unit_price", "").strip() or None
            annual_support = request.form.get("annual_support_cost", "").strip() or None

            if not product_name or not quantity:
                flash("Product name and quantity are required.", "danger")
            else:
                # Next line number
                last = query(
                    "SELECT COALESCE(MAX(line_number),0) AS mx FROM shared.license_entitlement_lines WHERE csi_id=%s",
                    (csi_id,), fetchall=False
                )
                next_line = (last["mx"] or 0) + 1
                try:
                    execute("""
                        INSERT INTO shared.license_entitlement_lines
                          (csi_id, line_number, product_name, product_family,
                           license_metric, quantity, unit_price, annual_support_cost)
                        VALUES (%s,%s,%s,%s::shared.product_family,%s::shared.license_metric,%s,%s,%s)
                    """, (csi_id, next_line, product_name, product_family,
                          metric, quantity, unit_price, annual_support))
                    flash(f"Line added: {product_name}.", "success")
                except Exception as e:
                    flash(f"Error adding line: {e}", "danger")

        elif action == "done":
            flash("Contract saved successfully.", "success")
            return redirect(url_for("contract_detail", csi_id=csi_id))

    lines = query(
        "SELECT * FROM shared.license_entitlement_lines WHERE csi_id=%s ORDER BY line_number",
        (csi_id,)
    )
    return render_template("add_contract_lines.html", contract=contract, lines=lines)


# ---------------------------------------------------------------------------
# Visibility — Versions
# ---------------------------------------------------------------------------

# Oracle DB version → Premier/Extended support end dates (yyyy-mm-dd)
# Source: Oracle Lifetime Support Policy
_WLS_SUPPORT = {
    "14.1": {"premier": "2027-01-31", "extended": "2030-01-31"},
    "12.2": {"premier": "2022-03-31", "extended": "2025-12-31"},
    "12.1": {"premier": "2019-12-31", "extended": "2023-12-31"},
    "10.3": {"premier": "2016-12-31", "extended": "2021-12-31"},
}

def _wls_support_status(wls_version):
    """Return ('supported'|'extended'|'warning'|'eol', end_date_str) for a WLS version string."""
    from datetime import date, timedelta
    if not wls_version:
        return "unknown", None
    v = (wls_version or "").strip()
    matched = None
    for key in sorted(_WLS_SUPPORT, key=len, reverse=True):
        if v.startswith(key):
            matched = _WLS_SUPPORT[key]
            break
    if not matched:
        return "unknown", None
    today = date.today()
    premier_end  = date.fromisoformat(matched["premier"])
    extended_end = date.fromisoformat(matched["extended"])
    warning_threshold = extended_end - timedelta(days=365)
    if today > extended_end:
        return "eol", matched["extended"]
    elif today >= warning_threshold:
        return "warning", matched["extended"]
    elif today > premier_end:
        return "extended", matched["extended"]
    return "supported", matched["premier"]

_ORACLE_DB_SUPPORT = {
    "23":   {"premier": "2028-04-30", "extended": "2031-04-30"},
    "21":   {"premier": "2024-04-30", "extended": "2027-04-30"},
    "19":   {"premier": "2029-12-31", "extended": "2032-12-31"},
    "18":   {"premier": "2021-06-30", "extended": "2021-06-30"},
    "12.2": {"premier": "2020-11-30", "extended": "2023-03-31"},
    "12.1": {"premier": "2018-07-31", "extended": "2022-07-31"},
    "11.2": {"premier": "2015-01-31", "extended": "2020-12-31"},
    "11.1": {"premier": "2010-08-31", "extended": "2013-08-31"},
    "10.2": {"premier": "2010-07-31", "extended": "2013-07-31"},
}

def _version_support_status(db_version):
    """Return ('supported'|'extended'|'warning'|'eol', end_date_str) for a db_version string."""
    from datetime import date, timedelta
    if not db_version:
        return "unknown", None
    v = db_version.strip()
    matched = None
    for key in sorted(_ORACLE_DB_SUPPORT, key=len, reverse=True):
        if v.startswith(key):
            matched = _ORACLE_DB_SUPPORT[key]
            break
    if not matched:
        return "unknown", None
    today = date.today()
    premier_end  = date.fromisoformat(matched["premier"])
    extended_end = date.fromisoformat(matched["extended"])
    warning_threshold = extended_end - timedelta(days=365)
    if today > extended_end:
        return "eol", matched["extended"]
    elif today >= warning_threshold:
        return "warning", matched["extended"]
    elif today > premier_end:
        return "extended", matched["extended"]
    return "supported", matched["premier"]

@app.route("/visibility/versions")
@login_required
def visibility_versions():
    active_clients = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    clients_data = []
    for c in active_clients:
        s = c["schema_name"]

        # ── Database servers: count per version ──────────────────────────────
        try:
            db_rows = query(f"""
                SELECT COALESCE(i.db_version, 'Unknown') AS db_version,
                       COUNT(DISTINCT srv.server_id) AS server_count
                FROM {s}.oracle_servers srv
                JOIN {s}.oracle_instances i ON i.server_id = srv.server_id AND i.is_active
                WHERE srv.is_active
                GROUP BY i.db_version
            """)
        except Exception as e:
            app.logger.error("visibility_versions db query failed for %s: %s", s, e)
            db_rows = []

        db_status_counts = {"supported": 0, "extended": 0, "warning": 0, "eol": 0, "unknown": 0}
        for r in db_rows:
            ver = r["db_version"] or "Unknown"
            cnt = int(r["server_count"] or 0)
            status, _ = _version_support_status(ver)
            db_status_counts[status] = db_status_counts.get(status, 0) + cnt

        # ── Middleware (WLS) servers: count per version ──────────────────────
        try:
            wls_rows = query(f"""
                SELECT COALESCE(d.wls_version, 'Unknown') AS wls_version,
                       COUNT(DISTINCT srv.server_id) AS server_count
                FROM {s}.oracle_servers srv
                JOIN {s}.wls_domains d ON d.server_id = srv.server_id AND d.is_active
                WHERE srv.is_active
                GROUP BY d.wls_version
            """)
        except Exception as e:
            app.logger.error("visibility_versions wls query failed for %s: %s", s, e)
            wls_rows = []

        mw_status_counts = {"supported": 0, "extended": 0, "warning": 0, "eol": 0, "unknown": 0}
        for r in wls_rows:
            ver = r["wls_version"] or "Unknown"
            cnt = int(r["server_count"] or 0)
            status, _ = _wls_support_status(ver)
            mw_status_counts[status] = mw_status_counts.get(status, 0) + cnt

        db_total = sum(db_status_counts.values())
        mw_total = sum(mw_status_counts.values())
        if db_total == 0 and mw_total == 0:
            continue

        clients_data.append({
            "client_name":      c["client_name"],
            "client_code":      c["client_code"],
            "db_status_counts": db_status_counts,
            "mw_status_counts": mw_status_counts,
            "db_total":         db_total,
            "mw_total":         mw_total,
        })
    return render_template("visibility_versions.html", clients=clients_data)


@app.route("/visibility/versions/<client_code>")
@login_required
def visibility_versions_client(client_code):
    rows_c = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE client_code = %s AND is_active",
        (client_code,)
    )
    if not rows_c:
        abort(404)
    client = rows_c[0]
    s = client["schema_name"]

    # ── Database servers ─────────────────────────────────────────────────────
    db_rows = query(f"""
        SELECT srv.hostname,
               COALESCE(srv.environment::TEXT, '') AS environment,
               COALESCE(srv.datacenter, '')         AS datacenter,
               COALESCE(i.db_version, 'Unknown')    AS db_version
        FROM {s}.oracle_servers srv
        LEFT JOIN {s}.oracle_instances i
               ON i.server_id = srv.server_id AND i.is_active
        WHERE srv.is_active
        ORDER BY srv.hostname
    """)
    db_servers = []
    for r in db_rows:
        ver = r["db_version"] or "Unknown"
        status, end_date = _version_support_status(ver)
        db_servers.append({
            "hostname":    r["hostname"],
            "environment": r["environment"],
            "datacenter":  r["datacenter"],
            "version":     ver,
            "status":      status,
            "end_date":    end_date,
        })

    # ── Middleware (WLS) servers ─────────────────────────────────────────────
    try:
        wls_rows = query(f"""
            SELECT DISTINCT ON (srv.server_id)
                   srv.hostname,
                   COALESCE(srv.environment::TEXT, '') AS environment,
                   COALESCE(srv.datacenter, '')         AS datacenter,
                   COALESCE(d.wls_version, 'Unknown')   AS wls_version,
                   d.wls_edition
            FROM {s}.oracle_servers srv
            JOIN {s}.wls_domains d ON d.server_id = srv.server_id AND d.is_active
            WHERE srv.is_active
            ORDER BY srv.server_id, d.domain_name
        """)
    except Exception:
        wls_rows = []

    mw_servers = []
    for r in wls_rows:
        ver = r["wls_version"] or "Unknown"
        status, end_date = _wls_support_status(ver)
        mw_servers.append({
            "hostname":    r["hostname"],
            "environment": r["environment"],
            "datacenter":  r["datacenter"],
            "version":     ver,
            "edition":     r.get("wls_edition") or "",
            "status":      status,
            "end_date":    end_date,
        })

    def _status_counts(srv_list):
        c = {"supported": 0, "extended": 0, "warning": 0, "eol": 0, "unknown": 0}
        for sv in srv_list:
            c[sv["status"]] = c.get(sv["status"], 0) + 1
        return c

    return render_template(
        "visibility_versions_client.html",
        client=client,
        db_servers=db_servers,
        mw_servers=mw_servers,
        db_counts=_status_counts(db_servers),
        mw_counts=_status_counts(mw_servers),
    )


# ---------------------------------------------------------------------------
# Visibility — Licence history (snapshot trend charts)
# ---------------------------------------------------------------------------
@app.route("/visibility/licence-history")
@login_required
def visibility_licence_history():
    """Client picker page — redirects to client-specific view."""
    clients_with_snaps = query("""
        SELECT c.client_id, c.client_code, c.client_name,
               COUNT(s.snapshot_id)      AS snapshot_count,
               MIN(s.snapshot_month)     AS earliest,
               MAX(s.snapshot_month)     AS latest
        FROM sam_admin.clients c
        JOIN sam_admin.licence_snapshots s ON s.client_id = c.client_id
        GROUP BY c.client_id, c.client_code, c.client_name
        ORDER BY c.client_name
    """)
    return render_template("visibility_licence_history.html",
                           clients=clients_with_snaps, selected=None,
                           chart_data=None)


@app.route("/visibility/licence-history/<client_code>")
@login_required
def visibility_licence_history_client(client_code):
    client = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE client_code = %s AND is_active",
        (client_code,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("visibility_licence_history"))

    # Enforce client-scope users can only see their own client
    u = current_user()
    if u.get("role") == "client" and u.get("client_code") != client_code:
        abort(403)

    # All clients list for the switcher
    clients_with_snaps = query("""
        SELECT c.client_id, c.client_code, c.client_name,
               COUNT(s.snapshot_id) AS snapshot_count
        FROM sam_admin.clients c
        JOIN sam_admin.licence_snapshots s ON s.client_id = c.client_id
        GROUP BY c.client_id, c.client_code, c.client_name
        ORDER BY c.client_name
    """)

    # Monthly aggregates
    monthly = query("""
        SELECT s.snapshot_month,
               COALESCE(SUM(l.licences_required), 0)  AS total_required,
               COALESCE(SUM(l.licences_assigned), 0)  AS total_assigned,
               COUNT(*) FILTER (WHERE l.compliance_status = 'under_licensed') AS under_count,
               COUNT(*) FILTER (WHERE l.compliance_status = 'compliant')      AS compliant_count,
               COUNT(*) FILTER (WHERE l.compliance_status = 'over_licensed')  AS over_count,
               COUNT(DISTINCT l.hostname)                                      AS server_count
        FROM sam_admin.licence_snapshots s
        JOIN sam_admin.licence_snapshot_lines l ON l.snapshot_id = s.snapshot_id
        WHERE s.client_id = %s
        GROUP BY s.snapshot_month
        ORDER BY s.snapshot_month
    """, (client["client_id"],))

    # Per-product-family monthly required (capped: DB + WLS only to stay <= 4 series)
    family_monthly = query("""
        SELECT s.snapshot_month, l.product_family,
               COALESCE(SUM(l.licences_required), 0) AS required
        FROM sam_admin.licence_snapshots s
        JOIN sam_admin.licence_snapshot_lines l ON l.snapshot_id = s.snapshot_id
        WHERE s.client_id = %s
          AND l.product_family IN ('oracle_database', 'oracle_weblogic')
        GROUP BY s.snapshot_month, l.product_family
        ORDER BY s.snapshot_month, l.product_family
    """, (client["client_id"],))

    # Latest snapshot detail lines (for the table at the bottom)
    latest_snap = query("""
        SELECT snapshot_id, snapshot_month, taken_at
        FROM sam_admin.licence_snapshots
        WHERE client_id = %s
        ORDER BY snapshot_month DESC LIMIT 1
    """, (client["client_id"],), fetchall=False)

    latest_lines = []
    if latest_snap:
        latest_lines = query("""
            SELECT hostname, environment, product_family, product_detail,
                   licences_required, licences_assigned, surplus_deficit, compliance_status,
                   csi_number
            FROM sam_admin.licence_snapshot_lines
            WHERE snapshot_id = %s
            ORDER BY hostname, product_family, product_detail
        """, (latest_snap["snapshot_id"],))

    def _row(r):
        d = dict(r)
        for k, v in d.items():
            if hasattr(v, 'isoformat'):
                d[k] = v.isoformat()
            elif hasattr(v, '__float__'):
                d[k] = float(v)
        return d

    chart_data = {
        "monthly":       [_row(r) for r in monthly],
        "family_monthly": [_row(r) for r in family_monthly],
        "latest_snap":   _row(latest_snap) if latest_snap else None,
    }

    return render_template("visibility_licence_history.html",
                           clients=clients_with_snaps,
                           selected=client,
                           chart_data=chart_data,
                           latest_lines=latest_lines)


# ---------------------------------------------------------------------------
# Visibility — WebLogic versions
# ---------------------------------------------------------------------------
@app.route("/visibility/wls-versions")
@login_required
def visibility_wls_versions():
    active_clients = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    clients_data = []
    for c in active_clients:
        s = c["schema_name"]
        try:
            rows = query(f"""
                SELECT
                    sv.hostname,
                    sv.environment::TEXT  AS environment,
                    sv.datacenter,
                    d.domain_name,
                    d.wls_version,
                    d.wls_edition,
                    d.last_seen::DATE     AS last_seen
                FROM {s}.oracle_servers sv
                JOIN {s}.wls_domains d ON d.server_id = sv.server_id AND d.is_active
                WHERE sv.is_active
                ORDER BY sv.hostname, d.domain_name
            """)
        except Exception as e:
            app.logger.error("visibility_wls_versions query failed for %s: %s", s, e)
            rows = []

        if not rows:
            continue

        # Group by version for the summary bar
        ver_counts = {}
        for r in rows:
            v = r["wls_version"] or "Unknown"
            ver_counts[v] = ver_counts.get(v, 0) + 1

        clients_data.append({
            "client_name":  c["client_name"],
            "client_code":  c["client_code"],
            "domains":      rows,
            "ver_counts":   sorted(ver_counts.items(), key=lambda x: x[0], reverse=True),
        })

    return render_template("visibility_wls_versions.html", clients=clients_data)


# ---------------------------------------------------------------------------
# Executive Summary
# ---------------------------------------------------------------------------
@app.route("/executive-summary")
@login_required
def executive_summary():
    today = date.today()

    active_clients = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )

    # ── Contract portfolio ───────────────────────────────────────────────────
    contract_stats = query("""
        SELECT
            cs.total_contracts,
            cs.ula_count,
            cs.standard_count,
            cs.active_count,
            cs.expired_count,
            cs.expiring_90d,
            cs.ula_expiring_90d,
            COALESCE(fin.total_licence_value, 0) AS total_licence_value,
            COALESCE(fin.total_support_cost, 0)  AS total_support_cost
        FROM (
            SELECT
                COUNT(*)                                          AS total_contracts,
                COUNT(*) FILTER (WHERE is_ula)                   AS ula_count,
                COUNT(*) FILTER (WHERE NOT is_ula)               AS standard_count,
                COUNT(*) FILTER (WHERE status = 'active')        AS active_count,
                COUNT(*) FILTER (WHERE status = 'expired')       AS expired_count,
                COUNT(*) FILTER (WHERE status = 'active'
                  AND NOT is_ula
                  AND support_expiry < CURRENT_DATE + INTERVAL '90 days') AS expiring_90d,
                COUNT(*) FILTER (WHERE status = 'active'
                  AND is_ula
                  AND ula_expiry < CURRENT_DATE + INTERVAL '90 days')     AS ula_expiring_90d
            FROM shared.csi_contracts
        ) cs
        CROSS JOIN (
            SELECT
                SUM(l.total_price)         AS total_licence_value,
                SUM(l.annual_support_cost) AS total_support_cost
            FROM shared.license_entitlement_lines l
            WHERE l.is_active
        ) fin
    """, fetchall=False)

    # ── Compliance posture across all clients ────────────────────────────────
    total_servers      = 0
    compliant_servers  = 0
    gap_servers        = 0
    unassigned_servers = 0
    total_licences_req = 0
    total_licences_lic = 0

    client_rows = []
    for cl in active_clients:
        s = cl["schema_name"]
        try:
            lp = query(f"""
                SELECT
                    COUNT(*)                                                    AS total,
                    COUNT(*) FILTER (WHERE compliance_status = 'compliant')     AS compliant,
                    COUNT(*) FILTER (WHERE compliance_status = 'under_licensed') AS gaps,
                    COALESCE(SUM(licences_required), 0)                         AS req,
                    COALESCE(SUM(COALESCE(total_licensed, 0)), 0)               AS lic
                FROM {s}.license_position
            """, fetchall=False)

            ua = query(f"""
                SELECT COUNT(*) AS cnt FROM {s}.oracle_servers sv
                WHERE sv.is_active
                  AND NOT EXISTS (
                    SELECT 1 FROM {s}.server_csi_map m WHERE m.server_id = sv.server_id
                  )
            """, fetchall=False)

            t  = int(lp["total"]    or 0)
            c  = int(lp["compliant"] or 0)
            g  = int(lp["gaps"]     or 0)
            ua_c = int(ua["cnt"]    or 0)
            req = float(lp["req"]   or 0)
            lic = float(lp["lic"]   or 0)

            total_servers      += t
            compliant_servers  += c
            gap_servers        += g
            unassigned_servers += ua_c
            total_licences_req += req
            total_licences_lic += lic

            score = round(100 * c / t) if t else 100
            rag = "green" if score == 100 and ua_c == 0 else ("amber" if score >= 80 else "red")
            client_rows.append({
                "client_name": cl["client_name"],
                "client_code": cl["client_code"],
                "total": t, "compliant": c, "gaps": g,
                "unassigned": ua_c, "score": score, "rag": rag,
                "req": req, "lic": lic,
            })
        except Exception:
            client_rows.append({
                "client_name": cl["client_name"],
                "client_code": cl["client_code"],
                "total": 0, "compliant": 0, "gaps": 0,
                "unassigned": 0, "score": 0, "rag": "amber",
                "req": 0, "lic": 0,
            })

    overall_score = round(100 * compliant_servers / total_servers) if total_servers else 100
    overall_rag   = "green" if overall_score == 100 and unassigned_servers == 0 \
                    else ("amber" if overall_score >= 80 else "red")

    # ── Top 5 products by licence count ─────────────────────────────────────
    top_products = query("""
        SELECT product_name,
               product_family::TEXT AS product_family,
               SUM(quantity) AS total_qty,
               SUM(annual_support_cost) AS total_support
        FROM shared.license_entitlement_lines
        WHERE is_active AND quantity > 0
        GROUP BY product_name, product_family
        ORDER BY total_qty DESC NULLS LAST
        LIMIT 5
    """)

    # ── Upcoming renewals (next 12 months) ───────────────────────────────────
    upcoming_renewals = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.is_ula, cs.currency,
               COALESCE(cs.ula_expiry, cs.support_expiry) AS expiry_date,
               COALESCE(SUM(l.annual_support_cost), 0)    AS renewal_cost,
               oc.client_name AS owning_client
        FROM shared.csi_contracts cs
        LEFT JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        WHERE cs.status = 'active'
          AND COALESCE(cs.ula_expiry, cs.support_expiry) BETWEEN CURRENT_DATE
                                                              AND CURRENT_DATE + INTERVAL '12 months'
        GROUP BY cs.csi_id, cs.csi_number, cs.contract_name, cs.is_ula,
                 cs.currency, cs.ula_expiry, cs.support_expiry, oc.client_name
        ORDER BY expiry_date
        LIMIT 10
    """)

    return render_template("executive_summary.html",
                           today=today,
                           contract_stats=contract_stats,
                           client_rows=client_rows,
                           overall_score=overall_score,
                           overall_rag=overall_rag,
                           total_servers=total_servers,
                           compliant_servers=compliant_servers,
                           gap_servers=gap_servers,
                           unassigned_servers=unassigned_servers,
                           total_licences_req=total_licences_req,
                           total_licences_lic=total_licences_lic,
                           top_products=top_products,
                           upcoming_renewals=upcoming_renewals)


# ---------------------------------------------------------------------------
# Audit Readiness Report
# ---------------------------------------------------------------------------
@app.route("/audit-readiness")
@login_required
def audit_readiness():
    today = date.today()
    active_clients = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )

    # ── 1. Licence gaps: servers where consumed > available ──────────────────
    licence_gaps = []
    for cl in active_clients:
        s = cl["schema_name"]
        try:
            rows = query(f"""
                SELECT %s AS client_name, hostname, environment::TEXT,
                       product_detail, product_family::TEXT AS product_family,
                       licences_required,
                       COALESCE(total_licensed, 0) AS total_licensed,
                       licences_required - COALESCE(total_licensed, 0) AS shortfall
                FROM {s}.license_position
                WHERE compliance_status = 'under_licensed'
                ORDER BY product_family, shortfall DESC
            """, (cl["client_name"],))
            licence_gaps.extend(rows)
        except Exception:
            pass

    # ── 2. Servers with no CSI assigned ─────────────────────────────────────
    unassigned_servers = []
    for cl in active_clients:
        s = cl["schema_name"]
        try:
            rows = query(f"""
                SELECT %s AS client_name, sv.hostname, sv.environment::TEXT,
                       sv.datacenter,
                       CASE WHEN EXISTS (
                           SELECT 1 FROM {s}.wls_domains d
                           WHERE d.server_id = sv.server_id AND d.is_active
                       ) THEN 'middleware' ELSE 'database' END AS server_type
                FROM {s}.oracle_servers sv
                WHERE sv.is_active
                  AND NOT EXISTS (
                      SELECT 1 FROM {s}.server_csi_map m WHERE m.server_id = sv.server_id
                  )
                ORDER BY sv.hostname
            """, (cl["client_name"],))
            unassigned_servers.extend(rows)
        except Exception:
            pass

    # ── 3. Expired or soon-expiring contracts (<90 days) ────────────────────
    contract_risks = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name, cs.status,
               cs.support_expiry, cs.ula_expiry, cs.is_ula,
               oc.client_name AS owning_client,
               CASE
                 WHEN cs.status != 'active' THEN 'expired'
                 WHEN cs.is_ula AND cs.ula_expiry < CURRENT_DATE THEN 'ula_expired'
                 WHEN cs.is_ula AND cs.ula_expiry < CURRENT_DATE + INTERVAL '90 days' THEN 'ula_expiring'
                 WHEN NOT cs.is_ula AND cs.support_expiry < CURRENT_DATE THEN 'expired'
                 WHEN NOT cs.is_ula AND cs.support_expiry < CURRENT_DATE + INTERVAL '90 days' THEN 'expiring'
                 ELSE 'ok'
               END AS risk_type
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        WHERE cs.status = 'active'
          AND (
            (cs.is_ula  AND cs.ula_expiry     < CURRENT_DATE + INTERVAL '90 days')
            OR
            (NOT cs.is_ula AND cs.support_expiry < CURRENT_DATE + INTERVAL '90 days')
          )
        ORDER BY COALESCE(cs.ula_expiry, cs.support_expiry)
    """)

    # ── 4. Contracts with no entitlement lines ───────────────────────────────
    empty_contracts = query("""
        SELECT cs.csi_id, cs.csi_number, cs.contract_name,
               oc.client_name AS owning_client
        FROM shared.csi_contracts cs
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        WHERE cs.is_ula = false
          AND NOT EXISTS (
              SELECT 1 FROM shared.license_entitlement_lines l
              WHERE l.csi_id = cs.csi_id AND l.is_active
          )
        ORDER BY cs.contract_name
    """)

    # ── 5. ULAs: servers assigned whose products exceed ULA scope ────────────
    ula_scope_violations = []
    ulas = query("""
        SELECT cs.csi_id, cs.contract_name, cs.owning_client_id,
               cl.schema_name, cl.client_name
        FROM shared.csi_contracts cs
        JOIN sam_admin.clients cl ON cl.client_id = cs.owning_client_id
        WHERE cs.is_ula AND cs.status = 'active'
    """)
    for ula in ulas:
        s = ula["schema_name"]
        covered = {r["product_name"] for r in query(
            "SELECT product_name FROM shared.ula_covered_products WHERE csi_id = %s",
            (ula["csi_id"],)
        )}
        if not covered:
            continue
        try:
            assigned = query(f"""
                SELECT DISTINCT sv.hostname, lp.product_detail
                FROM {s}.server_csi_map m
                JOIN {s}.oracle_servers sv ON sv.server_id = m.server_id
                JOIN {s}.license_position lp ON lp.server_id = m.server_id
                WHERE m.csi_id = %s
            """, (ula["csi_id"],))
            for row in assigned:
                if row["product_detail"] and row["product_detail"] not in covered:
                    ula_scope_violations.append({
                        "client_name": ula["client_name"],
                        "contract_name": ula["contract_name"],
                        "hostname": row["hostname"],
                        "product_detail": row["product_detail"],
                    })
        except Exception:
            pass

    # ── 6. Per-client summary scorecard ─────────────────────────────────────
    client_scores = []
    for cl in active_clients:
        s = cl["schema_name"]
        try:
            lp_rows = query(f"""
                SELECT COUNT(*) FILTER (WHERE compliance_status = 'under_licensed') AS gaps,
                       COUNT(*) AS total
                FROM {s}.license_position
            """, fetchall=False)
            ua_count = sum(1 for r in unassigned_servers if r["client_name"] == cl["client_name"])
            gaps = int(lp_rows["gaps"] or 0) if lp_rows else 0
            total = int(lp_rows["total"] or 0) if lp_rows else 0
            issues = gaps + ua_count
            if issues == 0:
                rag = "green"
            elif issues <= 3:
                rag = "amber"
            else:
                rag = "red"
            client_scores.append({
                "client_name": cl["client_name"],
                "client_code": cl["client_code"],
                "total_positions": total,
                "gaps": gaps,
                "unassigned": ua_count,
                "rag": rag,
            })
        except Exception:
            client_scores.append({
                "client_name": cl["client_name"],
                "client_code": cl["client_code"],
                "total_positions": 0,
                "gaps": 0,
                "unassigned": 0,
                "rag": "amber",
            })

    total_issues = (len(licence_gaps) + len(unassigned_servers)
                    + len(contract_risks) + len(empty_contracts)
                    + len(ula_scope_violations))

    return render_template("audit_readiness.html",
                           today=today,
                           client_scores=client_scores,
                           licence_gaps=licence_gaps,
                           unassigned_servers=unassigned_servers,
                           contract_risks=contract_risks,
                           empty_contracts=empty_contracts,
                           ula_scope_violations=ula_scope_violations,
                           total_issues=total_issues)


# ---------------------------------------------------------------------------
# Compliance dashboard
# ---------------------------------------------------------------------------
@app.route("/compliance")
@login_required
def compliance():
    active_clients = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    summary = []
    for c in active_clients:
        s = c["schema_name"]
        try:
            row = query(f"""
                WITH per_server AS (
                    SELECT
                        srv.server_id,
                        COALESCE(BOOL_OR(lp.compliance_status = 'under_licensed'), FALSE) AS is_non_compliant,
                        COALESCE(SUM(lp.licences_required), 0) AS licences_required,
                        COALESCE(SUM(lp.total_licensed), 0)    AS licences_assigned,
                        COALESCE(SUM(GREATEST(lp.licences_required - lp.total_licensed, 0)), 0) AS licences_short
                    FROM {s}.oracle_servers srv
                    LEFT JOIN {s}.license_position lp ON lp.server_id = srv.server_id
                    GROUP BY srv.server_id
                )
                SELECT
                    COUNT(*)                             AS total_servers,
                    COUNT(*) FILTER (WHERE is_non_compliant)  AS non_compliant_servers,
                    COALESCE(SUM(licences_required), 0) AS total_required,
                    COALESCE(SUM(licences_assigned), 0) AS total_assigned,
                    COALESCE(SUM(licences_short), 0)    AS total_short
                FROM per_server
            """, fetchall=False)
        except Exception as e:
            app.logger.error("compliance summary query failed for %s: %s", s, e)
            row = None
        if not row:
            continue
        total    = int(row["total_servers"] or 0)
        non_comp = int(row["non_compliant_servers"] or 0)
        comp     = total - non_comp
        score    = round(comp / total * 100, 1) if total > 0 else 100.0
        summary.append({
            "client_code":           c["client_code"],
            "client_name":           c["client_name"],
            "total_servers":         total,
            "compliant_servers":     comp,
            "non_compliant_servers": non_comp,
            "total_required":        int(row["total_required"] or 0),
            "total_assigned":        int(row["total_assigned"] or 0),
            "total_short":           int(row["total_short"] or 0),
            "score":                 score,
        })
    return render_template("compliance.html", summary=summary)


@app.route("/compliance/<client_code>")
@login_required
def compliance_client(client_code):
    client = query(
        "SELECT client_id, client_code, client_name, schema_name "
        "FROM sam_admin.clients WHERE client_code = %s",
        (client_code,), fetchall=False
    )
    if not client:
        flash("Client not found.", "danger")
        return redirect(url_for("compliance"))
    s = client["schema_name"]
    # Detect whether the license_position view has the NUP columns added by
    # migration 05_nup_license_position.sql so the page works before and after.
    _nup_cols_exist = query(
        """
        SELECT COUNT(*) AS n
        FROM   information_schema.columns
        WHERE  table_schema = %s
          AND  table_name   = 'license_position'
          AND  column_name  IN ('licence_metric','nup_minimum','nup_active_users')
        """,
        (s,), fetchall=False
    )
    _has_nup_cols = (_nup_cols_exist or {}).get("n", 0) == 3

    _nup_issue_fields = (
        """
                    'licence_metric', lp.licence_metric,
                    'nup_minimum',    lp.nup_minimum,
                    'nup_users',      lp.nup_active_users
        """
        if _has_nup_cols else
        """
                    'licence_metric', 'processor_perpetual'::TEXT,
                    'nup_minimum',    NULL::NUMERIC,
                    'nup_users',      NULL::NUMERIC
        """
    )

    try:
        servers = query(f"""
            SELECT
                srv.server_id, srv.hostname, srv.environment,
                COALESCE(BOOL_OR(lp.compliance_status = 'under_licensed'), FALSE) AS is_non_compliant,
                COALESCE(SUM(lp.licences_required), 0) AS total_required,
                COALESCE(SUM(lp.total_licensed), 0)    AS total_assigned,
                COALESCE(SUM(GREATEST(lp.licences_required - lp.total_licensed, 0)), 0) AS total_short,
                COALESCE(BOOL_OR(lp.product_family = 'oracle_database'), FALSE) AS has_db,
                COALESCE(BOOL_OR(lp.product_family != 'oracle_database'), FALSE) AS has_mw,
                COALESCE(BOOL_OR(lp.product_family = 'oracle_database'
                    AND lp.compliance_status = 'under_licensed'), FALSE) AS db_non_compliant,
                COALESCE(BOOL_OR(lp.product_family != 'oracle_database'
                    AND lp.compliance_status = 'under_licensed'), FALSE) AS mw_non_compliant,
                JSON_AGG(JSON_BUILD_OBJECT(
                    'product',        COALESCE(lp.product_detail, lp.product_family),
                    'product_family', lp.product_family,
                    'required',       lp.licences_required,
                    'assigned',       lp.total_licensed,
                    'short',          GREATEST(lp.licences_required - lp.total_licensed, 0),
                    {_nup_issue_fields}
                ) ORDER BY lp.product_family)
                    FILTER (WHERE lp.compliance_status = 'under_licensed') AS issues
            FROM {s}.oracle_servers srv
            LEFT JOIN {s}.license_position lp ON lp.server_id = srv.server_id
            GROUP BY srv.server_id, srv.hostname, srv.environment
            ORDER BY is_non_compliant DESC NULLS LAST, srv.hostname
        """)
    except Exception as e:
        app.logger.error("compliance_client query failed for %s: %s", s, e)
        servers = []

    servers = [dict(sv) for sv in servers]
    for sv in servers:
        issues = sv.get("issues")
        if issues and isinstance(issues, str):
            sv["issues"] = json.loads(issues)
        elif not issues:
            sv["issues"] = []

    total    = len(servers)
    non_comp = sum(1 for sv in servers if sv["is_non_compliant"])
    comp     = total - non_comp
    score    = round(comp / total * 100, 1) if total > 0 else 100.0

    db_servers = [sv for sv in servers if sv["has_db"]]
    mw_servers = [sv for sv in servers if sv["has_mw"]]
    db_total   = len(db_servers)
    mw_total   = len(mw_servers)
    db_comp    = sum(1 for sv in db_servers if not sv["db_non_compliant"])
    mw_comp    = sum(1 for sv in mw_servers if not sv["mw_non_compliant"])
    db_score   = round(db_comp / db_total * 100, 1) if db_total > 0 else 100.0
    mw_score   = round(mw_comp / mw_total * 100, 1) if mw_total > 0 else 100.0

    return render_template("compliance_client.html",
                           client=client,
                           servers=servers,
                           total_servers=total,
                           compliant_servers=comp,
                           non_compliant_servers=non_comp,
                           score=score,
                           db_servers=db_servers,
                           mw_servers=mw_servers,
                           db_total=db_total, db_comp=db_comp, db_score=db_score,
                           mw_total=mw_total, mw_comp=mw_comp, mw_score=mw_score)


# ---------------------------------------------------------------------------
# Compliance alerts
# ---------------------------------------------------------------------------
@app.route("/alerts")
@login_required
def alerts():
    role = current_role()
    u    = current_user()

    try:
        rows = query(
            "SELECT * FROM shared.compliance_alerts ORDER BY "
            "CASE severity WHEN 'HIGH' THEN 1 WHEN 'MEDIUM' THEN 2 ELSE 3 END, "
            "days_until NULLS LAST, client_name, object_name"
        )
    except Exception as e:
        flash(f"Could not load alerts: {e}", "danger")
        rows = []

    # Client users are hard-scoped to their own client — enforce server-side
    if role == "client":
        user_client_code = (query(
            "SELECT client_code FROM sam_admin.clients WHERE client_id = %s",
            (u.get("client_id"),), fetchall=False
        ) or {}).get("client_code")
        rows = [r for r in (rows or []) if r.get("client_code") == user_client_code]
        all_clients = []
        client_filter = user_client_code
    else:
        all_clients = query(
            "SELECT client_code, client_name FROM sam_admin.clients "
            "WHERE is_active ORDER BY client_name"
        ) or []
        client_filter = request.args.get("client", "ALL")
        if client_filter != "ALL":
            rows = [r for r in (rows or []) if r.get("client_code") == client_filter]

    severity_filter = request.args.get("severity", "ALL")
    type_filter     = request.args.get("type", "ALL")

    alert_types = sorted({r["alert_type"] for r in (rows or [])})

    if severity_filter != "ALL":
        rows = [r for r in rows if r["severity"] == severity_filter]
    if type_filter != "ALL":
        rows = [r for r in rows if r["alert_type"] == type_filter]

    high_count   = sum(1 for r in (rows or []) if r["severity"] == "HIGH")
    medium_count = sum(1 for r in (rows or []) if r["severity"] == "MEDIUM")

    return render_template("alerts.html", alerts=rows,
                           alert_types=alert_types,
                           severity_filter=severity_filter,
                           type_filter=type_filter,
                           client_filter=client_filter,
                           all_clients=all_clients,
                           high_count=high_count,
                           medium_count=medium_count)


# ---------------------------------------------------------------------------
# VMware cluster view
# ---------------------------------------------------------------------------
@app.route("/vmware")
@login_required
def vmware():
    clusters = query("""
        SELECT vle.*, c.client_name
        FROM   sam_admin.vmware_licence_exposure vle
        LEFT   JOIN sam_admin.clients c ON c.client_code = vle.client_code
        ORDER  BY vle.has_oracle_workloads DESC, vle.total_physical_cores DESC
    """)
    return render_template("vmware.html", clusters=clusters)


# ---------------------------------------------------------------------------
# LMS Audit Export (Excel workbook)
# ---------------------------------------------------------------------------
@app.route("/export/lms")
@login_required
def export_lms():
    schema = get_schema()
    wb = Workbook()

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="203864")
    warn_fill    = PatternFill("solid", fgColor="FFE599")
    danger_fill  = PatternFill("solid", fgColor="F4CCCC")

    def add_sheet(title, columns, rows):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for cell in ws[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(c.lower().replace(" ", "_").replace("/", "_")) for c in columns])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(col[0].value or "")),
                max((len(str(c.value or "")) for c in col[1:]), default=0)
            ) + 4
        return ws

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "Helios — LMS Audit Export"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum["A3"] = f"Client schema: {schema}"
    ws_sum["A5"] = "Sheet"
    ws_sum["B5"] = "Contents"
    ws_sum["A5"].font = Font(bold=True)
    ws_sum["B5"].font = Font(bold=True)
    for i, (s, d) in enumerate([
        ("Server Inventory", "All active Oracle servers with hardware details"),
        ("Processor Details", "CPU socket / core / factor data per server"),
        ("Oracle Instances", "Oracle DB instances with edition and version"),
        ("Options", "v$option flags per instance"),
        ("Licence Position", "Calculated licence requirements per product per server"),
        ("CSI Coverage", "Per-server/product breakdown: which CSI covers each licence requirement"),
        ("CSI Contracts", "All CSI contract headers and entitlement totals"),
        ("SE2 Violations", "Servers violating SE2 socket or RAC node limits"),
        ("CPU Validation", "Servers with unrecognised CPU models"),
        ("VMware Exposure", "vSphere clusters with Oracle VM workloads"),
    ], start=6):
        ws_sum[f"A{i}"] = s
        ws_sum[f"B{i}"] = d

    # Sheet 2: Server Inventory
    servers_data = query(f"""
        SELECT s.hostname, s.environment::TEXT AS environment, s.datacenter,
               s.ip_address::TEXT AS ip_address, s.os_family, s.os_distribution,
               s.os_version, s.total_ram_mb, s.last_seen::DATE AS last_seen,
               COALESCE(s.licence_metric_override, 'processor_perpetual') AS licence_metric
        FROM {schema}.oracle_servers s WHERE s.is_active ORDER BY s.hostname
    """)
    ws2 = wb.create_sheet("Server Inventory")
    cols2 = ["Hostname", "Environment", "Datacenter", "IP Address", "OS Family",
             "OS Distribution", "OS Version", "RAM (MB)", "Last Seen", "Licence Metric"]
    ws2.append(cols2)
    for cell in ws2[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in servers_data:
        ws2.append([r.get("hostname"), r.get("environment"), r.get("datacenter"),
                    r.get("ip_address"), r.get("os_family"), r.get("os_distribution"),
                    r.get("os_version"), r.get("total_ram_mb"),
                    str(r.get("last_seen", "")), r.get("licence_metric")])

    # Sheet 3: Processor Details
    proc_data = query(f"""
        SELECT s.hostname, p.cpu_model, p.cpu_architecture,
               p.cpu_sockets, p.cores_per_socket,
               p.cpu_sockets * p.cores_per_socket AS total_physical_cores,
               p.threads_per_core, p.virt_type::TEXT AS virt_type,
               p.is_vmware, p.is_exadata, p.vcpu_count,
               shared.cpu_core_factor_lookup(p.cpu_model) AS oracle_core_factor,
               p.recorded_at::DATE AS snapshot_date
        FROM {schema}.oracle_servers s
        JOIN LATERAL (
          SELECT * FROM {schema}.oracle_processors
          WHERE server_id = s.server_id ORDER BY recorded_at DESC LIMIT 1
        ) p ON TRUE
        WHERE s.is_active ORDER BY s.hostname
    """)
    ws3 = wb.create_sheet("Processor Details")
    cols3 = ["Hostname", "CPU Model", "Architecture", "Sockets", "Cores/Socket",
             "Total Physical Cores", "Threads/Core", "Virt Type", "Is VMware", "Is Exadata",
             "vCPU Count", "Oracle Core Factor", "Snapshot Date"]
    ws3.append(cols3)
    for cell in ws3[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in proc_data:
        row = [r.get("hostname"), r.get("cpu_model"), r.get("cpu_architecture"),
               r.get("cpu_sockets"), r.get("cores_per_socket"),
               r.get("total_physical_cores"), r.get("threads_per_core"),
               r.get("virt_type"), r.get("is_vmware"), r.get("is_exadata"), r.get("vcpu_count"),
               r.get("oracle_core_factor"), str(r.get("snapshot_date", ""))]
        ws3.append(row)
        if r.get("oracle_core_factor") is None:
            for cell in ws3[ws3.max_row]:
                cell.fill = warn_fill

    # Sheet 4: Oracle Instances
    inst_data = query(f"""
        SELECT s.hostname, i.oracle_sid, i.db_name, i.edition,
               i.db_version, i.platform_name, i.last_seen::DATE AS last_seen
        FROM {schema}.oracle_instances i
        JOIN {schema}.oracle_servers s ON s.server_id = i.server_id
        WHERE i.is_active ORDER BY s.hostname, i.oracle_sid
    """)
    ws4 = wb.create_sheet("Oracle Instances")
    cols4 = ["Hostname", "Oracle SID", "DB Name", "Edition", "Version",
             "Platform", "Last Seen"]
    ws4.append(cols4)
    for cell in ws4[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in inst_data:
        ws4.append([r.get("hostname"), r.get("oracle_sid"), r.get("db_name"),
                    r.get("edition"), r.get("db_version"), r.get("platform_name"),
                    str(r.get("last_seen", ""))])

    # Sheet 5: Options
    try:
        opt_data = query(f"""
            SELECT s.hostname, i.oracle_sid, o.option_name, o.is_active::TEXT AS installed
            FROM {schema}.oracle_options o
            JOIN {schema}.oracle_instances i ON i.instance_id = o.instance_id
            JOIN {schema}.oracle_servers   s ON s.server_id   = i.server_id
            WHERE o.is_active = TRUE
            ORDER BY s.hostname, i.oracle_sid, o.option_name
        """)
    except Exception:
        opt_data = []
    ws5 = wb.create_sheet("Options")
    cols5 = ["Hostname", "Oracle SID", "Option Name", "Installed"]
    ws5.append(cols5)
    for cell in ws5[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in opt_data:
        ws5.append([r.get("hostname"), r.get("oracle_sid"),
                    r.get("option_name"), r.get("installed")])

    # Sheet 6: Licence Position
    lp_data = query(f"""
        SELECT s.hostname, lp.product_family, lp.product_detail,
               lp.licences_required, lp.total_licensed,
               lp.licence_surplus_deficit, lp.compliance_status
        FROM {schema}.license_position lp
        JOIN {schema}.oracle_servers s ON s.server_id = lp.server_id
        ORDER BY s.hostname, lp.product_family
    """)
    ws6 = wb.create_sheet("Licence Position")
    cols6 = ["Hostname", "Product Family", "Product Detail", "Licences Required",
             "Total Licensed", "Surplus/Deficit", "Compliance Status"]
    ws6.append(cols6)
    for cell in ws6[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in lp_data:
        row = [r.get("hostname"), r.get("product_family"), r.get("product_detail"),
               r.get("licences_required"), r.get("total_licensed"),
               r.get("licence_surplus_deficit"), r.get("compliance_status")]
        ws6.append(row)
        if r.get("compliance_status") == "under_licensed":
            for cell in ws6[ws6.max_row]:
                cell.fill = danger_fill

    # Sheet 7: CSI Coverage — per-server/product/CSI breakdown across all clients
    active_clients = query(
        "SELECT client_name, schema_name FROM sam_admin.clients WHERE is_active ORDER BY client_name"
    )
    coverage_rows = []
    for cl in active_clients:
        s = cl["schema_name"]
        try:
            rows = query(f"""
                SELECT
                    sv.hostname,
                    COALESCE(m.product_detail, m.product_family::TEXT) AS product,
                    lp.licences_required,
                    cs.csi_number,
                    cs.contract_name,
                    COALESCE(m.licences_consumed, 0) AS licences_from_csi
                FROM {s}.server_csi_map m
                JOIN {s}.oracle_servers sv  ON sv.server_id  = m.server_id  AND sv.is_active
                JOIN shared.csi_contracts cs ON cs.csi_id    = m.csi_id
                LEFT JOIN {s}.license_position lp
                       ON lp.server_id      = m.server_id
                      AND lp.product_family = m.product_family
                      AND COALESCE(lp.product_detail,'') = COALESCE(m.product_detail,'')
                ORDER BY sv.hostname, product, cs.csi_number
            """)
            for r in rows:
                coverage_rows.append({
                    "client_name":       cl["client_name"],
                    "hostname":          r["hostname"],
                    "product":           r["product"],
                    "licences_required": r["licences_required"],
                    "csi_number":        r["csi_number"],
                    "contract_name":     r["contract_name"],
                    "licences_from_csi": r["licences_from_csi"],
                })
        except Exception as e:
            app.logger.warning("CSI coverage query failed for %s: %s", s, e)

    ws_cov = wb.create_sheet("CSI Coverage")
    cov_cols = ["Client", "Hostname", "Product", "Licences Required",
                "CSI Number", "Contract Name", "Licences from CSI"]
    ws_cov.append(cov_cols)
    for cell in ws_cov[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r in coverage_rows:
        ws_cov.append([
            r["client_name"], r["hostname"], r["product"],
            r["licences_required"], r["csi_number"],
            r["contract_name"], r["licences_from_csi"],
        ])
    for col in ws_cov.columns:
        ws_cov.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")),
            max((len(str(c.value or "")) for c in col[1:]), default=0)
        ) + 4

    # Sheet 8: CSI Contracts
    csi_data = query("""
        SELECT cs.csi_number, cs.contract_name, cs.purchase_date,
               cs.support_start, cs.support_expiry, cs.is_ula, cs.ula_expiry,
               cs.status, cs.sharing_policy,
               oc.client_code AS owning_client,
               COUNT(DISTINCT l.line_id) AS line_count,
               SUM(l.quantity) AS total_licences,
               SUM(l.total_price) AS total_licence_cost,
               STRING_AGG(DISTINCT l.product_name, ' | ' ORDER BY l.product_name) AS products
        FROM shared.csi_contracts cs
        LEFT JOIN shared.license_entitlement_lines l ON l.csi_id = cs.csi_id AND l.is_active
        LEFT JOIN sam_admin.clients oc ON oc.client_id = cs.owning_client_id
        GROUP BY cs.csi_id, cs.csi_number, cs.contract_name, cs.purchase_date,
                 cs.support_start, cs.support_expiry, cs.is_ula, cs.ula_expiry,
                 cs.status, cs.sharing_policy, oc.client_code
        ORDER BY cs.csi_number
    """)
    ws7 = wb.create_sheet("CSI Contracts")
    cols7 = ["CSI Number", "Contract Name", "Purchase Date", "Support Start",
             "Support Expiry", "Is ULA", "ULA Expiry", "Status", "Sharing Policy",
             "Owning Client", "Line Count", "Total Licences", "Total Cost", "Products"]
    ws7.append(cols7)
    for cell in ws7[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in csi_data:
        ws7.append([r.get("csi_number"), r.get("contract_name"),
                    str(r.get("purchase_date", "")), str(r.get("support_start", "")),
                    str(r.get("support_expiry", "")), r.get("is_ula"),
                    str(r.get("ula_expiry", "")), r.get("status"), r.get("sharing_policy"),
                    r.get("owning_client"), r.get("line_count"),
                    r.get("total_licences"), r.get("total_licence_cost"), r.get("products")])

    # Sheet 8: SE2 Violations
    try:
        se2_data = query(f"SELECT * FROM {schema}.se2_violations ORDER BY hostname")
    except Exception:
        se2_data = []
    ws8 = wb.create_sheet("SE2 Violations")
    cols8 = ["Hostname", "Oracle SID", "Edition", "CPU Sockets",
             "RAC Node Count", "Socket Violation", "RAC Violation", "Summary"]
    ws8.append(cols8)
    for cell in ws8[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in se2_data:
        row = [r.get("hostname"), r.get("oracle_sid"), r.get("edition"),
               r.get("cpu_sockets"), r.get("rac_node_count"),
               r.get("socket_violation"), r.get("rac_violation"),
               r.get("violation_summary")]
        ws8.append(row)
        for cell in ws8[ws8.max_row]:
            cell.fill = danger_fill

    # Sheet 9: CPU Validation
    try:
        cpu_data = query(f"SELECT * FROM {schema}.cpu_validation_report ORDER BY factor_unknown DESC, hostname")
    except Exception:
        cpu_data = []
    ws9 = wb.create_sheet("CPU Validation")
    cols9 = ["Hostname", "CPU Model", "Sockets", "Cores/Socket",
             "Total Cores", "Validated Factor", "Factor Unknown", "Status", "Last Snapshot"]
    ws9.append(cols9)
    for cell in ws9[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in cpu_data:
        row = [r.get("hostname"), r.get("cpu_model"), r.get("cpu_sockets"),
               r.get("cores_per_socket"), r.get("total_physical_cores"),
               r.get("validated_factor"), r.get("factor_unknown"),
               r.get("validation_status"), str(r.get("last_processor_snapshot", ""))]
        ws9.append(row)
        if r.get("factor_unknown"):
            for cell in ws9[ws9.max_row]:
                cell.fill = warn_fill

    # Sheet 10: VMware Exposure
    vmw_data = query("""
        SELECT cluster_name, vcenter_host, datacenter, client_code,
               host_count, total_sockets, total_physical_cores,
               oracle_db_vm_count, oracle_wls_vm_count, oracle_java_vm_count,
               has_oracle_workloads, last_seen
        FROM sam_admin.vmware_licence_exposure
        ORDER BY has_oracle_workloads DESC, total_physical_cores DESC
    """)
    ws10 = wb.create_sheet("VMware Exposure")
    cols10 = ["Cluster Name", "vCenter Host", "Datacenter", "Client",
              "Host Count", "Total Sockets", "Total Physical Cores",
              "Oracle DB VMs", "WLS VMs", "Java VMs", "Has Oracle Workloads", "Last Seen"]
    ws10.append(cols10)
    for cell in ws10[1]:
        cell.font = header_font; cell.fill = header_fill
    for r in vmw_data:
        row = [r.get("cluster_name"), r.get("vcenter_host"), r.get("datacenter"),
               r.get("client_code"), r.get("host_count"), r.get("total_sockets"),
               r.get("total_physical_cores"), r.get("oracle_db_vm_count"),
               r.get("oracle_wls_vm_count"), r.get("oracle_java_vm_count"),
               r.get("has_oracle_workloads"), str(r.get("last_seen", ""))]
        ws10.append(row)
        if r.get("has_oracle_workloads"):
            for cell in ws10[ws10.max_row]:
                cell.fill = warn_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"oracle_lms_export_{schema}_{date.today().isoformat()}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# Settings — alert channels
# ---------------------------------------------------------------------------
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "add_channel":
            ctype = request.form.get("channel_type")
            cname = request.form.get("channel_name", "").strip()
            min_sev = request.form.get("min_severity", "MEDIUM")
            if ctype == "email":
                cfg = {
                    "smtp_host":     request.form.get("smtp_host", ""),
                    "smtp_port":     int(request.form.get("smtp_port", 587)),
                    "smtp_user":     request.form.get("smtp_user", ""),
                    "smtp_password": request.form.get("smtp_password", ""),
                    "from_addr":     request.form.get("from_addr", ""),
                    "to_addrs":      [a.strip() for a in
                                      request.form.get("to_addrs", "").split(",") if a.strip()],
                }
            else:
                cfg = {"webhook_url": request.form.get("webhook_url", "")}
            execute(
                "INSERT INTO sam_admin.alert_channels "
                "(channel_type, channel_name, config, min_severity) VALUES (%s, %s, %s, %s)",
                (ctype, cname, json.dumps(cfg), min_sev)
            )
            flash(f"Alert channel '{cname}' added.", "success")

        elif action == "toggle":
            cid = request.form.get("channel_id")
            execute(
                "UPDATE sam_admin.alert_channels SET enabled = NOT enabled WHERE channel_id = %s",
                (cid,)
            )
            flash("Channel updated.", "success")

        elif action == "delete":
            cid = request.form.get("channel_id")
            execute("DELETE FROM sam_admin.alert_channels WHERE channel_id = %s", (cid,))
            flash("Channel deleted.", "success")

        return redirect(url_for("settings"))

    channels = query(
        "SELECT * FROM sam_admin.alert_channels ORDER BY channel_type, channel_name"
    )
    return render_template("settings.html", channels=channels)


@app.route("/settings/test/<int:channel_id>", methods=["POST"])
@login_required
def test_channel(channel_id):
    row = query(
        "SELECT * FROM sam_admin.alert_channels WHERE channel_id = %s",
        (channel_id,), fetchall=False
    )
    if not row:
        flash("Channel not found.", "danger")
        return redirect(url_for("settings"))

    test_alert = {
        "alert_type": "TEST",
        "severity": "LOW",
        "object_name": "SAM Alert System",
        "description": "This is a test alert from Helios.",
        "action_needed": "No action required — this is a connectivity test.",
    }
    ok, err = _send_to_channel(row, [test_alert])
    if ok:
        flash(f"Test alert sent successfully to '{row['channel_name']}'.", "success")
    else:
        flash(f"Test failed: {err}", "danger")
    return redirect(url_for("settings"))


# ---------------------------------------------------------------------------
# Alert dispatch (call from cron)
# ---------------------------------------------------------------------------
@app.route("/api/dispatch-alerts")
def dispatch_alerts():
    """Send pending compliance alerts to all enabled channels.
    Protect with DISPATCH_KEY env var:
      curl 'http://localhost:5000/api/dispatch-alerts?key=YOUR_KEY'
    """
    if DISPATCH_KEY and request.args.get("key", "") != DISPATCH_KEY:
        return jsonify({"error": "unauthorized"}), 401

    # Check all client schemas for Oracle features inactive > 1 year and
    # take automatic licence-position snapshots where found.
    dormancy_results = []
    try:
        rows = query("SELECT client_id, feature, last_active, snapshot_id"
                     "  FROM sam_admin.check_feature_dormancy()")
        for row in (rows or []):
            if row.get("snapshot_id"):
                dormancy_results.append({
                    "client_id": row["client_id"],
                    "feature":   row["feature"],
                    "snapshot_id": row["snapshot_id"],
                })
    except Exception as exc:
        dormancy_results = [{"error": str(exc)}]

    alerts_data = query(
        "SELECT * FROM shared.compliance_alerts ORDER BY severity, days_until NULLS LAST"
    )
    if not alerts_data:
        return jsonify({"status": "ok", "sent": 0, "message": "No alerts to dispatch"})

    channels = query(
        "SELECT * FROM sam_admin.alert_channels WHERE enabled ORDER BY channel_id"
    )
    results = []
    for ch in channels:
        # Filter by min_severity
        sev_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
        min_sev = sev_order.get(ch.get("min_severity", "MEDIUM"), 2)
        filtered = [a for a in alerts_data
                    if sev_order.get(a.get("severity", "LOW"), 1) >= min_sev]
        if not filtered:
            continue
        ok, err = _send_to_channel(ch, filtered)
        execute(
            "UPDATE sam_admin.alert_channels SET last_sent_at = NOW() WHERE channel_id = %s",
            (ch["channel_id"],)
        )
        results.append({"channel": ch["channel_name"], "ok": ok, "error": err,
                        "alert_count": len(filtered)})

    return jsonify({"status": "ok", "channels": results,
                    "dormancy_snapshots": dormancy_results})


def _send_to_channel(channel, alerts):
    """Send a list of alerts to a channel. Returns (success, error_message)."""
    ctype = channel["channel_type"]
    cfg   = channel["config"] if isinstance(channel["config"], dict) else json.loads(channel["config"])

    try:
        if ctype == "slack":
            _send_slack(cfg["webhook_url"], channel["channel_name"], alerts)
        elif ctype == "teams":
            _send_teams(cfg["webhook_url"], channel["channel_name"], alerts)
        elif ctype == "email":
            _send_email(cfg, channel["channel_name"], alerts)
        return True, None
    except Exception as exc:
        return False, str(exc)


def _send_slack(webhook_url, channel_name, alerts):
    high   = [a for a in alerts if a.get("severity") == "HIGH"]
    medium = [a for a in alerts if a.get("severity") == "MEDIUM"]
    blocks = [{"type": "header", "text": {"type": "plain_text",
               "text": f"Helios Compliance Alerts ({len(alerts)} total)"}}]
    for a in alerts[:20]:
        emoji = ":red_circle:" if a.get("severity") == "HIGH" else ":large_yellow_circle:"
        blocks.append({"type": "section", "text": {"type": "mrkdwn",
            "text": f"{emoji} *{a.get('alert_type')}* — {a.get('object_name')}\n"
                    f"{a.get('description')}\n_Action: {a.get('action_needed')}_"}})
    if len(alerts) > 20:
        blocks.append({"type": "section", "text": {"type": "mrkdwn",
            "text": f"… and {len(alerts) - 20} more alerts. Log in to Helios to review all."}})
    resp = requests.post(webhook_url, json={"blocks": blocks}, timeout=10)
    resp.raise_for_status()


def _send_teams(webhook_url, channel_name, alerts):
    facts = [{"name": f"{a.get('severity')} — {a.get('alert_type')}",
              "value": f"{a.get('object_name')}: {a.get('description')}"}
             for a in alerts[:25]]
    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": "FF0000" if any(a.get("severity") == "HIGH" for a in alerts) else "FFA500",
        "summary": f"Helios: {len(alerts)} compliance alert(s)",
        "sections": [{"activityTitle": f"Helios — {len(alerts)} Compliance Alert(s)",
                      "facts": facts}]
    }
    resp = requests.post(webhook_url, json=payload, timeout=10)
    resp.raise_for_status()


def _send_email(cfg, channel_name, alerts):
    body_lines = [f"Helios Compliance Alerts — {len(alerts)} item(s)\n",
                  "=" * 60]
    for a in alerts:
        body_lines += [
            f"\n[{a.get('severity')}] {a.get('alert_type')}",
            f"  Object  : {a.get('object_name')}",
            f"  Client  : {a.get('client_code', 'N/A')}",
            f"  Detail  : {a.get('description')}",
            f"  Action  : {a.get('action_needed')}",
        ]
    body = "\n".join(body_lines)

    msg = MIMEMultipart()
    msg["From"]    = cfg.get("from_addr", "sam@example.com")
    msg["To"]      = ", ".join(cfg.get("to_addrs", []))
    msg["Subject"] = f"Helios: {len(alerts)} compliance alert(s)"
    msg.attach(MIMEText(body, "plain"))

    ctx = ssl.create_default_context()
    with smtplib.SMTP(cfg["smtp_host"], int(cfg.get("smtp_port", 587))) as smtp:
        smtp.starttls(context=ctx)
        if cfg.get("smtp_user"):
            smtp.login(cfg["smtp_user"], cfg.get("smtp_password", ""))
        smtp.sendmail(cfg["from_addr"], cfg["to_addrs"], msg.as_string())


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------
@app.route("/healthz")
def healthz():
    return jsonify({"status": "ok"})


# ---------------------------------------------------------------------------
# Context processor — inject shared variables into every template
# ---------------------------------------------------------------------------
@app.context_processor
def inject_globals():
    today = date.today()
    schema = get_schema() if session.get("logged_in") else DEFAULT_CLIENT_SCHEMA

    stale_count = 0
    if session.get("logged_in"):
        try:
            active_clients = query(
                "SELECT schema_name FROM sam_admin.clients WHERE is_active"
            )
            for c in active_clients:
                s = c["schema_name"]
                try:
                    result = query(f"""
                        SELECT COUNT(*) AS n FROM {s}.oracle_servers
                        WHERE is_active AND last_seen < NOW() - INTERVAL '{STALE_THRESHOLD_DAYS} days'
                    """, fetchall=False)
                    stale_count += result["n"] if result else 0
                except Exception:
                    pass
        except Exception:
            pass

    return {
        "today":               today.isoformat(),
        "today_date":          today,
        "active_schema":       schema,
        "all_clients":         get_clients() if session.get("logged_in") else [],
        "has_ulas":            _client_has_ulas(schema) if session.get("logged_in") else False,
        "stale_servers_count": stale_count,
    }


@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403




if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
